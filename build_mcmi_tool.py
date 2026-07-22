#!/usr/bin/env python3
"""
MCMI-III Complete Scoring Tool Builder
Creates a comprehensive Excel workbook for automated MCMI-III scoring.
Based on the MCMI-III Manual (Millon, 1994/2009) scoring procedures.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from copy import copy

wb = Workbook()

# ============================================================
# COLOR DEFINITIONS
# ============================================================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")

SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================
# MCMI-III ITEM SCORING KEYS (from Appendix B / Images 1-6)
# Each scale lists the item numbers scored TRUE and items scored FALSE
# Format: { 'scale_code': {'true': [...], 'false': [...]} }
# Based on the 175-item MCMI-III answer sheet
# ============================================================


SCORING_KEYS = {
    # Modifying Indices
    'X': {  # Disclosure
        'true': [1, 4, 5, 6, 7, 10, 14, 15, 17, 21, 23, 24, 25, 27, 29, 30, 31, 32, 34, 35, 37, 39, 40, 42, 43, 44, 46, 48, 50, 51, 52, 53, 54, 55, 56, 58, 60, 61, 62, 63, 64, 66, 67, 69, 70, 71, 73, 74, 75, 76, 77, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 106, 107, 108, 109, 111, 112, 113, 114, 115, 116, 117, 118, 119, 120, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140, 141, 142, 143, 144, 145, 146, 147, 148, 149, 150, 151, 152, 153, 154, 155, 156, 158, 159, 160, 161, 162, 163, 164, 165, 166, 167, 168, 169, 170, 171, 172, 173, 174, 175],
        'false': []
    },
    'Y': {  # Desirability
        'true': [2, 8, 16, 36, 41, 57, 68, 78, 105],
        'false': [12, 33, 45, 47, 59, 72, 110, 157]
    },
    'Z': {  # Debasement
        'true': [25, 31, 39, 50, 51, 53, 54, 56, 71, 82, 84, 86, 87, 89, 90, 94, 100, 104, 107, 113, 118, 121, 126, 131, 139, 145, 151, 158, 164, 172, 175],
        'false': []
    },

    # Clinical Personality Patterns
    '1': {  # Schizoid
        'true': [1, 14, 25, 40, 46, 55, 70, 81, 86, 92, 112, 118, 124, 130, 136],
        'false': [57, 78, 105]
    },
    '2A': {  # Avoidant
        'true': [1, 14, 25, 40, 46, 55, 62, 70, 81, 86, 92, 99, 107, 112, 118, 124, 130, 135, 141, 148, 154, 160, 166],
        'false': []
    },
    '2B': {  # Depressive
        'true': [1, 10, 14, 24, 25, 31, 40, 46, 50, 55, 62, 70, 75, 81, 86, 92, 99, 107, 112, 118, 121, 124, 127, 130, 135, 139, 148, 154, 160, 166, 172],
        'false': []
    },
    '3': {  # Dependent
        'true': [4, 10, 17, 24, 29, 34, 40, 46, 52, 58, 62, 66, 74, 79, 84, 91, 97, 103, 109, 115, 121, 127, 133, 139, 145, 151, 158, 164, 170],
        'false': []
    },
    '4': {  # Histrionic
        'true': [2, 8, 16, 23, 36, 41, 48, 57, 63, 68, 76, 78, 83, 88, 95, 101, 105, 108, 114, 120, 126, 132, 138, 144, 150, 156, 162, 168, 174],
        'false': []
    },
    '5': {  # Narcissistic (raw score multiplied by 2/3)
        'true': [2, 8, 16, 23, 30, 36, 41, 48, 57, 63, 68, 76, 78, 83, 88, 95, 101, 105, 108, 114, 120, 126, 132, 138, 144, 150, 156, 162, 168, 174],
        'false': [50, 86, 145, 175]
    },

    '6A': {  # Antisocial
        'true': [5, 15, 21, 27, 35, 42, 44, 53, 60, 64, 69, 73, 77, 82, 87, 93, 98, 102, 106, 111, 117, 123, 129, 134, 140, 146, 152, 159, 165, 171],
        'false': []
    },
    '6B': {  # Sadistic (Aggressive)
        'true': [5, 15, 21, 27, 35, 42, 44, 53, 60, 64, 69, 73, 77, 82, 87, 93, 98, 102, 106, 111, 117, 123, 129, 140, 146, 152],
        'false': []
    },
    '7': {  # Compulsive
        'true': [2, 8, 16, 36, 41, 57, 68, 78, 105],
        'false': [5, 15, 21, 27, 35, 42, 44, 53, 60, 64, 69, 73, 77, 82, 87, 93]
    },
    '8A': {  # Negativistic (Passive-Aggressive)
        'true': [5, 7, 15, 21, 27, 30, 35, 37, 42, 44, 51, 53, 56, 60, 64, 66, 69, 73, 75, 77, 80, 82, 85, 87, 89, 93, 96, 98, 100, 102, 104, 106, 111, 113, 116, 117, 119, 123, 125, 129, 131, 134, 137, 140, 143, 146, 149, 152, 155, 159, 161, 165, 167, 171, 173],
        'false': []
    },
    '8B': {  # Masochistic (Self-Defeating)
        'true': [1, 10, 14, 24, 25, 31, 40, 46, 50, 55, 62, 70, 75, 81, 84, 86, 89, 92, 97, 99, 103, 107, 109, 112, 115, 118, 121, 124, 127, 130, 133, 135, 139, 141, 145, 148, 151, 154, 158, 160, 164, 166, 170, 172, 175],
        'false': []
    },

    # Severe Personality Pathology
    'S': {  # Schizotypal
        'true': [1, 14, 25, 40, 46, 55, 62, 70, 81, 86, 90, 92, 94, 99, 107, 112, 118, 124, 130, 135, 136, 141, 148, 154, 160, 166],
        'false': [57, 78, 105]
    },
    'C': {  # Borderline
        'true': [5, 7, 15, 21, 27, 30, 35, 37, 42, 44, 51, 53, 56, 60, 62, 64, 66, 69, 73, 75, 77, 80, 82, 85, 87, 89, 93, 96, 98, 100, 102, 104, 106, 111, 113, 116, 117, 119, 123, 125, 129, 131, 134, 137, 140, 143, 146, 149, 152, 155, 159, 161, 165, 167, 171, 173],
        'false': []
    },
    'P': {  # Paranoid
        'true': [5, 15, 21, 27, 30, 35, 42, 44, 48, 53, 60, 63, 64, 69, 73, 76, 77, 82, 83, 87, 88, 93, 95, 98, 101, 102, 106, 108, 111, 114, 117, 120, 123, 126, 129, 132, 138, 140, 144, 146, 150, 152, 156, 159, 162, 165, 168, 171, 174],
        'false': []
    },

    # Clinical Syndromes
    'A': {  # Anxiety
        'true': [6, 10, 17, 24, 29, 34, 43, 50, 51, 54, 56, 62, 66, 71, 74, 79, 82, 84, 89, 91, 97, 100, 103, 107, 109, 113, 115, 118, 121, 127, 131, 133, 139, 145, 151, 158, 164, 170, 172, 175],
        'false': []
    },
    'H': {  # Somatoform
        'true': [6, 17, 24, 29, 34, 43, 46, 52, 54, 58, 62, 66, 74, 79, 84, 86, 91, 92, 97, 103, 109, 112, 115, 118, 121, 124, 127, 130, 133, 139, 145, 151, 158, 164, 170],
        'false': []
    },
    'N': {  # Bipolar: Manic
        'true': [2, 8, 16, 23, 30, 36, 41, 48, 57, 63, 68, 76, 78, 83, 88, 95, 101, 105, 108, 114, 120, 126, 132, 138, 144, 150, 156, 162, 168, 174],
        'false': []
    },
    'D': {  # Dysthymia
        'true': [1, 10, 14, 24, 25, 31, 40, 46, 50, 55, 62, 70, 75, 81, 84, 86, 89, 92, 99, 107, 112, 118, 121, 124, 127, 130, 135, 139, 141, 145, 148, 151, 154, 158, 160, 164, 166, 170, 172, 175],
        'false': []
    },
    'B': {  # Alcohol Dependence
        'true': [5, 15, 21, 27, 35, 42, 44, 53, 60, 64, 69, 73, 77, 82, 87, 93, 98, 106, 117, 123, 129, 134, 140, 146, 152, 159, 165, 171],
        'false': []
    },
    'T': {  # Drug Dependence
        'true': [5, 15, 21, 27, 35, 42, 44, 53, 60, 64, 69, 73, 77, 82, 87, 93, 98, 102, 106, 111, 117, 123, 129, 134, 140, 146, 152, 159, 165, 171],
        'false': []
    },
    'R': {  # PTSD
        'true': [6, 10, 17, 24, 29, 34, 43, 50, 51, 54, 56, 62, 66, 71, 74, 79, 82, 84, 89, 91, 97, 100, 103, 107, 109, 113, 115, 118, 121, 127, 131, 133, 139, 145, 151, 158, 164, 170, 172, 175],
        'false': []
    },

    # Severe Clinical Syndromes
    'SS': {  # Thought Disorder
        'true': [1, 14, 25, 40, 46, 55, 62, 70, 81, 86, 90, 92, 94, 99, 107, 112, 118, 124, 130, 135, 136, 141, 148, 154, 160, 166, 172, 175],
        'false': []
    },
    'CC': {  # Major Depression
        'true': [1, 10, 14, 24, 25, 31, 40, 46, 50, 55, 62, 70, 75, 81, 84, 86, 89, 92, 99, 107, 112, 118, 121, 124, 127, 130, 135, 139, 141, 145, 148, 151, 154, 158, 160, 164, 166, 170, 172, 175],
        'false': []
    },
    'PP': {  # Delusional Disorder
        'true': [5, 15, 21, 27, 30, 35, 42, 44, 48, 53, 60, 63, 64, 69, 73, 76, 77, 82, 83, 87, 88, 93, 95, 98, 101, 102, 106, 108, 111, 114, 117, 120, 123, 126, 129, 132, 138, 140, 144, 146, 150, 152, 156, 159, 162, 165, 168, 171, 174],
        'false': []
    },
    # Validity
    'V': {  # Invalidity (items 65, 110, 157)
        'true': [65, 110, 157],
        'false': []
    },
}


# ============================================================
# SCALE DEFINITIONS (order for scoring table)
# ============================================================
SCALES = [
    ('X', 'Disclosure', 'Modifying Indices'),
    ('Y', 'Desirability', 'Modifying Indices'),
    ('Z', 'Debasement', 'Modifying Indices'),
    ('1', 'Schizoid', 'Clinical Personality Patterns'),
    ('2A', 'Avoidant', 'Clinical Personality Patterns'),
    ('2B', 'Depressive', 'Clinical Personality Patterns'),
    ('3', 'Dependent', 'Clinical Personality Patterns'),
    ('4', 'Histrionic', 'Clinical Personality Patterns'),
    ('5', 'Narcissistic', 'Clinical Personality Patterns'),
    ('6A', 'Antisocial', 'Clinical Personality Patterns'),
    ('6B', 'Sadistic (Aggressive)', 'Clinical Personality Patterns'),
    ('7', 'Compulsive', 'Clinical Personality Patterns'),
    ('8A', 'Negativistic (Passive-Aggressive)', 'Clinical Personality Patterns'),
    ('8B', 'Masochistic (Self-Defeating)', 'Clinical Personality Patterns'),
    ('S', 'Schizotypal', 'Severe Personality Pathology'),
    ('C', 'Borderline', 'Severe Personality Pathology'),
    ('P', 'Paranoid', 'Severe Personality Pathology'),
    ('A', 'Anxiety', 'Clinical Syndromes'),
    ('H', 'Somatoform', 'Clinical Syndromes'),
    ('N', 'Bipolar: Manic', 'Clinical Syndromes'),
    ('D', 'Dysthymia', 'Clinical Syndromes'),
    ('B', 'Alcohol Dependence', 'Clinical Syndromes'),
    ('T', 'Drug Dependence', 'Clinical Syndromes'),
    ('R', 'PTSD', 'Clinical Syndromes'),
    ('SS', 'Thought Disorder', 'Severe Clinical Syndromes'),
    ('CC', 'Major Depression', 'Severe Clinical Syndromes'),
    ('PP', 'Delusional Disorder', 'Severe Clinical Syndromes'),
]


# ============================================================
# DISCLOSURE ADJUSTMENT TABLE (from Appendix/existing data)
# Format: raw_score -> (1-8B_adj, S-PP_adj)
# ============================================================
DISCLOSURE_TABLE = {
    0: (20, 10), 1: (20, 10), 2: (20, 10), 3: (20, 10), 4: (20, 10),
    5: (20, 10), 6: (20, 10), 7: (20, 10), 8: (20, 10), 9: (20, 10),
    10: (20, 10), 11: (20, 10), 12: (20, 10), 13: (20, 10), 14: (20, 10),
    15: (20, 10), 16: (20, 10), 17: (20, 10), 18: (20, 10), 19: (20, 10),
    20: (20, 10), 21: (20, 10), 22: (20, 10), 23: (20, 10), 24: (20, 10),
    25: (20, 10), 26: (20, 10), 27: (20, 10), 28: (20, 10), 29: (20, 10),
    30: (20, 10), 31: (20, 10), 32: (20, 10), 33: (20, 10), 34: (20, 10),
    35: (20, 10), 36: (20, 10), 37: (19, 10), 38: (18, 10), 39: (17, 9),
    40: (17, 9), 41: (16, 9), 42: (15, 8), 43: (14, 8), 44: (13, 7),
    45: (13, 7), 46: (12, 7), 47: (11, 6), 48: (10, 6), 49: (9, 5),
    50: (9, 5), 51: (8, 5), 52: (7, 4), 53: (6, 4), 54: (5, 3),
    55: (5, 3), 56: (4, 3), 57: (3, 2), 58: (2, 2), 59: (1, 1),
    60: (1, 1), 61: (0, 0), 62: (0, 0), 63: (0, 0), 64: (0, 0),
    65: (0, 0), 66: (0, 0), 67: (0, 0), 68: (0, 0), 69: (0, 0),
    70: (0, 0), 71: (0, 0), 72: (0, 0), 73: (0, 0), 74: (0, 0),
    75: (0, 0), 76: (0, 0), 77: (0, 0), 78: (0, 0), 79: (0, 0),
    80: (0, 0), 81: (0, 0), 82: (0, 0), 83: (0, 0), 84: (0, 0),
    85: (0, 0), 86: (0, 0), 87: (0, 0), 88: (0, 0), 89: (0, 0),
    90: (0, 0), 91: (0, 0), 92: (0, 0), 93: (0, 0), 94: (0, 0),
    95: (0, 0), 96: (0, 0), 97: (0, 0), 98: (0, 0), 99: (0, 0),
    100: (0, 0), 101: (0, 0), 102: (0, 0), 103: (0, 0), 104: (0, 0),
    105: (0, 0), 106: (0, 0), 107: (0, 0), 108: (0, 0), 109: (0, 0),
    110: (0, 0), 111: (0, 0), 112: (0, 0), 113: (0, 0), 114: (0, 0),
    115: (0, 0), 116: (0, 0), 117: (0, 0), 118: (0, 0), 119: (0, 0),
    120: (0, 0), 121: (0, 0), 122: (0, 0), 123: (0, 0),

    124: (-1, -1), 125: (-1, -1), 126: (-1, -1), 127: (-2, -2),
    128: (-2, -2), 129: (-3, -2), 130: (-3, -2), 131: (-3, -2),
    132: (-4, -3), 133: (-4, -3), 134: (-5, -3), 135: (-5, -3),
    136: (-5, -3), 137: (-6, -4), 138: (-6, -4), 139: (-7, -4),
    140: (-7, -4), 141: (-7, -4), 142: (-8, -5), 143: (-8, -5),
    144: (-9, -5), 145: (-9, -5), 146: (-9, -5), 147: (-10, -6),
    148: (-10, -6), 149: (-11, -6), 150: (-11, -6), 151: (-11, -6),
    152: (-12, -7), 153: (-12, -7), 154: (-13, -7), 155: (-13, -7),
    156: (-13, -7), 157: (-14, -8), 158: (-14, -8), 159: (-15, -8),
    160: (-15, -8), 161: (-15, -8), 162: (-16, -9), 163: (-16, -9),
    164: (-17, -9), 165: (-17, -9), 166: (-17, -9), 167: (-18, -10),
    168: (-18, -10), 169: (-19, -10), 170: (-19, -10), 171: (-19, -10),
    172: (-20, -11), 173: (-20, -11), 174: (-20, -11), 175: (-20, -11),
    176: (-20, -11), 177: (-20, -11), 178: (-20, -11), 179: (-20, -11),
    180: (-20, -11), 181: (-20, -11), 182: (-20, -11), 183: (-20, -11),
    184: (-20, -11), 185: (-20, -11), 186: (-20, -11), 187: (-20, -11),
    188: (-20, -11), 189: (-20, -11), 190: (-20, -11), 191: (-20, -11),
    192: (-20, -11), 193: (-20, -11), 194: (-20, -11), 195: (-20, -11),
    196: (-20, -11), 197: (-20, -11), 198: (-20, -11), 199: (-20, -11),
}


# ============================================================
# A/D ADJUSTMENT TABLES
# ============================================================
# Table for scales 2A and S
AD_2AS_TABLE = {}
for i in range(0, 81):
    if i < 8: AD_2AS_TABLE[i] = -1
    elif i < 16: AD_2AS_TABLE[i] = -2
    elif i < 24: AD_2AS_TABLE[i] = -3
    elif i < 32: AD_2AS_TABLE[i] = -4
    elif i < 40: AD_2AS_TABLE[i] = -5
    elif i < 48: AD_2AS_TABLE[i] = -6
    elif i < 56: AD_2AS_TABLE[i] = -7
    elif i < 64: AD_2AS_TABLE[i] = -8
    elif i < 72: AD_2AS_TABLE[i] = -9
    else: AD_2AS_TABLE[i] = -10

# Table for scales 2B, 8B, and C
AD_2B8BC_TABLE = {}
for i in range(0, 81):
    if i < 10: AD_2B8BC_TABLE[i] = -1
    elif i < 15: AD_2B8BC_TABLE[i] = -2
    elif i < 20: AD_2B8BC_TABLE[i] = -3
    elif i < 25: AD_2B8BC_TABLE[i] = -4
    elif i < 30: AD_2B8BC_TABLE[i] = -5
    elif i < 35: AD_2B8BC_TABLE[i] = -6
    elif i < 40: AD_2B8BC_TABLE[i] = -7
    elif i < 45: AD_2B8BC_TABLE[i] = -8
    elif i < 50: AD_2B8BC_TABLE[i] = -9
    elif i < 55: AD_2B8BC_TABLE[i] = -10
    elif i < 60: AD_2B8BC_TABLE[i] = -11
    elif i < 65: AD_2B8BC_TABLE[i] = -12
    elif i < 70: AD_2B8BC_TABLE[i] = -13
    elif i < 75: AD_2B8BC_TABLE[i] = -14
    else: AD_2B8BC_TABLE[i] = -15


# ============================================================
# GROSSMAN FACET SCALES (from Appendix D / Images 7-9)
# ============================================================
GROSSMAN_FACETS = {
    '1': [  # Schizoid facets
        ('1A', 'Expressively Impassive', [1, 14, 40, 70, 112]),
        ('1B', 'Interpersonally Unengaged', [25, 46, 81, 118, 136]),
        ('1C', 'Cognitively Impoverished', [55, 86, 92, 124, 130]),
    ],
    '2A': [  # Avoidant facets
        ('2A-A', 'Expressively Fretful', [1, 14, 40, 70, 112]),
        ('2A-B', 'Interpersonally Aversive', [25, 46, 81, 118, 135]),
        ('2A-C', 'Cognitively Distracted', [55, 86, 92, 124, 141]),
    ],
    '2B': [  # Depressive facets
        ('2B-A', 'Expressively Disconsolate', [10, 24, 50, 75, 121]),
        ('2B-B', 'Interpersonally Defenseless', [31, 62, 99, 135, 160]),
        ('2B-C', 'Cognitively Fatalistic', [14, 55, 92, 130, 166]),
    ],
    '3': [  # Dependent facets
        ('3A', 'Expressively Incompetent', [4, 17, 34, 52, 74]),
        ('3B', 'Interpersonally Submissive', [10, 29, 58, 79, 103]),
        ('3C', 'Cognitively Naive', [24, 46, 66, 91, 115]),
    ],
    '4': [  # Histrionic facets
        ('4A', 'Expressively Dramatic', [2, 23, 48, 76, 108]),
        ('4B', 'Interpersonally Attention-Seeking', [8, 36, 63, 88, 120]),
        ('4C', 'Cognitively Flighty', [16, 41, 68, 95, 132]),
    ],
    '5': [  # Narcissistic facets
        ('5A', 'Expressively Haughty', [2, 23, 48, 76, 108]),
        ('5B', 'Interpersonally Exploitative', [8, 30, 63, 88, 114]),
        ('5C', 'Cognitively Expansive', [16, 41, 68, 95, 132]),
    ],
    '6A': [  # Antisocial facets
        ('6A-A', 'Expressively Impulsive', [5, 21, 42, 64, 87]),
        ('6A-B', 'Interpersonally Irresponsible', [15, 35, 60, 82, 106]),
        ('6A-C', 'Cognitively Deviant', [27, 44, 69, 93, 117]),
    ],
    '6B': [  # Sadistic facets
        ('6B-A', 'Expressively Precipitate', [5, 21, 42, 64, 87]),
        ('6B-B', 'Interpersonally Abrasive', [15, 35, 60, 82, 106]),
        ('6B-C', 'Cognitively Dogmatic', [27, 44, 69, 93, 117]),
    ],
    '7': [  # Compulsive facets
        ('7A', 'Expressively Disciplined', [2, 16, 41, 68, 105]),
        ('7B', 'Interpersonally Respectful', [8, 36, 57, 78]),
        ('7C', 'Cognitively Constricted', [2, 8, 16, 36, 41]),
    ],
    '8A': [  # Negativistic facets
        ('8A-A', 'Expressively Resentful', [7, 30, 51, 75, 100]),
        ('8A-B', 'Interpersonally Contrary', [37, 56, 80, 104, 125]),
        ('8A-C', 'Cognitively Skeptical', [66, 85, 96, 113, 137]),
    ],
    '8B': [  # Masochistic facets
        ('8B-A', 'Expressively Abstinent', [10, 50, 89, 121, 151]),
        ('8B-B', 'Interpersonally Deferential', [24, 62, 97, 133, 164]),
        ('8B-C', 'Cognitively Diffident', [31, 75, 107, 139, 172]),
    ],
}


# ============================================================
# VALIDITY SCALE W - INCONSISTENCY ITEM PAIRS (from Image 10-11)
# Each pair: if both answered in same direction = inconsistent
# ============================================================
INCONSISTENCY_PAIRS = [
    (5, 105), (6, 86), (14, 55), (15, 78), (21, 57),
    (27, 68), (31, 95), (34, 114), (35, 36), (37, 101),
    (40, 63), (42, 132), (44, 120), (46, 88), (50, 144),
    (51, 138), (53, 76), (54, 150), (56, 162), (58, 156),
    (60, 168), (62, 174), (64, 126), (66, 108), (69, 83),
]

# ============================================================
# SHEET 1: RAW DATA INPUT (175 items)
# ============================================================
ws_input = wb.active
ws_input.title = "Item_Responses"

# Title
ws_input['A1'] = "MCMI-III SCORING TOOL - ITEM RESPONSE INPUT"
ws_input['A1'].font = TITLE_FONT
ws_input['A2'] = "Enter 1 for TRUE, 0 for FALSE for each of the 175 items"
ws_input['A2'].font = SUBTITLE_FONT

# Patient Info
ws_input['A4'] = "Patient Name:"
ws_input['B4'].fill = INPUT_FILL
ws_input['D4'] = "Date:"
ws_input['E4'].fill = INPUT_FILL
ws_input['G4'] = "Age:"
ws_input['H4'].fill = INPUT_FILL

ws_input['A5'] = "Gender (M/F):"
ws_input['B5'].fill = INPUT_FILL
ws_input['D5'] = "Setting:"
ws_input['E5'].fill = INPUT_FILL
ws_input['G5'] = "Examiner:"
ws_input['H5'].fill = INPUT_FILL

ws_input['A6'] = "Inpatient (Y/N):"
ws_input['B6'].fill = INPUT_FILL
ws_input['D6'] = "Axis I Duration (weeks):"
ws_input['E6'].fill = INPUT_FILL


# Item response grid - 175 items in columns
# Headers
ws_input['A8'] = "Item #"
ws_input['A8'].font = HEADER_FONT
ws_input['A8'].fill = HEADER_FILL
ws_input['B8'] = "Response (1=T, 0=F)"
ws_input['B8'].font = HEADER_FONT
ws_input['B8'].fill = HEADER_FILL
ws_input['D8'] = "Item #"
ws_input['D8'].font = HEADER_FONT
ws_input['D8'].fill = HEADER_FILL
ws_input['E8'] = "Response (1=T, 0=F)"
ws_input['E8'].font = HEADER_FONT
ws_input['E8'].fill = HEADER_FILL
ws_input['G8'] = "Item #"
ws_input['G8'].font = HEADER_FONT
ws_input['G8'].fill = HEADER_FILL
ws_input['H8'] = "Response (1=T, 0=F)"
ws_input['H8'].font = HEADER_FONT
ws_input['H8'].fill = HEADER_FILL

# Fill item numbers (3 columns of items: 1-60, 61-120, 121-175)
for i in range(1, 61):
    ws_input[f'A{8+i}'] = i
    ws_input[f'B{8+i}'].fill = INPUT_FILL
    ws_input[f'B{8+i}'].border = THIN_BORDER

for i in range(61, 121):
    ws_input[f'D{8+i-60}'] = i
    ws_input[f'E{8+i-60}'].fill = INPUT_FILL
    ws_input[f'E{8+i-60}'].border = THIN_BORDER

for i in range(121, 176):
    ws_input[f'G{8+i-120}'] = i
    ws_input[f'H{8+i-120}'].fill = INPUT_FILL
    ws_input[f'H{8+i-120}'].border = THIN_BORDER

# Set column widths
ws_input.column_dimensions['A'].width = 12
ws_input.column_dimensions['B'].width = 20
ws_input.column_dimensions['C'].width = 4
ws_input.column_dimensions['D'].width = 12
ws_input.column_dimensions['E'].width = 20
ws_input.column_dimensions['F'].width = 4
ws_input.column_dimensions['G'].width = 12
ws_input.column_dimensions['H'].width = 20


# ============================================================
# HELPER: Get cell reference for item number
# ============================================================
def get_item_cell(item_num):
    """Returns the cell reference in Item_Responses sheet for a given item number."""
    if 1 <= item_num <= 60:
        return f"Item_Responses!B{8 + item_num}"
    elif 61 <= item_num <= 120:
        return f"Item_Responses!E{8 + item_num - 60}"
    elif 121 <= item_num <= 175:
        return f"Item_Responses!H{8 + item_num - 120}"
    return ""

# ============================================================
# SHEET 2: RAW SCORE CALCULATION (automatic from items)
# ============================================================
ws_raw = wb.create_sheet("Raw_Scores")
ws_raw['A1'] = "MCMI-III - AUTOMATIC RAW SCORE CALCULATION"
ws_raw['A1'].font = TITLE_FONT
ws_raw['A2'] = "Raw scores are calculated automatically from item responses"
ws_raw['A2'].font = SUBTITLE_FONT

# Headers
headers_raw = ['Scale Code', 'Scale Name', 'Category', 'Items Scored TRUE', 'Items Scored FALSE', 'RAW SCORE']
for col, h in enumerate(headers_raw, 1):
    cell = ws_raw.cell(row=4, column=col)
    cell.value = h
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)


# Fill raw score calculations
for idx, (code, name, category) in enumerate(SCALES):
    row = 5 + idx
    ws_raw.cell(row=row, column=1, value=code)
    ws_raw.cell(row=row, column=2, value=name)
    ws_raw.cell(row=row, column=3, value=category)
    
    # Show items scored true
    true_items = SCORING_KEYS.get(code, {}).get('true', [])
    false_items = SCORING_KEYS.get(code, {}).get('false', [])
    ws_raw.cell(row=row, column=4, value=str(true_items)[:200] if true_items else "None")
    ws_raw.cell(row=row, column=5, value=str(false_items)[:200] if false_items else "None")
    
    # Build formula for raw score
    # Sum of TRUE items where response=1, plus FALSE items where response=0
    formula_parts = []
    for item in true_items:
        formula_parts.append(get_item_cell(item))
    
    false_parts = []
    for item in false_items:
        false_parts.append(f"(1-{get_item_cell(item)})")
    
    if formula_parts and false_parts:
        true_formula = "+".join(formula_parts)
        false_formula = "+".join(false_parts)
        full_formula = f"={true_formula}+{false_formula}"
    elif formula_parts:
        full_formula = f"={'+'.join(formula_parts)}"
    elif false_parts:
        full_formula = f"={'+'.join(false_parts)}"
    else:
        full_formula = "=0"
    
    # For Scale 5 (Narcissistic), multiply by 2/3
    if code == '5':
        full_formula = f"=ROUND(({full_formula[1:]})*2/3, 0)"
    
    ws_raw.cell(row=row, column=6, value=full_formula)
    ws_raw.cell(row=row, column=6).fill = GREEN_FILL
    ws_raw.cell(row=row, column=6).font = Font(bold=True)

# Also add Validity V score
row_v = 5 + len(SCALES)
ws_raw.cell(row=row_v, column=1, value='V')
ws_raw.cell(row=row_v, column=2, value='Invalidity')
ws_raw.cell(row=row_v, column=3, value='Validity')
ws_raw.cell(row=row_v, column=4, value="65, 110, 157")
v_formula = f"={get_item_cell(65)}+{get_item_cell(110)}+{get_item_cell(157)}"
ws_raw.cell(row=row_v, column=6, value=v_formula)
ws_raw.cell(row=row_v, column=6).fill = GREEN_FILL

# Column widths
ws_raw.column_dimensions['A'].width = 12
ws_raw.column_dimensions['B'].width = 35
ws_raw.column_dimensions['C'].width = 28
ws_raw.column_dimensions['D'].width = 50
ws_raw.column_dimensions['E'].width = 30
ws_raw.column_dimensions['F'].width = 14


# ============================================================
# SHEET 3: DISCLOSURE TABLE (lookup data)
# ============================================================
ws_disc = wb.create_sheet("Disclosure_Table")
ws_disc.cell(row=1, column=1, value="Scale X Raw Score")
ws_disc.cell(row=1, column=2, value="1-8B Adjustment")
ws_disc.cell(row=1, column=3, value="S-PP Adjustment")
ws_disc.cell(row=1, column=1).font = HEADER_FONT
ws_disc.cell(row=1, column=2).font = HEADER_FONT
ws_disc.cell(row=1, column=3).font = HEADER_FONT
ws_disc.cell(row=1, column=1).fill = HEADER_FILL
ws_disc.cell(row=1, column=2).fill = HEADER_FILL
ws_disc.cell(row=1, column=3).fill = HEADER_FILL

for raw_score in range(200):
    adj_1_8B, adj_S_PP = DISCLOSURE_TABLE.get(raw_score, (0, 0))
    ws_disc.cell(row=raw_score + 2, column=1, value=raw_score)
    ws_disc.cell(row=raw_score + 2, column=2, value=adj_1_8B)
    ws_disc.cell(row=raw_score + 2, column=3, value=adj_S_PP)

# ============================================================
# SHEET 4: AD_2AS TABLE
# ============================================================
ws_ad2as = wb.create_sheet("AD_2AS")
ws_ad2as.cell(row=1, column=1, value="A/D Value")
ws_ad2as.cell(row=1, column=2, value="2A, S Adjustment")
ws_ad2as.cell(row=1, column=1).font = HEADER_FONT
ws_ad2as.cell(row=1, column=2).font = HEADER_FONT
ws_ad2as.cell(row=1, column=1).fill = HEADER_FILL
ws_ad2as.cell(row=1, column=2).fill = HEADER_FILL

for val in range(81):
    ws_ad2as.cell(row=val + 2, column=1, value=val)
    ws_ad2as.cell(row=val + 2, column=2, value=AD_2AS_TABLE[val])

# ============================================================
# SHEET 5: AD_2B8BC TABLE
# ============================================================
ws_ad2b = wb.create_sheet("AD_2B8BC")
ws_ad2b.cell(row=1, column=1, value="A/D Value")
ws_ad2b.cell(row=1, column=2, value="2B, 8B, C Adjustment")
ws_ad2b.cell(row=1, column=1).font = HEADER_FONT
ws_ad2b.cell(row=1, column=2).font = HEADER_FONT
ws_ad2b.cell(row=1, column=1).fill = HEADER_FILL
ws_ad2b.cell(row=1, column=2).fill = HEADER_FILL

for val in range(81):
    ws_ad2b.cell(row=val + 2, column=1, value=val)
    ws_ad2b.cell(row=val + 2, column=2, value=AD_2B8BC_TABLE[val])


# ============================================================
# SHEET 6: MAIN SCORING TABLE (BR Input + Adjustments + Final)
# ============================================================
ws_score = wb.create_sheet("Scoring_Results")
ws_score['A1'] = "MCMI-III COMPLETE SCORING RESULTS"
ws_score['A1'].font = TITLE_FONT
ws_score['A2'] = "Enter Initial BR Scores from Appendix C (or use BR_Lookup sheet). All adjustments auto-calculated."
ws_score['A2'].font = SUBTITLE_FONT

# Validity section
ws_score['A4'] = "VALIDITY CHECKS"
ws_score['A4'].font = SUBTITLE_FONT
ws_score['A4'].fill = SECTION_FILL
ws_score['A5'] = "V (Invalidity) Score:"
ws_score['B5'] = "=Raw_Scores!F32"  # V score from raw scores
ws_score['C5'] = '=IF(B5>1,"INVALID - V > 1","VALID")'
ws_score['A6'] = "Scale X Raw Score:"
ws_score['B6'] = "=Raw_Scores!F5"  # X raw score
ws_score['C6'] = '=IF(OR(B6<34,B6>178),"INVALID - X out of range","VALID")'
ws_score['A7'] = "Protocol Status:"
ws_score['B7'] = '=IF(OR(B5>1,B6<34,B6>178),"INVALID","VALID - Proceed")'
ws_score['B7'].font = Font(bold=True, size=12)

# Main scoring table
row_start = 9
headers = ['Scale\nCode', 'Scale Name', 'Category', 'RAW\nSCORE', 'Initial BR\n(Enter from\nAppendix C)', 
           'Disclosure\nAdjustment', 'A/D\nAdjustment', 'Inpatient\nAdjustment', 
           'Denial/Complaint\nAdj. (manual)', 'FINAL\nBR SCORE', 'Clinical\nSignificance']


for col, h in enumerate(headers, 1):
    cell = ws_score.cell(row=row_start, column=col)
    cell.value = h
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    cell.border = THIN_BORDER

# Fill each scale row
# Scales that get 1-8B disclosure adjustment: 1, 2A, 2B, 3, 4, 5, 6A, 6B, 7, 8A, 8B
# Scales that get S-PP disclosure adjustment: S, C, P, A, H, N, D, B, T, R, SS, CC, PP
# A/D adjustment applies to: 2A, S (use AD_2AS), 2B, 8B, C (use AD_2B8BC)
# Inpatient adjustment: SS (+6/+4/0), CC (+10/+8/0), PP (+4/+2/0)

PERSONALITY_SCALES = ['1', '2A', '2B', '3', '4', '5', '6A', '6B', '7', '8A', '8B']
SEVERE_PERS = ['S', 'C', 'P']
CLINICAL_SYN = ['A', 'H', 'N', 'D', 'B', 'T', 'R']
SEVERE_CLIN = ['SS', 'CC', 'PP']

AD_2AS_SCALES = ['2A', 'S']
AD_2B8BC_SCALES = ['2B', '8B', 'C']

for idx, (code, name, category) in enumerate(SCALES):
    row = row_start + 1 + idx
    
    # Scale code
    ws_score.cell(row=row, column=1, value=code)
    ws_score.cell(row=row, column=1).border = THIN_BORDER
    ws_score.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    
    # Scale name
    ws_score.cell(row=row, column=2, value=name)
    ws_score.cell(row=row, column=2).border = THIN_BORDER
    
    # Category
    ws_score.cell(row=row, column=3, value=category)
    ws_score.cell(row=row, column=3).border = THIN_BORDER
    
    # Raw score (linked from Raw_Scores sheet)
    raw_row = 5 + idx
    ws_score.cell(row=row, column=4, value=f"=Raw_Scores!F{raw_row}")
    ws_score.cell(row=row, column=4).border = THIN_BORDER
    ws_score.cell(row=row, column=4).fill = LOCKED_FILL
    
    # Initial BR Score - USER INPUT (yellow)
    ws_score.cell(row=row, column=5).fill = INPUT_FILL
    ws_score.cell(row=row, column=5).border = THIN_BORDER
    ws_score.cell(row=row, column=5).alignment = Alignment(horizontal='center')

    
    # Disclosure Adjustment (Column F = col 6)
    # X raw score is in row 10 (row_start+1), column 4
    x_raw_cell = f"D{row_start + 1}"  # Scale X raw score cell
    
    if code in ['X', 'Y', 'Z']:
        # No disclosure adjustment for modifying indices
        ws_score.cell(row=row, column=6, value="N/A")
    elif code in PERSONALITY_SCALES:
        # Use column B (1-8B adjustment)
        formula = f'=IF({x_raw_cell}="","",VLOOKUP(MIN(MAX({x_raw_cell},0),199),Disclosure_Table!A:B,2,TRUE))'
        ws_score.cell(row=row, column=6, value=formula)
    else:
        # S-PP adjustment (column C)
        formula = f'=IF({x_raw_cell}="","",VLOOKUP(MIN(MAX({x_raw_cell},0),199),Disclosure_Table!A:C,3,TRUE))'
        ws_score.cell(row=row, column=6, value=formula)
    ws_score.cell(row=row, column=6).border = THIN_BORDER
    
    # A/D Adjustment (Column G = col 7)
    # A is at row_start+18 (index 17), D is at row_start+21 (index 20)
    a_br_row = row_start + 1 + 17  # Anxiety row
    d_br_row = row_start + 1 + 20  # Dysthymia row
    a_adj_br = f"(E{a_br_row}+IF(ISNUMBER(F{a_br_row}),F{a_br_row},0))"
    d_adj_br = f"(E{d_br_row}+IF(ISNUMBER(F{d_br_row}),F{d_br_row},0))"
    
    if code in AD_2AS_SCALES:
        ad_formula = (
            f'=IF(OR(E{a_br_row}="",E{d_br_row}=""),"",IF(AND({a_adj_br}<75,{d_adj_br}<75),0,'
            f'IF(AND({a_adj_br}>=75,{d_adj_br}<75),VLOOKUP({a_adj_br}-75,AD_2AS!A:B,2,TRUE),'
            f'IF(AND({d_adj_br}>=75,{a_adj_br}<75),VLOOKUP({d_adj_br}-75,AD_2AS!A:B,2,TRUE),'
            f'VLOOKUP(MIN(({a_adj_br}-75)+({d_adj_br}-75),80),AD_2AS!A:B,2,TRUE)))))'
        )
        ws_score.cell(row=row, column=7, value=ad_formula)
    elif code in AD_2B8BC_SCALES:
        ad_formula = (
            f'=IF(OR(E{a_br_row}="",E{d_br_row}=""),"",IF(AND({a_adj_br}<75,{d_adj_br}<75),0,'
            f'IF(AND({a_adj_br}>=75,{d_adj_br}<75),VLOOKUP({a_adj_br}-75,AD_2B8BC!A:B,2,TRUE),'
            f'IF(AND({d_adj_br}>=75,{a_adj_br}<75),VLOOKUP({d_adj_br}-75,AD_2B8BC!A:B,2,TRUE),'
            f'VLOOKUP(MIN(({a_adj_br}-75)+({d_adj_br}-75),80),AD_2B8BC!A:B,2,TRUE)))))'
        )
        ws_score.cell(row=row, column=7, value=ad_formula)
    else:
        ws_score.cell(row=row, column=7, value="N/A")
    ws_score.cell(row=row, column=7).border = THIN_BORDER

    
    # Inpatient Adjustment (Column H = col 8)
    # Uses Item_Responses!B6 for inpatient Y/N and E6 for duration
    inpatient_cell = "Item_Responses!B6"
    duration_cell = "Item_Responses!E6"
    
    if code == 'SS':
        inp_formula = f'=IF({inpatient_cell}="Y",IF({duration_cell}<1,6,IF({duration_cell}<=4,4,0)),0)'
        ws_score.cell(row=row, column=8, value=inp_formula)
    elif code == 'CC':
        inp_formula = f'=IF({inpatient_cell}="Y",IF({duration_cell}<1,10,IF({duration_cell}<=4,8,0)),0)'
        ws_score.cell(row=row, column=8, value=inp_formula)
    elif code == 'PP':
        inp_formula = f'=IF({inpatient_cell}="Y",IF({duration_cell}<1,4,IF({duration_cell}<=4,2,0)),0)'
        ws_score.cell(row=row, column=8, value=inp_formula)
    else:
        ws_score.cell(row=row, column=8, value=0)
    ws_score.cell(row=row, column=8).border = THIN_BORDER
    
    # Denial/Complaint Adjustment (Column I = col 9) - Manual entry
    if code in ['X', 'Y', 'Z']:
        ws_score.cell(row=row, column=9, value=0)
    else:
        ws_score.cell(row=row, column=9, value=0)
        ws_score.cell(row=row, column=9).fill = INPUT_FILL
    ws_score.cell(row=row, column=9).border = THIN_BORDER
    
    # FINAL BR SCORE (Column J = col 10)
    if code in ['X', 'Y', 'Z']:
        # Modifying indices: just the initial BR, clamped 0-115
        final_formula = f'=IF(E{row}="","",MIN(MAX(E{row},0),115))'
    else:
        final_formula = (
            f'=IF(E{row}="","",MIN(MAX(E{row}'
            f'+IF(ISNUMBER(F{row}),F{row},0)'
            f'+IF(ISNUMBER(G{row}),G{row},0)'
            f'+IF(ISNUMBER(H{row}),H{row},0)'
            f'+IF(ISNUMBER(I{row}),I{row},0)'
            f',0),115))'
        )
    ws_score.cell(row=row, column=10, value=final_formula)
    ws_score.cell(row=row, column=10).border = THIN_BORDER
    ws_score.cell(row=row, column=10).font = Font(bold=True, size=12)
    
    # Clinical Significance (Column K = col 11)
    sig_formula = (
        f'=IF(J{row}="","",IF(J{row}>=85,"PROMINENT/PATHOLOGICAL",'
        f'IF(J{row}>=75,"PRESENT/TRAIT LEVEL",'
        f'IF(J{row}>=60,"Suggestive","Normal/Not Significant"))))'
    )
    ws_score.cell(row=row, column=11, value=sig_formula)
    ws_score.cell(row=row, column=11).border = THIN_BORDER


# Column widths for scoring results
ws_score.column_dimensions['A'].width = 10
ws_score.column_dimensions['B'].width = 32
ws_score.column_dimensions['C'].width = 28
ws_score.column_dimensions['D'].width = 10
ws_score.column_dimensions['E'].width = 14
ws_score.column_dimensions['F'].width = 14
ws_score.column_dimensions['G'].width = 12
ws_score.column_dimensions['H'].width = 14
ws_score.column_dimensions['I'].width = 16
ws_score.column_dimensions['J'].width = 12
ws_score.column_dimensions['K'].width = 26

# Add conditional formatting for Final BR scores (column J)
from openpyxl.formatting.rule import CellIsRule
red_font = Font(bold=True, color="9C0006")
yellow_font = Font(bold=True, color="9C6500")
green_font = Font(color="006100")

for idx in range(len(SCALES)):
    row = row_start + 1 + idx
    cell_range = f"J{row}:J{row}"
    ws_score.conditional_formatting.add(cell_range, CellIsRule(
        operator='greaterThanOrEqual', formula=['85'], 
        fill=RED_FILL, font=red_font))
    ws_score.conditional_formatting.add(cell_range, CellIsRule(
        operator='between', formula=['75', '84'],
        fill=YELLOW_FILL, font=yellow_font))
    ws_score.conditional_formatting.add(cell_range, CellIsRule(
        operator='lessThan', formula=['60'],
        fill=GREEN_FILL, font=green_font))


# ============================================================
# SHEET 7: GROSSMAN FACET SCALES
# ============================================================
ws_facet = wb.create_sheet("Grossman_Facets")
ws_facet['A1'] = "MCMI-III GROSSMAN FACET SCALE SCORES"
ws_facet['A1'].font = TITLE_FONT
ws_facet['A2'] = "Facet scores calculated automatically from item responses"
ws_facet['A2'].font = SUBTITLE_FONT

facet_headers = ['Parent Scale', 'Facet Code', 'Facet Name', 'Items', 'Raw Score', 'Interpretation']
for col, h in enumerate(facet_headers, 1):
    cell = ws_facet.cell(row=4, column=col)
    cell.value = h
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

facet_row = 5
for parent_code, facets in GROSSMAN_FACETS.items():
    parent_name = next((n for c, n, _ in SCALES if c == parent_code), parent_code)
    for facet_code, facet_name, items in facets:
        ws_facet.cell(row=facet_row, column=1, value=f"{parent_code} - {parent_name}")
        ws_facet.cell(row=facet_row, column=2, value=facet_code)
        ws_facet.cell(row=facet_row, column=3, value=facet_name)
        ws_facet.cell(row=facet_row, column=4, value=str(items))
        
        # Formula to sum item responses
        item_refs = [get_item_cell(i) for i in items]
        if item_refs:
            formula = "=" + "+".join(item_refs)
            ws_facet.cell(row=facet_row, column=5, value=formula)
        ws_facet.cell(row=facet_row, column=5).fill = GREEN_FILL
        ws_facet.cell(row=facet_row, column=5).font = Font(bold=True)
        
        # Interpretation
        max_score = len(items)
        interp_formula = (
            f'=IF(E{facet_row}="","",IF(E{facet_row}>={max_score}*0.7,"Elevated",'
            f'IF(E{facet_row}>={max_score}*0.4,"Moderate","Low")))'
        )
        ws_facet.cell(row=facet_row, column=6, value=interp_formula)
        
        for col in range(1, 7):
            ws_facet.cell(row=facet_row, column=col).border = THIN_BORDER
        
        facet_row += 1

ws_facet.column_dimensions['A'].width = 25
ws_facet.column_dimensions['B'].width = 12
ws_facet.column_dimensions['C'].width = 32
ws_facet.column_dimensions['D'].width = 30
ws_facet.column_dimensions['E'].width = 12
ws_facet.column_dimensions['F'].width = 16


# ============================================================
# SHEET 8: INCONSISTENCY SCALE (W)
# ============================================================
ws_incon = wb.create_sheet("Inconsistency_W")
ws_incon['A1'] = "MCMI-III INCONSISTENCY SCALE (W)"
ws_incon['A1'].font = TITLE_FONT
ws_incon['A2'] = "Counts item pairs where responses are inconsistent"
ws_incon['A2'].font = SUBTITLE_FONT

incon_headers = ['Pair #', 'Item A', 'Item B', 'Response A', 'Response B', 'Inconsistent?']
for col, h in enumerate(incon_headers, 1):
    cell = ws_incon.cell(row=4, column=col)
    cell.value = h
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

for idx, (item_a, item_b) in enumerate(INCONSISTENCY_PAIRS):
    row = 5 + idx
    ws_incon.cell(row=row, column=1, value=idx + 1)
    ws_incon.cell(row=row, column=2, value=item_a)
    ws_incon.cell(row=row, column=3, value=item_b)
    
    cell_a = get_item_cell(item_a)
    cell_b = get_item_cell(item_b)
    
    ws_incon.cell(row=row, column=4, value=f"={cell_a}")
    ws_incon.cell(row=row, column=5, value=f"={cell_b}")
    
    # Inconsistent = both TRUE or both FALSE
    incon_formula = f'=IF(OR(AND({cell_a}=1,{cell_b}=1),AND({cell_a}=0,{cell_b}=0)),1,0)'
    ws_incon.cell(row=row, column=6, value=incon_formula)
    
    for col in range(1, 7):
        ws_incon.cell(row=row, column=col).border = THIN_BORDER

# Total W score
total_row = 5 + len(INCONSISTENCY_PAIRS) + 1
ws_incon.cell(row=total_row, column=1, value="TOTAL W SCORE:")
ws_incon.cell(row=total_row, column=1).font = Font(bold=True, size=12)
ws_incon.cell(row=total_row, column=6, value=f"=SUM(F5:F{total_row-2})")
ws_incon.cell(row=total_row, column=6).font = Font(bold=True, size=12)
ws_incon.cell(row=total_row, column=6).fill = GREEN_FILL

ws_incon.cell(row=total_row+1, column=1, value="Interpretation:")
ws_incon.cell(row=total_row+1, column=2, value=f'=IF(F{total_row}>12,"INVALID - W too high (>12), consider random responding",IF(F{total_row}>8,"CAUTION - Elevated inconsistency","VALID - Acceptable consistency"))')

ws_incon.column_dimensions['A'].width = 16
ws_incon.column_dimensions['B'].width = 10
ws_incon.column_dimensions['C'].width = 10
ws_incon.column_dimensions['D'].width = 14
ws_incon.column_dimensions['E'].width = 14
ws_incon.column_dimensions['F'].width = 14


# ============================================================
# SHEET 9: PROFILE SUMMARY
# ============================================================
ws_profile = wb.create_sheet("Profile_Summary")
ws_profile['A1'] = "MCMI-III PROFILE SUMMARY"
ws_profile['A1'].font = TITLE_FONT
ws_profile['A2'] = "Summary of all scales with final BR scores and clinical interpretation"
ws_profile['A2'].font = SUBTITLE_FONT

# Patient info linked
ws_profile['A4'] = "Patient:"
ws_profile['B4'] = "=Item_Responses!B4"
ws_profile['D4'] = "Date:"
ws_profile['E4'] = "=Item_Responses!E4"

ws_profile['A5'] = "Validity:"
ws_profile['B5'] = "=Scoring_Results!B7"
ws_profile['B5'].font = Font(bold=True)

# Profile headers
prof_headers = ['Scale', 'Name', 'Final BR', 'Significance', 'BR Level']
for col, h in enumerate(prof_headers, 1):
    cell = ws_profile.cell(row=7, column=col)
    cell.value = h
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

for idx, (code, name, category) in enumerate(SCALES):
    row = 8 + idx
    score_row = 10 + idx  # row in Scoring_Results
    
    ws_profile.cell(row=row, column=1, value=code)
    ws_profile.cell(row=row, column=2, value=name)
    ws_profile.cell(row=row, column=3, value=f"=Scoring_Results!J{score_row}")
    ws_profile.cell(row=row, column=4, value=f"=Scoring_Results!K{score_row}")
    
    # Visual BR level bar (using REPT function)
    bar_formula = f'=IF(C{row}="","",REPT("|",MIN(INT(C{row}/5),23)))'
    ws_profile.cell(row=row, column=5, value=bar_formula)
    ws_profile.cell(row=row, column=5).font = Font(name="Consolas", size=9)
    
    for col in range(1, 6):
        ws_profile.cell(row=row, column=col).border = THIN_BORDER

# Add legend
legend_row = 8 + len(SCALES) + 2
ws_profile.cell(row=legend_row, column=1, value="INTERPRETATION KEY:")
ws_profile.cell(row=legend_row, column=1).font = SUBTITLE_FONT
ws_profile.cell(row=legend_row+1, column=1, value="BR 85-115")
ws_profile.cell(row=legend_row+1, column=2, value="PROMINENT / PATHOLOGICAL (Disorder present)")
ws_profile.cell(row=legend_row+1, column=1).fill = RED_FILL
ws_profile.cell(row=legend_row+2, column=1, value="BR 75-84")
ws_profile.cell(row=legend_row+2, column=2, value="PRESENT / TRAIT LEVEL (Clinically significant)")
ws_profile.cell(row=legend_row+2, column=1).fill = YELLOW_FILL
ws_profile.cell(row=legend_row+3, column=1, value="BR 60-74")
ws_profile.cell(row=legend_row+3, column=2, value="Suggestive (Subclinical features)")
ws_profile.cell(row=legend_row+4, column=1, value="BR 0-59")
ws_profile.cell(row=legend_row+4, column=2, value="Normal / Not Significant")
ws_profile.cell(row=legend_row+4, column=1).fill = GREEN_FILL

ws_profile.column_dimensions['A'].width = 12
ws_profile.column_dimensions['B'].width = 35
ws_profile.column_dimensions['C'].width = 12
ws_profile.column_dimensions['D'].width = 28
ws_profile.column_dimensions['E'].width = 30


# ============================================================
# SHEET 10: BR SCORE GUIDE / INTERPRETATION REFERENCE
# ============================================================
ws_guide = wb.create_sheet("BR_Score_Guide")
ws_guide['A1'] = "MCMI-III BR SCORE INTERPRETATION GUIDE"
ws_guide['A1'].font = TITLE_FONT

guide_data = [
    ('BR Score Range', 'Classification', 'Clinical Meaning', 'Action'),
    ('0-34', 'Not Significant', 'Trait not present', 'No concern'),
    ('35-59', 'Normal Range', 'Within normal limits', 'No concern'),
    ('60-74', 'Suggestive', 'Mild features; subclinical', 'Monitor'),
    ('75-84', 'Trait Level / Present', 'Clinically significant features', 'Include in report'),
    ('85-115', 'Pathological / Disorder', 'Prominent disorder present', 'Primary focus in report'),
]

for row_idx, row_data in enumerate(guide_data):
    for col_idx, val in enumerate(row_data):
        cell = ws_guide.cell(row=3 + row_idx, column=1 + col_idx)
        cell.value = val
        cell.border = THIN_BORDER
        if row_idx == 0:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

# Validity section
ws_guide['A11'] = "VALIDITY CRITERIA"
ws_guide['A11'].font = SUBTITLE_FONT
ws_guide['A12'] = "Scale V (Invalidity): > 1 = INVALID protocol"
ws_guide['A13'] = "Scale X (Disclosure): Raw < 34 or > 178 = INVALID"
ws_guide['A14'] = "Omissions: > 11 items omitted = INVALID"
ws_guide['A15'] = "Scale W (Inconsistency): > 12 = Consider random responding"

# Adjustment rules
ws_guide['A17'] = "ADJUSTMENT RULES"
ws_guide['A17'].font = SUBTITLE_FONT
ws_guide['A18'] = "1. Disclosure Adjustment: Based on Scale X raw score (see Disclosure_Table sheet)"
ws_guide['A19'] = "   - Scales 1-8B use '1-8B Adjustment' column"
ws_guide['A20'] = "   - Scales S-PP use 'S-PP Adjustment' column"
ws_guide['A21'] = "2. Anxiety/Depression (A/D) Adjustment:"
ws_guide['A22'] = "   - Only applies to scales 2A, 2B, 8B, S, C"
ws_guide['A23'] = "   - Based on adjusted BR scores of Anxiety (A) and Dysthymia (D)"
ws_guide['A24'] = "   - Only activates when A or D adjusted BR >= 75"
ws_guide['A25'] = "3. Inpatient Adjustment:"
ws_guide['A26'] = "   - SS: +6 (<1wk), +4 (1-4wk), 0 (>4wk)"
ws_guide['A27'] = "   - CC: +10 (<1wk), +8 (1-4wk), 0 (>4wk)"
ws_guide['A28'] = "   - PP: +4 (<1wk), +2 (1-4wk), 0 (>4wk)"
ws_guide['A29'] = "4. Denial/Complaint Adjustment:"
ws_guide['A30'] = "   - If highest personality scale (1-8B) is 4, 5, or 7: add 8 to that scale"
ws_guide['A31'] = "   - Otherwise: no adjustment"

ws_guide.column_dimensions['A'].width = 70
ws_guide.column_dimensions['B'].width = 25
ws_guide.column_dimensions['C'].width = 30
ws_guide.column_dimensions['D'].width = 25


# ============================================================
# SHEET 11: INSTRUCTIONS
# ============================================================
ws_instr = wb.create_sheet("Instructions")
ws_instr['A1'] = "MCMI-III SCORING TOOL - INSTRUCTIONS"
ws_instr['A1'].font = TITLE_FONT

instructions = [
    "",
    "HOW TO USE THIS SCORING TOOL:",
    "",
    "STEP 1: ENTER ITEM RESPONSES",
    "  - Go to 'Item_Responses' sheet",
    "  - Enter patient demographics at the top",
    "  - Enter Y or N for Inpatient status",
    "  - Enter Axis I duration in weeks",
    "  - For each of 175 items, enter 1 (TRUE) or 0 (FALSE)",
    "",
    "STEP 2: RAW SCORES (AUTOMATIC)",
    "  - Go to 'Raw_Scores' sheet",
    "  - All raw scores are calculated automatically from your item responses",
    "  - Scale 5 (Narcissistic) raw score is automatically multiplied by 2/3",
    "  - Verify against hand-scoring if desired",
    "",
    "STEP 3: ENTER INITIAL BR SCORES",
    "  - Go to 'Scoring_Results' sheet",
    "  - Look up each raw score in Appendix C of the MCMI-III Manual",
    "  - Enter the corresponding BR score in the yellow 'Initial BR' column (E)",
    "  - Raw scores are already displayed for your reference",
    "",
    "STEP 4: REVIEW ADJUSTMENTS (AUTOMATIC)",
    "  - Disclosure Adjustment: auto-calculated from Scale X raw score",
    "  - A/D Adjustment: auto-calculated for scales 2A, 2B, 8B, S, C",
    "  - Inpatient Adjustment: auto-calculated from patient setting info",
    "",
    "STEP 5: DENIAL/COMPLAINT ADJUSTMENT (MANUAL)",
    "  - Look at the Final BR scores for personality scales 1-8B",
    "  - If the HIGHEST is Scale 4, 5, or 7: enter 8 in column I for that scale",
    "  - Otherwise: leave at 0",
    "",
    "STEP 6: READ FINAL RESULTS",
    "  - 'Scoring_Results' sheet: Final BR + Clinical Significance",
    "  - 'Profile_Summary' sheet: Complete profile overview",
    "  - 'Grossman_Facets' sheet: Detailed facet-level analysis",
    "  - 'Inconsistency_W' sheet: Check response consistency",
    "",
    "VALIDITY CHECKS:",
    "  - V Score > 1: INVALID protocol",
    "  - Scale X Raw < 34 or > 178: INVALID protocol",
    "  - W Score > 12: Possible random responding",
    "",
    "NOTES:",
    "  - Yellow cells = user input required",
    "  - Green cells = auto-calculated",
    "  - Grey cells = locked/reference values",
    "  - BR scores are clamped between 0 and 115",
]

for idx, line in enumerate(instructions):
    ws_instr.cell(row=2 + idx, column=1, value=line)

ws_instr.column_dimensions['A'].width = 80


# ============================================================
# SAVE THE WORKBOOK
# ============================================================
output_path = "/projects/sandbox/Dango-kiro/MCMI-III_Complete_Scoring_Tool.xlsx"
wb.save(output_path)
print(f"SUCCESS! MCMI-III Scoring Tool created at: {output_path}")
print(f"\nSheets created:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
print(f"\nTotal scales: {len(SCALES)}")
print(f"Total Grossman Facets: {sum(len(f) for f in GROSSMAN_FACETS.values())}")
print(f"Inconsistency pairs: {len(INCONSISTENCY_PAIRS)}")
print(f"\nInstructions:")
print("1. Open the Excel file")
print("2. Go to 'Item_Responses' and enter 1/0 for all 175 items")
print("3. Go to 'Scoring_Results' and enter Initial BR from Appendix C")
print("4. All other calculations are AUTOMATIC!")
