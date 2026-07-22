#!/usr/bin/env python3
"""
MCMI-III SINGLE-SHEET AUTO-SCORING TOOL
========================================
Everything on ONE sheet:
- Items 1-175 entry (T/F) at TOP
- All scores calculated BELOW on same sheet
- No cross-sheet references. Just one sheet.
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

from mcmi_item_keys import (
    SCALE_ITEMS, SCALE_V_ITEMS, SCALE_W_PAIRS,
    CLINICAL_PERSONALITY, SEVERE_PERSONALITY,
    CLINICAL_SYNDROMES, SEVERE_SYNDROMES
)
from mcmi_br_tables import BR_CONVERSION_ANCHORS



# ===================== STYLES =====================
TITLE = Font(bold=True, size=14, color="1F4E79")
HDR = Font(bold=True, size=11, color="FFFFFF")
BOLD = Font(bold=True, size=11)
SMALL = Font(italic=True, size=9, color="888888")
HFILL = PatternFill("solid", fgColor="1F4E79")
YFILL = PatternFill("solid", fgColor="FFFDE6")
GFILL = PatternFill("solid", fgColor="C6EFCE")
CTR = Alignment(horizontal='center', vertical='center')
BDR = Border(left=Side('thin'), right=Side('thin'),
             top=Side('thin'), bottom=Side('thin'))


# ===================== HELPERS =====================

# Item responses are in column B, rows 5 to 179
# Item 1 = B5, Item 2 = B6, ... Item N = B(N+4)
def R(item_num):
    """Cell ref for item N response (same sheet, column B)."""
    return f"B{item_num + 4}"


def raw_formula(scale_items):
    """Build raw score SUM formula from item weights."""
    parts = []
    for item_num, weight in sorted(scale_items):
        ref = R(item_num)
        if weight == 1:
            parts.append(f'IF({ref}="T",1,0)')
        else:
            parts.append(f'IF({ref}="T",{weight},0)')
    return "=" + "+".join(parts)


def br_formula(scale_name, raw_cell):
    """Piecewise linear interpolation for BR conversion."""
    if scale_name not in BR_CONVERSION_ANCHORS:
        return "=0"
    anchors = BR_CONVERSION_ANCHORS[scale_name]
    inner = str(anchors[0][1])
    for i in range(len(anchors) - 1):
        lo_r, lo_b = anchors[i]
        hi_r, hi_b = anchors[i + 1]
        if hi_r == lo_r:
            continue
        slope = round((hi_b - lo_b) / (hi_r - lo_r), 4)
        seg = f"{lo_b}+{slope}*({raw_cell}-{lo_r})"
        inner = f"IF({raw_cell}>={lo_r},{seg},{inner})"
    return f"=MIN(115,MAX(0,ROUND({inner},0)))"


def disc_formula(x_cell, which):
    """Disclosure adjustment lookup. which='18B' or 'SPP'."""
    if which == '18B':
        ranges = [
            (37,20),(38,19),(39,18),(41,17),(42,16),(43,15),
            (44,14),(45,13),(46,12),(47,11),(48,10),(50,9),
            (51,8),(52,7),(53,6),(54,5),(55,4),(56,3),(58,2),(60,1),
        ]
        neg = [
            (125,-1),(128,-2),(130,-3),(132,-4),(134,-5),
            (137,-6),(139,-7),(141,-8),(144,-9),(147,-10),
            (151,-11),(153,-12),(156,-13),(158,-14),
            (161,-15),(163,-16),(166,-17),(168,-18),(171,-19),(178,-20),
        ]
    else:
        ranges = [
            (38,10),(42,9),(44,8),(45,7),(47,6),(48,6),
            (50,5),(51,4),(52,4),(55,3),(58,2),(60,1),
        ]
        neg = [
            (125,-1),(128,-2),(130,-2),(132,-3),(135,-4),
            (139,-5),(141,-5),(144,-6),(147,-7),(150,-8),
            (153,-9),(156,-9),(158,-10),(161,-10),
            (163,-11),(166,-12),(168,-12),(171,-13),(178,-14),
        ]
    pos = "0"
    for mx, adj in reversed(ranges):
        pos = f"IF({x_cell}<={mx},{adj},{pos})"
    ng = str(neg[-1][1])
    for mx, adj in reversed(neg):
        ng = f"IF({x_cell}<={mx},{adj},{ng})"
    return (f'=IF(OR({x_cell}<34,{x_cell}>178),0,'
            f'IF({x_cell}<=60,{pos},IF({x_cell}<=123,0,{ng})))')



def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "MCMI-III"
    ws.freeze_panes = 'C5'

    # Column widths
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 4
    ws.column_dimensions['G'].width = 7
    ws.column_dimensions['H'].width = 10

    # === HEADER ===
    ws['A1'] = "MCMI-III AUTO-SCORING"
    ws['A1'].font = TITLE
    ws['D1'] = "Age:"
    ws['D1'].font = BOLD
    ws['E1'] = ""
    ws['E1'].fill = YFILL
    ws['E1'].border = BDR
    ws['F1'] = "Setting:"
    ws['F1'].font = BOLD
    ws['G1'] = "OPD"
    ws['G1'].fill = YFILL
    ws['G1'].border = BDR
    ws['A2'] = "Enter T or F in column B"
    ws['A2'].font = SMALL
    ws['D2'] = "Axis I wks:"
    ws['D2'].font = BOLD
    ws['E2'] = ""
    ws['E2'].fill = YFILL
    ws['E2'].border = BDR
    ws['F2'] = "(for IPD)"
    ws['F2'].font = SMALL

    # Setting validation
    dv_s = DataValidation(type="list", formula1='"OPD,IPD"')
    ws.add_data_validation(dv_s)
    dv_s.add(ws['G1'])

    # === ITEM ENTRY HEADERS (row 4) ===
    ws['A4'] = "Item"
    ws['B4'] = "Resp"
    ws['A4'].font = HDR
    ws['B4'].font = HDR
    ws['A4'].fill = HFILL
    ws['B4'].fill = HFILL
    ws['A4'].alignment = CTR
    ws['B4'].alignment = CTR

    # Also columns D-E for items 89-175 side by side
    ws['D4'] = "Item"
    ws['E4'] = "Resp"
    ws['D4'].font = HDR
    ws['E4'].font = HDR
    ws['D4'].fill = HFILL
    ws['E4'].fill = HFILL
    ws['D4'].alignment = CTR
    ws['E4'].alignment = CTR

    # Also columns G-H for overflow display
    ws['G4'] = "Item"
    ws['H4'] = "Resp"
    ws['G4'].font = HDR
    ws['H4'].font = HDR
    ws['G4'].fill = HFILL
    ws['H4'].fill = HFILL
    ws['G4'].alignment = CTR
    ws['H4'].alignment = CTR

    # T/F validation
    dv = DataValidation(type="list", formula1='"T,F"')
    dv.error = "Enter T or F"
    ws.add_data_validation(dv)

    # Items 1-60 in cols A-B (rows 5-64)
    for i in range(1, 61):
        r = i + 4
        ws.cell(row=r, column=1, value=i).alignment = CTR
        c = ws.cell(row=r, column=2)
        c.alignment = CTR
        c.fill = YFILL
        c.border = BDR
        dv.add(c)

    # Items 61-120 in cols D-E (rows 5-64)
    for i in range(61, 121):
        r = (i - 60) + 4
        ws.cell(row=r, column=4, value=i).alignment = CTR
        c = ws.cell(row=r, column=5)
        c.alignment = CTR
        c.fill = YFILL
        c.border = BDR
        dv.add(c)

    # Items 121-175 in cols G-H (rows 5-59)
    for i in range(121, 176):
        r = (i - 120) + 4
        ws.cell(row=r, column=7, value=i).alignment = CTR
        c = ws.cell(row=r, column=8)
        c.alignment = CTR
        c.fill = YFILL
        c.border = BDR
        dv.add(c)

    # IMPORTANT: All formulas reference column B rows 5-179
    # But items are split across 3 column pairs visually.
    # We need a HIDDEN column that consolidates all 175 responses.
    # Let's use column J (hidden) as the master response column.
    # Item N -> J(N+4)

    # Actually simpler: put ALL items in column B rows 5-179 (single column)
    # and use cols D-E, G-H just as a VISUAL MIRROR.
    # This way formulas only reference B5:B179.

    # Let me redo: items ALL go in column B (rows 5-179).
    # We show them in 3 visual groups but the actual data is ONE column.
    
    # CLEAR the visual split approach - use single column B for all 175
    # Remove the D-E and G-H entries
    for r in range(5, 65):
        ws.cell(row=r, column=4, value=None)
        ws.cell(row=r, column=5, value=None)
    for r in range(5, 60):
        ws.cell(row=r, column=7, value=None)
        ws.cell(row=r, column=8, value=None)
    
    # Clear headers for D-H
    for col in [4, 5, 7, 8]:
        ws.cell(row=4, column=col, value=None)
        ws.cell(row=4, column=col).fill = PatternFill()
        ws.cell(row=4, column=col).font = Font()

    # Put ALL 175 items in column A-B, rows 5-179
    for i in range(61, 176):
        r = i + 4
        ws.cell(row=r, column=1, value=i).alignment = CTR
        c = ws.cell(row=r, column=2)
        c.alignment = CTR
        c.fill = YFILL
        c.border = BDR
        dv.add(c)

    print("  175 items in column B (rows 5-179)")

    # === SCORING SECTION starts at row 182 ===
    SR = 182  # Scoring start row

    return wb, ws, SR


def add_scoring(ws, SR):
    """Add all scoring formulas starting at row SR."""
    row = SR

    # ========== VALIDITY ==========
    ws.cell(row=row, column=1, value="VALIDITY CHECKS").font = TITLE
    row += 1
    ws.cell(row=row, column=1, value="Check").font = BOLD
    ws.cell(row=row, column=2, value="Score").font = BOLD
    ws.cell(row=row, column=3, value="").font = BOLD
    ws.cell(row=row, column=4, value="Status").font = BOLD
    row += 1

    # Omits
    ws.cell(row=row, column=1, value="Omits")
    ws.cell(row=row, column=2, value="=175-COUNTA(B5:B179)")
    ws.cell(row=row, column=4, value=f'=IF(B{row}>11,"INVALID","OK")')
    omit_r = row
    row += 1

    # V
    v_parts = [f'IF({R(i)}="T",1,0)' for i in SCALE_V_ITEMS]
    ws.cell(row=row, column=1, value="V (Invalidity)")
    ws.cell(row=row, column=2, value="=" + "+".join(v_parts))
    ws.cell(row=row, column=4, 
            value=f'=IF(B{row}>=2,"INVALID",IF(B{row}=1,"CAUTION","OK"))')
    v_r = row
    row += 1

    # W - BOTH items TRUE = 1 point
    w_parts = []
    for a, b in SCALE_W_PAIRS:
        w_parts.append(f'IF(AND({R(a)}="T",{R(b)}="T"),1,0)')
    ws.cell(row=row, column=1, value="W (Inconsistency)")
    ws.cell(row=row, column=2, value="=" + "+".join(w_parts))
    ws.cell(row=row, column=4, 
            value=f'=IF(B{row}>=10,"INVALID",IF(B{row}>=8,"CAUTION","OK"))')
    w_r = row
    row += 1

    # Overall
    ws.cell(row=row, column=1, value="OVERALL").font = BOLD
    ws.cell(row=row, column=2, 
            value=f'=IF(OR(D{omit_r}="INVALID",D{v_r}="INVALID",D{w_r}="INVALID"),'
                  f'"INVALID",IF(OR(D{v_r}="CAUTION",D{w_r}="CAUTION"),"QUESTIONABLE","VALID"))')
    ws.cell(row=row, column=2).font = Font(bold=True, size=13, color="C00000")
    row += 2

    # ========== SCALE X ==========
    ws.cell(row=row, column=1, value="SCALE X (Disclosure)").font = BOLD
    row += 1
    
    # First compute raw scores for personality scales we need for X
    # We'll put raw score formulas in column D for temp use
    # Actually let's just build X formula directly

    # We need raw scores for 1,2A,2B,3,4,5,6A,6B,7,8A,8B first
    # Let's put the scale table below and reference it
    # Better: compute X inline using the item formulas directly
    
    # For X, we need to sum the raw scores of each personality scale
    # Let's compute each personality scale raw in a helper area first
    
    ws.cell(row=row, column=1, value="Scale")
    ws.cell(row=row, column=2, value="Raw")
    ws.cell(row=row, column=1).font = BOLD
    ws.cell(row=row, column=2).font = BOLD
    row += 1
    
    # Personality scale raws (for X calculation AND for scoring)
    pers_raw_rows = {}
    personality_order = ['1','2A','2B','3','4','5','6A','6B','7','8A','8B']
    for sid in personality_order:
        ws.cell(row=row, column=1, value=sid)
        ws.cell(row=row, column=2, value=raw_formula(SCALE_ITEMS[sid]))
        pers_raw_rows[sid] = row
        row += 1
    
    # X = sum of personality scales (Scale 5 * 2/3)
    x_terms = [f"B{pers_raw_rows[s]}" for s in personality_order if s != '5']
    x_terms.append(f"B{pers_raw_rows['5']}*2/3")
    ws.cell(row=row, column=1, value="X Raw").font = Font(bold=True, color="C00000")
    ws.cell(row=row, column=2, value=f"=ROUND({'+'.join(x_terms)},0)")
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    ws.cell(row=row, column=4, 
            value=f'=IF(OR(B{row}<34,B{row}>178),"INVALID","VALID")')
    x_r = row
    row += 2

    # Disclosure adjustments
    ws.cell(row=row, column=1, value="Disc Adj 1-8B:").font = BOLD
    ws.cell(row=row, column=2, value=disc_formula(f"B{x_r}", '18B'))
    adj18_r = row
    row += 1
    ws.cell(row=row, column=1, value="Disc Adj S-PP:").font = BOLD
    ws.cell(row=row, column=2, value=disc_formula(f"B{x_r}", 'SPP'))
    adjsp_r = row
    row += 1
    
    # Inpatient adjustments
    ws.cell(row=row, column=1, value="Inp SS:").font = BOLD
    ws.cell(row=row, column=2, 
            value='=IF(G1="IPD",IF(E2="",0,IF(E2<1,6,IF(E2<=4,4,0))),0)')
    inp_ss = row
    row += 1
    ws.cell(row=row, column=1, value="Inp CC:").font = BOLD
    ws.cell(row=row, column=2, 
            value='=IF(G1="IPD",IF(E2="",0,IF(E2<1,10,IF(E2<=4,8,0))),0)')
    inp_cc = row
    row += 1
    ws.cell(row=row, column=1, value="Inp PP:").font = BOLD
    ws.cell(row=row, column=2, 
            value='=IF(G1="IPD",IF(E2="",0,IF(E2<1,4,IF(E2<=4,2,0))),0)')
    inp_pp = row
    row += 2

    print("  Validity + X + Adjustments done")

    # ========== FULL SCORING TABLE ==========
    ws.cell(row=row, column=1, value="COMPLETE SCORES").font = TITLE
    row += 1
    # Headers
    hdrs = ['Scale', 'Raw', 'BR', 'Adj', 'Final BR', 'Interpretation']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HDR
        cell.fill = HFILL
        cell.alignment = CTR
        cell.border = BDR
    row += 1
    first_score_row = row

    # All scales
    all_scales = [
        ('Y','Desirability'), ('Z','Debasement'),
        ('1','Schizoid'), ('2A','Avoidant'), ('2B','Depressive'),
        ('3','Dependent'), ('4','Histrionic'), ('5','Narcissistic'),
        ('6A','Antisocial'), ('6B','Aggressive/Sadistic'),
        ('7','Compulsive'), ('8A','Negativistic'), ('8B','Masochistic'),
        ('S','Schizotypal'), ('C','Borderline'), ('P','Paranoid'),
        ('A','Anxiety'), ('H','Somatoform'), ('N','Bipolar:Manic'),
        ('D','Dysthymia'), ('B','Alcohol Dep'), ('T','Drug Dep'),
        ('R','PTSD'),
        ('SS','Thought Disorder'), ('CC','Major Depression'), ('PP','Delusional'),
    ]

    scale_rows = {}
    for sid, sname in all_scales:
        ws.cell(row=row, column=1, value=f"{sid} {sname}").font = BOLD
        
        # Raw score
        if sid in pers_raw_rows:
            # Already computed above, reference it
            ws.cell(row=row, column=2, value=f"=B{pers_raw_rows[sid]}")
        elif sid in SCALE_ITEMS:
            ws.cell(row=row, column=2, value=raw_formula(SCALE_ITEMS[sid]))
        ws.cell(row=row, column=2).alignment = CTR
        
        # BR (from raw)
        ws.cell(row=row, column=3, value=br_formula(sid, f"B{row}"))
        ws.cell(row=row, column=3).alignment = CTR
        
        # Adjustment
        if sid in CLINICAL_PERSONALITY:
            ws.cell(row=row, column=4, value=f"=B{adj18_r}")
        elif sid in SEVERE_PERSONALITY + CLINICAL_SYNDROMES + SEVERE_SYNDROMES:
            ws.cell(row=row, column=4, value=f"=B{adjsp_r}")
        else:
            ws.cell(row=row, column=4, value=0)
        ws.cell(row=row, column=4).alignment = CTR
        
        # Final BR = BR + Adj (+ inpatient for SS/CC/PP)
        if sid == 'SS':
            ws.cell(row=row, column=5, value=f"=MIN(115,MAX(0,C{row}+D{row}+B{inp_ss}))")
        elif sid == 'CC':
            ws.cell(row=row, column=5, value=f"=MIN(115,MAX(0,C{row}+D{row}+B{inp_cc}))")
        elif sid == 'PP':
            ws.cell(row=row, column=5, value=f"=MIN(115,MAX(0,C{row}+D{row}+B{inp_pp}))")
        else:
            ws.cell(row=row, column=5, value=f"=MIN(115,MAX(0,C{row}+D{row}))")
        ws.cell(row=row, column=5).alignment = CTR
        
        # Interpretation
        ws.cell(row=row, column=6, 
                value=f'=IF(E{row}>=85,"PROMINENT ***",'
                      f'IF(E{row}>=75,"SIGNIFICANT **",'
                      f'IF(E{row}>=60,"Trait *","Not Sig")))')
        
        # Borders
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BDR
        
        scale_rows[sid] = row
        row += 1

    last_score_row = row - 1
    print("  All 26 scale scores done")

    # Conditional formatting on Final BR (column E)
    rng = f"E{first_score_row}:E{last_score_row}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='greaterThanOrEqual', formula=['85'],
        fill=PatternFill("solid", fgColor="FF0000"),
        font=Font(bold=True, color="FFFFFF")))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='between', formula=['75','84'],
        fill=PatternFill("solid", fgColor="FF8C00"),
        font=Font(bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='between', formula=['60','74'],
        fill=PatternFill("solid", fgColor="FFFF00")))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='lessThan', formula=['60'],
        fill=PatternFill("solid", fgColor="C6EFCE")))

    print("  Color formatting applied")
    return scale_rows



if __name__ == '__main__':
    print("=" * 55)
    print("  BUILDING MCMI-III SINGLE-SHEET SCORING TOOL")
    print("=" * 55)
    
    wb, ws, SR = build()
    add_scoring(ws, SR)
    
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "MCMI_III_SingleSheet.xlsx")
    wb.save(out)
    
    print(f"\n  SAVED: {out}")
    print(f"\n  HOW TO USE:")
    print(f"  1. Open MCMI_III_SingleSheet.xlsx")
    print(f"  2. Enter age (E1), setting OPD/IPD (G1)")
    print(f"  3. If IPD: enter Axis I weeks in E2")
    print(f"  4. Type T or F for items 1-175 in column B (rows 5-179)")
    print(f"  5. Scroll down to row 182 - ALL SCORES ARE THERE!")
    print(f"  6. Red=BR>=85, Orange=75-84, Yellow=60-74, Green<60")
    print("=" * 55)
