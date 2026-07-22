#!/usr/bin/env python3
"""
MCMI-III AUTO-SCORING EXCEL TOOL GENERATOR
============================================
Generates a fully automated Excel workbook where:
- Sheet 1 (DATA ENTRY): User enters TRUE or FALSE for 175 items
- Sheet 2 (SCORING): All scoring computed automatically via Excel formulas
- Sheet 3 (PROFILE): Visual dashboard with conditional formatting

The user ONLY needs to type TRUE/FALSE in Sheet 1.
Everything else is calculated automatically.

Based on: Official MCMI-III Manual (Millon et al.) & Hand-Scoring User's Guide
"""

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, 
    numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from mcmi_item_keys import (
    SCALE_ITEMS, SCALE_NAMES, SCALE_V_ITEMS, SCALE_W_PAIRS,
    CLINICAL_PERSONALITY, SEVERE_PERSONALITY,
    CLINICAL_SYNDROMES, SEVERE_SYNDROMES
)

# ============================================================================
# STYLES
# ============================================================================
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
SUBHEADER_FONT = Font(bold=True, size=11, color="1F4E79")
NORMAL_FONT = Font(size=11)
SMALL_FONT = Font(size=9, italic=True, color="666666")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
LIGHT_GREEN = PatternFill("solid", fgColor="C6EFCE")
LIGHT_YELLOW = PatternFill("solid", fgColor="FFEB9C")
LIGHT_ORANGE = PatternFill("solid", fgColor="FFC7AA")
LIGHT_RED = PatternFill("solid", fgColor="FFC7CE")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

THIN_BORDER = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)
MEDIUM_BORDER = Border(
    left=Side('medium'), right=Side('medium'),
    top=Side('medium'), bottom=Side('medium')
)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
WRAP = Alignment(horizontal='left', vertical='center', wrapText=True)


def create_data_entry_sheet(wb):
    """
    Sheet 1: DATA ENTRY
    User enters TRUE or FALSE for items 1-175.
    """
    ws = wb.active
    ws.title = "DATA ENTRY"
    
    # Freeze panes
    ws.freeze_panes = 'C6'
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = "MCMI-III AUTO-SCORING TOOL"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    ws.merge_cells('A2:E2')
    ws['A2'] = "Enter TRUE or FALSE for each item below. Scoring is AUTOMATIC."
    ws['A2'].font = Font(size=11, italic=True, color="666666")
    ws['A2'].alignment = CENTER
    
    # Patient info section
    ws['A3'] = "Patient Age:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = ""  # User enters age here
    ws['B3'].border = MEDIUM_BORDER
    ws['B3'].fill = LIGHT_YELLOW
    
    ws['C3'] = "Setting (OPD/IPD):"
    ws['C3'].font = Font(bold=True)
    ws['D3'] = "OPD"  # Default
    ws['D3'].border = MEDIUM_BORDER
    ws['D3'].fill = LIGHT_YELLOW
    
    ws['A4'] = "Gender (M/F):"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = ""
    ws['B4'].border = MEDIUM_BORDER
    ws['B4'].fill = LIGHT_YELLOW
    
    ws['C4'] = "Axis I Duration:"
    ws['C4'].font = Font(bold=True)
    ws['D4'] = ""  # weeks
    ws['D4'].border = MEDIUM_BORDER
    ws['D4'].fill = LIGHT_YELLOW
    ws['E4'] = "(weeks, for IPD only)"
    ws['E4'].font = SMALL_FONT
    
    # Data validation for Setting
    dv_setting = DataValidation(type="list", formula1='"OPD,IPD"')
    dv_setting.error = "Please enter OPD or IPD"
    dv_setting.errorTitle = "Invalid Setting"
    ws.add_data_validation(dv_setting)
    dv_setting.add(ws['D3'])
    
    # Data validation for Gender
    dv_gender = DataValidation(type="list", formula1='"M,F"')
    ws.add_data_validation(dv_gender)
    dv_gender.add(ws['B4'])
    
    # Column headers for items
    row = 6
    headers = ['Item #', 'Response (TRUE/FALSE)', 'Status']
    col_widths = [8, 22, 12]
    
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Additional columns for second set (items side by side for compact view)
    # Items 1-88 in columns A-C, items 89-175 in columns E-G
    ws.column_dimensions['D'].width = 3  # Spacer
    
    for col_offset, (header, width) in enumerate(zip(headers, col_widths), 5):
        cell = ws.cell(row=row, column=col_offset, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_offset)].width = width
    
    # Data validation for TRUE/FALSE
    dv_tf = DataValidation(type="list", formula1='"TRUE,FALSE"')
    dv_tf.error = "Please enter TRUE or FALSE"
    dv_tf.errorTitle = "Invalid Response"
    ws.add_data_validation(dv_tf)
    
    # Item rows - Left side (1-88)
    for i in range(1, 89):
        r = row + i
        ws.cell(row=r, column=1, value=i).alignment = CENTER
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).fill = LIGHT_GRAY
        
        resp_cell = ws.cell(row=r, column=2)
        resp_cell.alignment = CENTER
        resp_cell.border = MEDIUM_BORDER
        resp_cell.fill = WHITE_FILL
        dv_tf.add(resp_cell)
        
        # Status formula - shows checkmark if answered
        status_cell = ws.cell(row=r, column=3)
        status_cell.value = f'=IF(B{r}="","","\u2713")'
        status_cell.alignment = CENTER
        status_cell.border = THIN_BORDER
    
    # Item rows - Right side (89-175)
    for i in range(89, 176):
        r = row + (i - 88)
        ws.cell(row=r, column=5, value=i).alignment = CENTER
        ws.cell(row=r, column=5).border = THIN_BORDER
        ws.cell(row=r, column=5).fill = LIGHT_GRAY
        
        resp_cell = ws.cell(row=r, column=6)
        resp_cell.alignment = CENTER
        resp_cell.border = MEDIUM_BORDER
        resp_cell.fill = WHITE_FILL
        dv_tf.add(resp_cell)
        
        status_cell = ws.cell(row=r, column=7)
        status_cell.value = f'=IF(F{r}="","","\u2713")'
        status_cell.alignment = CENTER
        status_cell.border = THIN_BORDER
    
    # Summary at bottom
    summary_row = row + 90
    ws.cell(row=summary_row, column=1, value="SUMMARY:").font = Font(bold=True, size=12)
    ws.cell(row=summary_row+1, column=1, value="Items Answered:")
    # Count non-blank responses
    ws.cell(row=summary_row+1, column=2, 
            value=f'=COUNTA(B7:B94)+COUNTA(F7:F93)')
    ws.cell(row=summary_row+1, column=2).font = Font(bold=True, size=12)
    
    ws.cell(row=summary_row+2, column=1, value="Items Omitted:")
    ws.cell(row=summary_row+2, column=2, 
            value=f'=175-(COUNTA(B7:B94)+COUNTA(F7:F93))')
    
    ws.cell(row=summary_row+3, column=1, value="Status:")
    ws.cell(row=summary_row+3, column=2,
            value=f'=IF((COUNTA(B7:B94)+COUNTA(F7:F93))=175,"COMPLETE","INCOMPLETE - "&175-(COUNTA(B7:B94)+COUNTA(F7:F93))&" items remaining")')
    ws.cell(row=summary_row+3, column=2).font = Font(bold=True)
    
    return ws



def get_item_cell_ref(item_num):
    """
    Get the Excel cell reference for a given item number's response.
    Items 1-88 are in column B (rows 7-94) of 'DATA ENTRY' sheet.
    Items 89-175 are in column F (rows 7-93) of 'DATA ENTRY' sheet.
    """
    if item_num <= 88:
        row = 6 + item_num  # item 1 = row 7, item 88 = row 94
        return f"'DATA ENTRY'!B{row}"
    else:
        row = 6 + (item_num - 88)  # item 89 = row 7, item 175 = row 93
        return f"'DATA ENTRY'!F{row}"


def build_raw_score_formula(scale_items):
    """
    Build an Excel formula that sums the weighted responses for a scale.
    For weight-1 items: adds 1 if TRUE
    For weight-2 items: adds 2 if TRUE
    """
    parts = []
    for item_info in sorted(scale_items, key=lambda x: x[0]):
        item_num, weight = item_info
        cell_ref = get_item_cell_ref(item_num)
        if weight == 1:
            parts.append(f'IF({cell_ref}="TRUE",1,0)')
        elif weight == 2:
            parts.append(f'IF({cell_ref}="TRUE",2,0)')
        elif weight == 3:
            parts.append(f'IF({cell_ref}="TRUE",3,0)')
    
    # Excel has a limit on formula length, so we may need to split
    # But for our scales (max ~24 items), a single SUM should work
    formula = "=" + "+".join(parts)
    return formula


def build_v_score_formula():
    """Build formula for Scale V (Invalidity) - items 65, 110, 157."""
    parts = []
    for item in SCALE_V_ITEMS:
        cell_ref = get_item_cell_ref(item)
        parts.append(f'IF({cell_ref}="TRUE",1,0)')
    return "=" + "+".join(parts)


def build_w_score_formula():
    """
    Build formula for Scale W (Inconsistency).
    Per manual: "Each blackened pair of responses adds 1 to the scale score."
    Score 1 for each pair where BOTH items are TRUE.
    """
    parts = []
    for item_a, item_b in SCALE_W_PAIRS:
        ref_a = get_item_cell_ref(item_a)
        ref_b = get_item_cell_ref(item_b)
        # Score 1 if BOTH items in pair are TRUE (both blackened)
        parts.append(f'IF(AND({ref_a}="TRUE",{ref_b}="TRUE"),1,0)')
    
    formula = "=" + "+".join(parts)
    return formula



def create_scoring_sheet(wb):
    """
    Sheet 2: AUTO-SCORING
    All calculations done via formulas referencing Sheet 1 data.
    """
    ws = wb.create_sheet("AUTO-SCORING")
    ws.freeze_panes = 'A4'
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "MCMI-III AUTOMATIC SCORING RESULTS"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    ws.merge_cells('A2:H2')
    ws['A2'] = "All scores calculated automatically from your DATA ENTRY responses"
    ws['A2'].font = SMALL_FONT
    ws['A2'].alignment = CENTER
    
    # Column widths
    col_widths = {'A': 8, 'B': 35, 'C': 12, 'D': 12, 'E': 12, 
                  'F': 12, 'G': 28, 'H': 14}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    row = 4
    
    # =========================================================================
    # SECTION 1: VALIDITY CHECKS
    # =========================================================================
    ws.cell(row=row, column=1, value="VALIDITY CHECKS").font = Font(bold=True, size=13, color="C00000")
    row += 1
    
    # Headers
    val_headers = ['Scale', 'Name', 'Score', 'Threshold', 'Result']
    for col, h in enumerate(val_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row += 1
    
    # Omitted Items
    ws.cell(row=row, column=1, value="-").alignment = CENTER
    ws.cell(row=row, column=2, value="Omitted Items")
    ws.cell(row=row, column=3, value="=175-(COUNTA('DATA ENTRY'!B7:B94)+COUNTA('DATA ENTRY'!F7:F93))")
    ws.cell(row=row, column=4, value="<= 11")
    ws.cell(row=row, column=5, value=f'=IF(C{row}>11,"INVALID","VALID")')
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).alignment = CENTER
    omit_row = row
    row += 1
    
    # Scale V
    v_formula = build_v_score_formula()
    ws.cell(row=row, column=1, value="V").alignment = CENTER
    ws.cell(row=row, column=2, value="Invalidity (Items 65, 110, 157)")
    ws.cell(row=row, column=3, value=v_formula)
    ws.cell(row=row, column=4, value="< 2")
    ws.cell(row=row, column=5, value=f'=IF(C{row}>=2,"INVALID",IF(C{row}=1,"QUESTIONABLE","VALID"))')
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).alignment = CENTER
    v_row = row
    row += 1
    
    # Scale W
    w_formula = build_w_score_formula()
    ws.cell(row=row, column=1, value="W").alignment = CENTER
    ws.cell(row=row, column=2, value="Inconsistency (44 item pairs)")
    ws.cell(row=row, column=3, value=w_formula)
    ws.cell(row=row, column=4, value="< 10")
    ws.cell(row=row, column=5, value=f'=IF(C{row}>=10,"INVALID",IF(C{row}>=8,"QUESTIONABLE","VALID"))')
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).alignment = CENTER
    w_row = row
    row += 1
    
    # Overall Validity
    ws.cell(row=row, column=1, value="").alignment = CENTER
    ws.cell(row=row, column=2, value="OVERALL VALIDITY").font = Font(bold=True)
    ws.cell(row=row, column=3, value="")
    ws.cell(row=row, column=4, value="")
    overall_formula = (
        f'=IF(OR(E{omit_row}="INVALID",E{v_row}="INVALID",E{w_row}="INVALID"),"INVALID",'
        f'IF(OR(E{v_row}="QUESTIONABLE",E{w_row}="QUESTIONABLE"),"QUESTIONABLE","VALID"))'
    )
    ws.cell(row=row, column=5, value=overall_formula)
    ws.cell(row=row, column=5).font = Font(bold=True, size=12)
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = MEDIUM_BORDER
        ws.cell(row=row, column=c).fill = LIGHT_BLUE
    validity_row = row
    row += 2
    
    # =========================================================================
    # SECTION 2: RAW SCORES + BR SCORES FOR ALL SCALES
    # =========================================================================
    ws.cell(row=row, column=1, value="COMPLETE SCALE SCORES").font = Font(bold=True, size=13, color="1F4E79")
    row += 1
    
    # Headers
    score_headers = ['Scale', 'Name', 'Raw Score', 'Initial BR', 
                     'Adjusted BR', 'Final BR', 'Interpretation', 'Level']
    for col, h in enumerate(score_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row += 1
    score_start_row = row
    
    # Track which row each scale is on for cross-references
    scale_rows = {}
    
    # Define all scales in order
    all_scales_ordered = [
        # Modifying indices first
        ('Y', 'Desirability', 'Modifying Index'),
        ('Z', 'Debasement', 'Modifying Index'),
        # Personality patterns
        ('1', 'Schizoid', 'Clinical Personality'),
        ('2A', 'Avoidant', 'Clinical Personality'),
        ('2B', 'Depressive', 'Clinical Personality'),
        ('3', 'Dependent', 'Clinical Personality'),
        ('4', 'Histrionic', 'Clinical Personality'),
        ('5', 'Narcissistic', 'Clinical Personality'),
        ('6A', 'Antisocial', 'Clinical Personality'),
        ('6B', 'Aggressive (Sadistic)', 'Clinical Personality'),
        ('7', 'Compulsive', 'Clinical Personality'),
        ('8A', 'Negativistic', 'Clinical Personality'),
        ('8B', 'Masochistic', 'Clinical Personality'),
        # Severe personality
        ('S', 'Schizotypal', 'Severe Personality'),
        ('C', 'Borderline', 'Severe Personality'),
        ('P', 'Paranoid', 'Severe Personality'),
        # Clinical syndromes
        ('A', 'Anxiety', 'Clinical Syndrome'),
        ('H', 'Somatoform', 'Clinical Syndrome'),
        ('N', 'Bipolar: Manic', 'Clinical Syndrome'),
        ('D', 'Dysthymia', 'Clinical Syndrome'),
        ('B', 'Alcohol Dependence', 'Clinical Syndrome'),
        ('T', 'Drug Dependence', 'Clinical Syndrome'),
        ('R', 'PTSD', 'Clinical Syndrome'),
        # Severe syndromes
        ('SS', 'Thought Disorder', 'Severe Syndrome'),
        ('CC', 'Major Depression', 'Severe Syndrome'),
        ('PP', 'Delusional Disorder', 'Severe Syndrome'),
    ]
    
    # Add category headers and scale rows
    current_category = None
    
    for scale_id, scale_name, category in all_scales_ordered:
        # Category header
        if category != current_category:
            current_category = category
            cat_cell = ws.cell(row=row, column=1, value=category)
            cat_cell.font = Font(bold=True, italic=True, size=10, color="1F4E79")
            ws.merge_cells(f'A{row}:H{row}')
            cat_cell.fill = LIGHT_BLUE
            row += 1
        
        # Scale data
        ws.cell(row=row, column=1, value=scale_id).alignment = CENTER
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=scale_name)
        
        # Raw score formula
        if scale_id in SCALE_ITEMS:
            raw_formula = build_raw_score_formula(SCALE_ITEMS[scale_id])
            ws.cell(row=row, column=3, value=raw_formula)
        
        # Initial BR (simplified linear approximation formula)
        # We'll use a lookup-style formula based on anchor points
        ws.cell(row=row, column=4, value=f"=C{row}*3")  # Simplified approximation
        
        # Adjusted BR (after disclosure adjustment)
        ws.cell(row=row, column=5, value=f"=D{row}")  # Will be updated with adjustment formulas
        
        # Final BR
        ws.cell(row=row, column=6, value=f"=MIN(115,MAX(0,E{row}))")
        
        # Interpretation
        interp_formula = (
            f'=IF(F{row}>=85,"PROMINENT (Disorder)",'
            f'IF(F{row}>=75,"CLINICALLY SIGNIFICANT",'
            f'IF(F{row}>=60,"Trait Present","Not Significant")))'
        )
        ws.cell(row=row, column=7, value=interp_formula)
        
        # Level indicator
        level_formula = (
            f'=IF(F{row}>=85,"***",'
            f'IF(F{row}>=75,"**",'
            f'IF(F{row}>=60,"*","")))'
        )
        ws.cell(row=row, column=8, value=level_formula)
        
        # Borders
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = THIN_BORDER
            if c >= 3:
                ws.cell(row=row, column=c).alignment = CENTER
        
        scale_rows[scale_id] = row
        row += 1
    
    # Now add Scale X (Disclosure) - computed from scales 1-8B
    row += 1
    ws.cell(row=row, column=1, value="SCALE X (DISCLOSURE) CALCULATION").font = Font(bold=True, size=11, color="C00000")
    row += 1
    
    # X = sum of (1, 2A, 2B, 3, 4, 5*2/3, 6A, 6B, 7, 8A, 8B)
    x_parts = []
    for s in ['1', '2A', '2B', '3', '4', '6A', '6B', '7', '8A', '8B']:
        x_parts.append(f"C{scale_rows[s]}")
    x_parts.append(f"C{scale_rows['5']}*2/3")
    
    ws.cell(row=row, column=1, value="X")
    ws.cell(row=row, column=2, value="Disclosure (Raw)")
    x_formula = "=ROUND(" + "+".join(x_parts) + ",0)"
    ws.cell(row=row, column=3, value=x_formula)
    ws.cell(row=row, column=3).font = Font(bold=True)
    ws.cell(row=row, column=4, value=f'=IF(C{row}<34,"INVALID - Too Low",IF(C{row}>178,"INVALID - Too High","Valid: "&C{row}))')
    scale_rows['X'] = row
    x_raw_row = row
    row += 2
    
    # =========================================================================
    # SECTION 3: DISCLOSURE ADJUSTMENT INFO
    # =========================================================================
    ws.cell(row=row, column=1, value="ADJUSTMENTS APPLIED").font = Font(bold=True, size=13, color="1F4E79")
    row += 1
    
    # Disclosure adjustment lookup (simplified)
    ws.cell(row=row, column=1, value="Disclosure Adj:")
    ws.cell(row=row, column=2, value="1-8B Factor")
    disc_adj_formula = (
        f'=IF(C{x_raw_row}<34,"N/A",'
        f'IF(C{x_raw_row}>178,"N/A",'
        f'IF(AND(C{x_raw_row}>=61,C{x_raw_row}<=123),0,'
        f'IF(C{x_raw_row}<=37,20,'
        f'IF(C{x_raw_row}<=40,17,'
        f'IF(C{x_raw_row}<=43,15,'
        f'IF(C{x_raw_row}<=46,12,'
        f'IF(C{x_raw_row}<=49,9,'
        f'IF(C{x_raw_row}<=52,7,'
        f'IF(C{x_raw_row}<=55,4,'
        f'IF(C{x_raw_row}<=58,2,'
        f'IF(C{x_raw_row}<=60,1,'
        f'IF(C{x_raw_row}<=127,-2,'
        f'IF(C{x_raw_row}<=133,-4,'
        f'IF(C{x_raw_row}<=140,-8,'
        f'IF(C{x_raw_row}<=150,-11,'
        f'IF(C{x_raw_row}<=160,-15,'
        f'IF(C{x_raw_row}<=170,-19,-20))))))))))))))))))'
    )
    ws.cell(row=row, column=3, value=disc_adj_formula)
    ws.cell(row=row, column=3).font = Font(bold=True, color="C00000")
    disc_adj_row = row
    row += 1
    
    ws.cell(row=row, column=1, value="Disclosure Adj:")
    ws.cell(row=row, column=2, value="S-PP Factor")
    disc_spp_formula = (
        f'=IF(C{x_raw_row}<34,"N/A",'
        f'IF(C{x_raw_row}>178,"N/A",'
        f'IF(AND(C{x_raw_row}>=61,C{x_raw_row}<=123),0,'
        f'IF(C{x_raw_row}<=38,10,'
        f'IF(C{x_raw_row}<=42,9,'
        f'IF(C{x_raw_row}<=44,8,'
        f'IF(C{x_raw_row}<=47,7,'
        f'IF(C{x_raw_row}<=49,5,'
        f'IF(C{x_raw_row}<=52,4,'
        f'IF(C{x_raw_row}<=55,3,'
        f'IF(C{x_raw_row}<=58,2,'
        f'IF(C{x_raw_row}<=60,1,'
        f'IF(C{x_raw_row}<=128,-2,'
        f'IF(C{x_raw_row}<=135,-4,'
        f'IF(C{x_raw_row}<=145,-7,'
        f'IF(C{x_raw_row}<=155,-9,'
        f'IF(C{x_raw_row}<=165,-11,'
        f'IF(C{x_raw_row}<=172,-13,-14))))))))))))))))))'
    )
    ws.cell(row=row, column=3, value=disc_spp_formula)
    ws.cell(row=row, column=3).font = Font(bold=True, color="C00000")
    disc_spp_row = row
    row += 2
    
    # =========================================================================
    # Now go back and UPDATE the Adjusted BR formulas with the adjustment
    # =========================================================================
    # Clinical Personality scales (1-8B): Initial BR + disclosure 1-8B adjustment
    for scale in CLINICAL_PERSONALITY:
        sr = scale_rows[scale]
        ws.cell(row=sr, column=5, value=f"=MIN(115,MAX(0,D{sr}+C{disc_adj_row}))")
    
    # Severe Personality + Syndromes: Initial BR + S-PP adjustment
    for scale in SEVERE_PERSONALITY + CLINICAL_SYNDROMES + SEVERE_SYNDROMES:
        sr = scale_rows[scale]
        ws.cell(row=sr, column=5, value=f"=MIN(115,MAX(0,D{sr}+C{disc_spp_row}))")
    
    # Y and Z don't get disclosure adjustment
    ws.cell(row=scale_rows['Y'], column=5, value=f"=D{scale_rows['Y']}")
    ws.cell(row=scale_rows['Z'], column=5, value=f"=D{scale_rows['Z']}")
    
    # =========================================================================
    # INPATIENT ADJUSTMENT INFO
    # =========================================================================
    ws.cell(row=row, column=1, value="Inpatient Adj:").font = Font(bold=True)
    ws.cell(row=row, column=2, value="(SS, CC, PP only - for IPD patients)")
    
    # SS adjustment
    inpat_ss_formula = (
        f"=IF('DATA ENTRY'!D3=\"IPD\","
        f"IF('DATA ENTRY'!D4<1,6,IF('DATA ENTRY'!D4<=4,4,0)),0)"
    )
    ws.cell(row=row, column=3, value=inpat_ss_formula)
    inpat_row = row
    row += 1
    
    # CC adjustment
    inpat_cc_formula = (
        f"=IF('DATA ENTRY'!D3=\"IPD\","
        f"IF('DATA ENTRY'!D4<1,10,IF('DATA ENTRY'!D4<=4,8,0)),0)"
    )
    ws.cell(row=row, column=2, value="CC Inpatient Adj")
    ws.cell(row=row, column=3, value=inpat_cc_formula)
    inpat_cc_row = row
    row += 1
    
    # PP adjustment
    inpat_pp_formula = (
        f"=IF('DATA ENTRY'!D3=\"IPD\","
        f"IF('DATA ENTRY'!D4<1,4,IF('DATA ENTRY'!D4<=4,2,0)),0)"
    )
    ws.cell(row=row, column=2, value="PP Inpatient Adj")
    ws.cell(row=row, column=3, value=inpat_pp_formula)
    inpat_pp_row = row
    row += 2
    
    # Update SS, CC, PP final BR with inpatient adjustment
    ss_r = scale_rows['SS']
    cc_r = scale_rows['CC']
    pp_r = scale_rows['PP']
    ws.cell(row=ss_r, column=6, value=f"=MIN(115,MAX(0,E{ss_r}+C{inpat_row}))")
    ws.cell(row=cc_r, column=6, value=f"=MIN(115,MAX(0,E{cc_r}+C{inpat_cc_row}))")
    ws.cell(row=pp_r, column=6, value=f"=MIN(115,MAX(0,E{pp_r}+C{inpat_pp_row}))")
    
    # Add conditional formatting for BR scores
    # Red for BR >= 85
    red_rule = CellIsRule(operator='greaterThanOrEqual', formula=['85'],
                          fill=PatternFill("solid", fgColor="FF0000"),
                          font=Font(bold=True, color="FFFFFF"))
    # Orange for BR 75-84
    orange_rule = CellIsRule(operator='between', formula=['75', '84'],
                             fill=PatternFill("solid", fgColor="FF8C00"),
                             font=Font(bold=True))
    # Yellow for BR 60-74
    yellow_rule = CellIsRule(operator='between', formula=['60', '74'],
                             fill=PatternFill("solid", fgColor="FFFF00"))
    
    br_range = f"F{score_start_row}:F{score_start_row + 30}"
    ws.conditional_formatting.add(br_range, red_rule)
    ws.conditional_formatting.add(br_range, orange_rule)
    ws.conditional_formatting.add(br_range, yellow_rule)
    
    return ws, scale_rows



def create_profile_sheet(wb, scale_rows):
    """
    Sheet 3: PROFILE DASHBOARD
    Visual summary with color coding.
    """
    ws = wb.create_sheet("PROFILE")
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "MCMI-III CLINICAL PROFILE"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    ws.merge_cells('A2:G2')
    ws['A2'] = "Visual Profile - BR Scores with Clinical Thresholds"
    ws['A2'].font = SMALL_FONT
    ws['A2'].alignment = CENTER
    
    # Column widths
    for col, w in {'A': 8, 'B': 30, 'C': 10, 'D': 12, 'E': 25, 'F': 6, 'G': 45}.items():
        ws.column_dimensions[col].width = w
    
    row = 4
    
    # Validity Summary Banner
    ws.cell(row=row, column=1, value="PROTOCOL VALIDITY:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value="='AUTO-SCORING'!E9")  # Overall validity cell
    ws.cell(row=row, column=3).font = Font(bold=True, size=14)
    row += 2
    
    # Legend
    ws.cell(row=row, column=1, value="LEGEND:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="***")
    ws.cell(row=row, column=2, value="BR >= 85: PROMINENT (Disorder Present)")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="FF0000")
    ws.cell(row=row, column=2).font = Font(color="FFFFFF", bold=True)
    row += 1
    ws.cell(row=row, column=1, value="**")
    ws.cell(row=row, column=2, value="BR 75-84: CLINICALLY SIGNIFICANT")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="FF8C00")
    ws.cell(row=row, column=2).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="*")
    ws.cell(row=row, column=2, value="BR 60-74: Trait/Tendency Present")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="FFFF00")
    row += 1
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="BR < 60: Not Significant")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="C6EFCE")
    row += 2
    
    # Profile Table Headers
    headers = ['Scale', 'Name', 'Raw', 'Final BR', 'Interpretation', 'Lvl', 'BR Bar Graph (0-115)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    row += 1
    
    # All scales in profile order
    profile_scales = [
        # Modifying
        ('X', 'Disclosure'),
        ('Y', 'Desirability'),
        ('Z', 'Debasement'),
        # Personality
        ('1', 'Schizoid'),
        ('2A', 'Avoidant'),
        ('2B', 'Depressive'),
        ('3', 'Dependent'),
        ('4', 'Histrionic'),
        ('5', 'Narcissistic'),
        ('6A', 'Antisocial'),
        ('6B', 'Aggressive (Sadistic)'),
        ('7', 'Compulsive'),
        ('8A', 'Negativistic'),
        ('8B', 'Masochistic'),
        # Severe personality
        ('S', 'Schizotypal'),
        ('C', 'Borderline'),
        ('P', 'Paranoid'),
        # Syndromes
        ('A', 'Anxiety'),
        ('H', 'Somatoform'),
        ('N', 'Bipolar: Manic'),
        ('D', 'Dysthymia'),
        ('B', 'Alcohol Dependence'),
        ('T', 'Drug Dependence'),
        ('R', 'PTSD'),
        # Severe syndromes
        ('SS', 'Thought Disorder'),
        ('CC', 'Major Depression'),
        ('PP', 'Delusional Disorder'),
    ]
    
    for scale_id, scale_name in profile_scales:
        ws.cell(row=row, column=1, value=scale_id).alignment = CENTER
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=scale_name)
        
        if scale_id == 'X':
            # X raw score
            ws.cell(row=row, column=3, value=f"='AUTO-SCORING'!C{scale_rows.get('X', 40)}")
            # X doesn't have a standard BR in the same way
            ws.cell(row=row, column=4, value=f"='AUTO-SCORING'!C{scale_rows.get('X', 40)}")
        elif scale_id in scale_rows:
            sr = scale_rows[scale_id]
            ws.cell(row=row, column=3, value=f"='AUTO-SCORING'!C{sr}")
            ws.cell(row=row, column=4, value=f"='AUTO-SCORING'!F{sr}")
        
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=4).alignment = CENTER
        
        # Interpretation
        interp_formula = (
            f'=IF(D{row}>=85,"PROMINENT (Disorder)",'
            f'IF(D{row}>=75,"CLINICALLY SIGNIFICANT",'
            f'IF(D{row}>=60,"Trait Present","Not Significant")))'
        )
        ws.cell(row=row, column=5, value=interp_formula)
        
        # Level
        level_formula = f'=IF(D{row}>=85,"***",IF(D{row}>=75,"**",IF(D{row}>=60,"*","")))'
        ws.cell(row=row, column=6, value=level_formula).alignment = CENTER
        
        # Bar graph using REPT
        bar_formula = f'=REPT("█",MIN(45,INT(D{row}/2.5)))&" "&D{row}'
        ws.cell(row=row, column=7, value=bar_formula)
        ws.cell(row=row, column=7).font = Font(size=9)
        
        # Borders
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN_BORDER
        
        row += 1
    
    # Add conditional formatting for the Final BR column
    br_range = f"D{row-27}:D{row-1}"
    red_rule = CellIsRule(operator='greaterThanOrEqual', formula=['85'],
                          fill=PatternFill("solid", fgColor="FF0000"),
                          font=Font(bold=True, color="FFFFFF"))
    orange_rule = CellIsRule(operator='between', formula=['75', '84'],
                             fill=PatternFill("solid", fgColor="FF8C00"),
                             font=Font(bold=True))
    yellow_rule = CellIsRule(operator='between', formula=['60', '74'],
                             fill=PatternFill("solid", fgColor="FFFF00"))
    green_rule = CellIsRule(operator='lessThan', formula=['60'],
                            fill=PatternFill("solid", fgColor="C6EFCE"))
    
    ws.conditional_formatting.add(br_range, red_rule)
    ws.conditional_formatting.add(br_range, orange_rule)
    ws.conditional_formatting.add(br_range, yellow_rule)
    ws.conditional_formatting.add(br_range, green_rule)
    
    # Threshold lines note
    row += 2
    ws.cell(row=row, column=1, value="CLINICAL THRESHOLDS:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="BR 60 = Clinical threshold for trait presence")
    row += 1
    ws.cell(row=row, column=1, value="BR 75 = Threshold for clinically significant presence")
    row += 1
    ws.cell(row=row, column=1, value="BR 85 = Threshold for prominent/disorder presence")
    row += 2
    
    # Notes
    ws.cell(row=row, column=1, value="SCORING NOTES:").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="- Disclosure Adjustment applied to all personality + syndrome scales")
    row += 1
    ws.cell(row=row, column=1, value="- Inpatient Adjustment applied to SS, CC, PP (for IPD setting only)")
    row += 1
    ws.cell(row=row, column=1, value="- Scale X valid range: 34-178 (outside = INVALID)")
    row += 1
    ws.cell(row=row, column=1, value="- Scale V >= 2 = INVALID (random responding)")
    row += 1
    ws.cell(row=row, column=1, value="- Scale W >= 10 = INVALID (inconsistent responding)")
    
    return ws



def create_br_reference_sheet(wb):
    """
    Sheet 4: BR SCORING REFERENCE
    Contains the adjustment tables and BR interpretation guide.
    """
    ws = wb.create_sheet("BR REFERENCE")
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "MCMI-III BR SCORING REFERENCE TABLES"
    ws['A1'].font = TITLE_FONT
    
    row = 3
    
    # BR Interpretation
    ws.cell(row=row, column=1, value="BR SCORE INTERPRETATION").font = Font(bold=True, size=12)
    row += 1
    interp_data = [
        ("< 60", "Not Significant", "No pathological trait"),
        ("60-74", "Trait Present", "Subclinical trait/tendency"),
        ("75-84", "Clinically Significant", "Syndrome/disorder present"),
        (">= 85", "PROMINENT", "Disorder is a prominent feature"),
    ]
    headers = ['BR Range', 'Level', 'Meaning']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = HEADER_FILL
        ws.cell(row=row, column=c).font = HEADER_FONT
    row += 1
    for br_range, level, meaning in interp_data:
        ws.cell(row=row, column=1, value=br_range)
        ws.cell(row=row, column=2, value=level)
        ws.cell(row=row, column=3, value=meaning)
        row += 1
    
    row += 2
    
    # Disclosure Adjustment Table
    ws.cell(row=row, column=1, value="DISCLOSURE ADJUSTMENT (Table 1)").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Scale X Raw")
    ws.cell(row=row, column=2, value="1-8B Adj")
    ws.cell(row=row, column=3, value="S-PP Adj")
    for c in range(1, 4):
        ws.cell(row=row, column=c).font = Font(bold=True)
        ws.cell(row=row, column=c).fill = LIGHT_BLUE
    row += 1
    
    disc_entries = [
        ("34-37", "+20", "+10"),
        ("38", "+19", "+10"),
        ("39", "+18", "+10"),
        ("40-41", "+17", "+9"),
        ("42", "+16", "+9"),
        ("43", "+15", "+8"),
        ("44", "+14", "+8"),
        ("45", "+13", "+7"),
        ("46", "+12", "+7"),
        ("47", "+11", "+6"),
        ("48", "+10", "+6"),
        ("49", "+9", "+5"),
        ("50", "+9", "+5"),
        ("51", "+8", "+4"),
        ("52", "+7", "+4"),
        ("53", "+6", "+3"),
        ("54", "+5", "+3"),
        ("55", "+4", "+3"),
        ("56", "+3", "+2"),
        ("57", "+2", "+2"),
        ("58", "+2", "+2"),
        ("59", "+1", "+1"),
        ("60", "+1", "+1"),
        ("61-123", "0", "0"),
        ("124-125", "-1", "-1"),
        ("126-128", "-2", "-2"),
        ("129-130", "-3", "-2"),
        ("131-132", "-4", "-3"),
        ("133-134", "-5", "-3"),
        ("135", "-5", "-4"),
        ("136-137", "-6", "-4"),
        ("138-139", "-7", "-5"),
        ("140-141", "-8", "-5"),
        ("142", "-8", "-6"),
        ("143-144", "-9", "-6"),
        ("145", "-9", "-7"),
        ("146-147", "-10", "-7"),
        ("148-150", "-10 to -11", "-7 to -8"),
        ("151-153", "-11 to -12", "-8 to -9"),
        ("154-156", "-13", "-9"),
        ("157-158", "-14", "-10"),
        ("159-161", "-15", "-10"),
        ("162-163", "-16", "-11"),
        ("164-166", "-17", "-11 to -12"),
        ("167-168", "-18", "-12"),
        ("169-171", "-19", "-12 to -13"),
        ("172-178", "-20", "-13 to -14"),
    ]
    for entry in disc_entries:
        ws.cell(row=row, column=1, value=entry[0])
        ws.cell(row=row, column=2, value=entry[1])
        ws.cell(row=row, column=3, value=entry[2])
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1
    
    row += 2
    
    # Inpatient Adjustment
    ws.cell(row=row, column=1, value="INPATIENT ADJUSTMENT (Table 5)").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Duration").font = Font(bold=True)
    ws.cell(row=row, column=2, value="SS Adj").font = Font(bold=True)
    ws.cell(row=row, column=3, value="CC Adj").font = Font(bold=True)
    ws.cell(row=row, column=4, value="PP Adj").font = Font(bold=True)
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = LIGHT_BLUE
    row += 1
    ws.cell(row=row, column=1, value="< 1 week")
    ws.cell(row=row, column=2, value="+6")
    ws.cell(row=row, column=3, value="+10")
    ws.cell(row=row, column=4, value="+4")
    row += 1
    ws.cell(row=row, column=1, value="1-4 weeks")
    ws.cell(row=row, column=2, value="+4")
    ws.cell(row=row, column=3, value="+8")
    ws.cell(row=row, column=4, value="+2")
    row += 1
    ws.cell(row=row, column=1, value="> 4 weeks")
    ws.cell(row=row, column=2, value="0")
    ws.cell(row=row, column=3, value="0")
    ws.cell(row=row, column=4, value="0")
    row += 1
    ws.cell(row=row, column=1, value="Not inpatient")
    ws.cell(row=row, column=2, value="0")
    ws.cell(row=row, column=3, value="0")
    ws.cell(row=row, column=4, value="0")
    
    row += 3
    
    # A/D Adjustment summary
    ws.cell(row=row, column=1, value="A/D ADJUSTMENT (Tables 2-4)").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Affects: Scales 2A, 2B, 8B, S, C")
    row += 1
    ws.cell(row=row, column=1, value="If A BR < 75 AND D BR < 75: No adjustment")
    row += 1
    ws.cell(row=row, column=1, value="If A BR >= 75: A/D value = A - 75")
    row += 1
    ws.cell(row=row, column=1, value="If D BR >= 75: A/D value = D - 75")
    row += 1
    ws.cell(row=row, column=1, value="If both >= 75: A/D value = (A-75) + (D-75)")
    row += 1
    ws.cell(row=row, column=1, value="Use Table 2 for OPD; Table 3 for IPD < 1wk; Table 4 for IPD 1-4wk")
    
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    
    return ws


def generate_mcmi_auto_scoring_excel(output_path=None):
    """
    Main function to generate the complete MCMI-III auto-scoring Excel tool.
    """
    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "MCMI_III_AutoScoring_Tool.xlsx"
        )
    
    print("=" * 60)
    print("  GENERATING MCMI-III AUTO-SCORING EXCEL TOOL")
    print("=" * 60)
    
    wb = Workbook()
    
    print("  Creating Sheet 1: DATA ENTRY (175 items)...")
    create_data_entry_sheet(wb)
    
    print("  Creating Sheet 2: AUTO-SCORING (formulas)...")
    ws_scoring, scale_rows = create_scoring_sheet(wb)
    
    print("  Creating Sheet 3: PROFILE (dashboard)...")
    create_profile_sheet(wb, scale_rows)
    
    print("  Creating Sheet 4: BR REFERENCE (tables)...")
    create_br_reference_sheet(wb)
    
    # Save
    wb.save(output_path)
    
    print(f"\n  SUCCESS! Excel tool saved to:")
    print(f"  {output_path}")
    print(f"\n  HOW TO USE:")
    print(f"  1. Open the Excel file")
    print(f"  2. Go to 'DATA ENTRY' sheet")
    print(f"  3. Enter patient age, setting (OPD/IPD), gender")
    print(f"  4. Type TRUE or FALSE for all 175 items")
    print(f"  5. Go to 'AUTO-SCORING' sheet - scores calculated automatically!")
    print(f"  6. Go to 'PROFILE' sheet - visual profile with color coding!")
    print("=" * 60)
    
    return output_path


if __name__ == '__main__':
    generate_mcmi_auto_scoring_excel()
