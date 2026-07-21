"""
MCMI-III EXCEL WORKBOOK GENERATOR
==================================
Generates a multi-sheet Excel workbook that mirrors the
5-sheet structure requested:
  Sheet 1: Data Entry (175 items)
  Sheet 2: Raw Score Calculation
  Sheet 3: Adjustment Calculations
  Sheet 4: BR Conversion Tables
  Sheet 5: Final Profile Output

Requires: openpyxl
"""

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from mcmi_item_keys import (
    SCALE_ITEMS, SCALE_NAMES, SCALE_V_ITEMS,
    CLINICAL_PERSONALITY, SEVERE_PERSONALITY,
    CLINICAL_SYNDROMES, SEVERE_SYNDROMES
)
from mcmi_br_tables import (
    DISCLOSURE_ADJUSTMENT_TABLE, INPATIENT_ADJUSTMENT,
    BR_CONVERSION_ANCHORS
)



# Style definitions
HEADER_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF", size=11)
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="B4D7FF", end_color="B4D7FF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def generate_excel_workbook(results, output_path="MCMI_III_Scoring_Workbook.xlsx"):
    """
    Generate a complete MCMI-III scoring workbook in Excel format.
    
    Parameters:
        results: The results dictionary from MCMIIIScorer.score()
        output_path: Path to save the Excel file
    """
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
        return None
    
    wb = Workbook()
    
    # Sheet 1: Data Entry
    _create_data_entry_sheet(wb, results)
    
    # Sheet 2: Raw Score Calculation
    _create_raw_score_sheet(wb, results)
    
    # Sheet 3: Adjustment Calculations
    _create_adjustment_sheet(wb, results)
    
    # Sheet 4: BR Conversion Tables
    _create_br_tables_sheet(wb, results)
    
    # Sheet 5: Final Profile Output
    _create_profile_sheet(wb, results)
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    wb.save(output_path)
    print(f"  Excel workbook saved to: {output_path}")
    return output_path



def _create_data_entry_sheet(wb, results):
    """Sheet 1: Data Entry (175 items with validation)."""
    ws = wb.create_sheet("1-Data Entry", 0)
    
    # Title
    ws['A1'] = "MCMI-III DATA ENTRY SHEET"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')
    
    ws['A3'] = "Instructions:"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = "Enter 1 for TRUE or 0 for FALSE for each item"
    ws['A5'] = "Leave blank if item was omitted"
    ws['A6'] = f"Total items: 175 | Max omissions allowed: 11"
    
    # Headers
    headers = ['Item #', 'Response (1/0)', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    
    # Item rows
    for item in range(1, 176):
        row = item + 8
        ws.cell(row=row, column=1, value=item).alignment = Alignment(horizontal='center')
        
        # Response cell (would be data-entry in live version)
        resp_cell = ws.cell(row=row, column=2)
        resp_cell.alignment = Alignment(horizontal='center')
        resp_cell.border = THIN_BORDER
        
        # Status formula placeholder
        status_cell = ws.cell(row=row, column=3)
        status_cell.alignment = Alignment(horizontal='center')
    
    # Summary section
    summary_row = 185
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = HEADER_FONT
    ws.cell(row=summary_row+1, column=1, value="Items Answered:")
    ws.cell(row=summary_row+2, column=1, value="Items Omitted:")
    ws.cell(row=summary_row+3, column=1, value="Protocol Status:")
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12


def _create_raw_score_sheet(wb, results):
    """Sheet 2: Raw Score Calculation."""
    ws = wb.create_sheet("2-Raw Scores", 1)
    
    # Title
    ws['A1'] = "MCMI-III RAW SCORE CALCULATION"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')
    
    # Validity section
    ws['A3'] = "VALIDITY INDICES"
    ws['A3'].font = HEADER_FONT
    
    headers = ['Scale', 'Name', 'Items', 'Raw Score', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
    
    # V Scale
    v_score = results['raw_scores'].get('V', 'N/A')
    ws.cell(row=5, column=1, value="V")
    ws.cell(row=5, column=2, value="Invalidity")
    ws.cell(row=5, column=3, value="65, 110, 157")
    ws.cell(row=5, column=4, value=v_score)
    ws.cell(row=5, column=5, value="VALID" if v_score == 0 else 
            "QUESTIONABLE" if v_score == 1 else "INVALID")
    
    # W Scale
    w_score = results['raw_scores'].get('W', 'N/A')
    ws.cell(row=6, column=1, value="W")
    ws.cell(row=6, column=2, value="Inconsistency")
    ws.cell(row=6, column=3, value="44 item pairs")
    ws.cell(row=6, column=4, value=w_score)
    ws.cell(row=6, column=5, value="VALID" if w_score < 8 else 
            "QUESTIONABLE" if w_score < 10 else "INVALID")
    
    # All scales raw scores
    row = 9
    ws.cell(row=row, column=1, value="ALL SCALE RAW SCORES").font = HEADER_FONT
    row += 1
    
    # Headers
    headers2 = ['Scale', 'Full Name', 'Category', 'Raw Score', 'Max Possible']
    for col, header in enumerate(headers2, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
    row += 1
    
    # Scale categories
    categories = {
        'X': 'Modifying Index', 'Y': 'Modifying Index', 'Z': 'Modifying Index',
    }
    for s in CLINICAL_PERSONALITY:
        categories[s] = 'Clinical Personality'
    for s in SEVERE_PERSONALITY:
        categories[s] = 'Severe Personality'
    for s in CLINICAL_SYNDROMES:
        categories[s] = 'Clinical Syndrome'
    for s in SEVERE_SYNDROMES:
        categories[s] = 'Severe Syndrome'
    
    all_scales = ['X', 'Y', 'Z'] + CLINICAL_PERSONALITY + SEVERE_PERSONALITY + CLINICAL_SYNDROMES + SEVERE_SYNDROMES
    
    for scale in all_scales:
        raw = results['raw_scores'].get(scale, 'N/A')
        ws.cell(row=row, column=1, value=scale)
        ws.cell(row=row, column=2, value=SCALE_NAMES.get(scale, ''))
        ws.cell(row=row, column=3, value=categories.get(scale, ''))
        ws.cell(row=row, column=4, value=raw)
        
        # Color code
        if isinstance(raw, (int, float)) and scale in results.get('final_br_scores', {}):
            br = results['final_br_scores'][scale]
            if br >= 85:
                ws.cell(row=row, column=4).fill = RED_FILL
            elif br >= 75:
                ws.cell(row=row, column=4).fill = ORANGE_FILL
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14



def _create_adjustment_sheet(wb, results):
    """Sheet 3: Adjustment Calculations."""
    ws = wb.create_sheet("3-Adjustments", 2)
    
    ws['A1'] = "MCMI-III ADJUSTMENT CALCULATIONS"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    row = 3
    
    # Disclosure Adjustment
    ws.cell(row=row, column=1, value="1. DISCLOSURE ADJUSTMENT").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Scale X Raw Score:")
    ws.cell(row=row, column=2, value=results['raw_scores'].get('X', 'N/A'))
    row += 1
    ws.cell(row=row, column=1, value="Valid Range:")
    ws.cell(row=row, column=2, value="34-178")
    row += 1
    
    # Find the disclosure adjustment from log
    for log_entry in results.get('adjustment_log', []):
        if 'Disclosure Adjustment' in log_entry:
            ws.cell(row=row, column=1, value="Applied:")
            ws.cell(row=row, column=2, value=log_entry)
            row += 1
            break
    
    row += 2
    
    # A/D Adjustment
    ws.cell(row=row, column=1, value="2. ANXIETY/DEPRESSION (A/D) ADJUSTMENT").font = HEADER_FONT
    row += 1
    
    setting = results['patient_info']['setting']
    duration = results['patient_info']['axis_duration_weeks']
    
    ws.cell(row=row, column=1, value="Patient Setting:")
    ws.cell(row=row, column=2, value=f"{setting} ({'Outpatient' if setting=='OPD' else 'Inpatient'})")
    row += 1
    
    if setting == 'IPD':
        ws.cell(row=row, column=1, value="Axis I Duration:")
        ws.cell(row=row, column=2, value=f"{duration} weeks")
        row += 1
    
    ws.cell(row=row, column=1, value="Table Used:")
    if setting == 'OPD' or (duration and duration > 4):
        ws.cell(row=row, column=2, value="Table 2 (Non-Inpatient / IPD > 4 weeks)")
    elif duration and duration < 1:
        ws.cell(row=row, column=2, value="Table 3 (Inpatient < 1 week)")
    elif duration and 1 <= duration <= 4:
        ws.cell(row=row, column=2, value="Table 4 (Inpatient 1-4 weeks)")
    row += 1
    
    for log_entry in results.get('adjustment_log', []):
        if 'A/D Adjustment' in log_entry:
            ws.cell(row=row, column=1, value="Applied:")
            ws.cell(row=row, column=2, value=log_entry)
            row += 1
            break
    
    row += 2
    
    # Inpatient Adjustment
    ws.cell(row=row, column=1, value="3. INPATIENT ADJUSTMENT (Table 5)").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Affects: Scales SS, CC, PP")
    row += 1
    
    for log_entry in results.get('adjustment_log', []):
        if 'Inpatient Adjustment' in log_entry:
            ws.cell(row=row, column=1, value="Applied:")
            ws.cell(row=row, column=2, value=log_entry)
            row += 1
            break
    
    row += 2
    
    # Denial/Complaint Adjustment
    ws.cell(row=row, column=1, value="4. DENIAL/COMPLAINT ADJUSTMENT").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Affects: Highest of Scales 4, 5, 7 (+8 points)")
    row += 1
    
    for log_entry in results.get('adjustment_log', []):
        if 'Denial/Complaint' in log_entry:
            ws.cell(row=row, column=1, value="Applied:")
            ws.cell(row=row, column=2, value=log_entry)
            row += 1
            break
    
    row += 2
    
    # Complete Adjustment Log
    ws.cell(row=row, column=1, value="COMPLETE ADJUSTMENT LOG").font = HEADER_FONT
    row += 1
    for entry in results.get('adjustment_log', []):
        ws.cell(row=row, column=1, value=entry)
        row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 60


def _create_br_tables_sheet(wb, results):
    """Sheet 4: BR Conversion Tables."""
    ws = wb.create_sheet("4-BR Tables", 3)
    
    ws['A1'] = "MCMI-III BASE RATE (BR) CONVERSION REFERENCE"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    row = 3
    ws.cell(row=row, column=1, value="BR SCORE INTERPRETATION").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="BR < 60")
    ws.cell(row=row, column=2, value="Not significant")
    row += 1
    ws.cell(row=row, column=1, value="BR 60-74")
    ws.cell(row=row, column=2, value="Trait/Tendency present")
    ws.cell(row=row, column=2).fill = YELLOW_FILL
    row += 1
    ws.cell(row=row, column=1, value="BR 75-84")
    ws.cell(row=row, column=2, value="Clinically Significant")
    ws.cell(row=row, column=2).fill = ORANGE_FILL
    row += 1
    ws.cell(row=row, column=1, value="BR >= 85")
    ws.cell(row=row, column=2, value="PROMINENT - Disorder Present")
    ws.cell(row=row, column=2).fill = RED_FILL
    row += 2
    
    # Disclosure Adjustment Table
    ws.cell(row=row, column=1, value="DISCLOSURE ADJUSTMENT TABLE (Table 1)").font = HEADER_FONT
    row += 1
    headers = ['Scale X Raw Score', '1-8B Adjustment', 'S-PP Adjustment']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
    row += 1
    
    # Key entries from the disclosure table
    key_entries = [
        (34, "+20", "+10"), (40, "+17", "+9"), (45, "+13", "+7"),
        (50, "+9", "+5"), (55, "+4", "+3"), (60, "+1", "+1"),
        ("61-123", "0", "0"),
        (124, "-1", "-1"), (130, "-3", "-2"), (140, "-8", "-5"),
        (150, "-11", "-8"), (160, "-15", "-10"), (170, "-19", "-13"),
        (178, "-20", "-14"),
    ]
    for entry in key_entries:
        ws.cell(row=row, column=1, value=entry[0])
        ws.cell(row=row, column=2, value=entry[1])
        ws.cell(row=row, column=3, value=entry[2])
        row += 1
    
    row += 2
    
    # Inpatient Adjustment Table
    ws.cell(row=row, column=1, value="INPATIENT ADJUSTMENT TABLE (Table 5)").font = HEADER_FONT
    row += 1
    headers = ['Duration', 'SS Adjustment', 'CC Adjustment', 'PP Adjustment']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
    row += 1
    ws.cell(row=row, column=1, value="< 1 week")
    ws.cell(row=row, column=2, value="+6")
    ws.cell(row=row, column=3, value="+10")
    ws.cell(row=row, column=4, value="+4")
    row += 1
    ws.cell(row=row, column=1, value="1-4 weeks")
    ws.cell(row=row, column=2, value="+4")
    ws.cell(row=row, column=3, value="+8")
    ws.cell(row=row, column=4, value="+2")
    row += 1
    ws.cell(row=row, column=1, value="> 4 weeks")
    ws.cell(row=row, column=2, value="0")
    ws.cell(row=row, column=3, value="0")
    ws.cell(row=row, column=4, value="0")
    
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18



def _create_profile_sheet(wb, results):
    """Sheet 5: Final Profile Output (Dashboard)."""
    ws = wb.create_sheet("5-Final Profile", 4)
    
    ws['A1'] = "MCMI-III FINAL SCORING PROFILE"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')
    
    # Validity Banner
    validity = results['validity']['status']
    ws['A3'] = f"VALIDITY STATUS: {validity}"
    ws['A3'].font = Font(bold=True, size=14)
    if validity == 'INVALID':
        ws['A3'].fill = RED_FILL
    elif validity == 'QUESTIONABLE':
        ws['A3'].fill = ORANGE_FILL
    else:
        ws['A3'].fill = GREEN_FILL
    
    # Patient Info
    pi = results['patient_info']
    ws['A5'] = f"Age: {pi['age']} | Gender: {pi['gender'] or 'N/A'} | Setting: {pi['setting']}"
    if pi['setting'] == 'IPD':
        ws['A5'].value += f" | Axis I Duration: {pi['axis_duration_weeks']} wks"
    
    # Main scoring table
    row = 7
    headers = ['Scale', 'Name', 'Category', 'Raw Score', 
               'Initial BR', 'Final BR', 'Interpretation']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    if results['validity']['status'] == 'INVALID':
        ws.cell(row=row, column=1, value="PROTOCOL IS INVALID - CANNOT SCORE")
        ws.cell(row=row, column=1).font = Font(bold=True, color="FF0000")
        ws.merge_cells(f'A{row}:G{row}')
        return
    
    # Category labels
    categories = {
        'Modifying Indices': ['X', 'Y', 'Z'],
        'Clinical Personality Patterns': CLINICAL_PERSONALITY,
        'Severe Personality Pathology': SEVERE_PERSONALITY,
        'Clinical Syndromes': CLINICAL_SYNDROMES,
        'Severe Clinical Syndromes': SEVERE_SYNDROMES,
    }
    
    for cat_name, scales in categories.items():
        # Category header
        ws.cell(row=row, column=1, value=cat_name).font = Font(bold=True, italic=True)
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = LIGHT_BLUE_FILL
        row += 1
        
        for scale in scales:
            if scale not in results.get('final_br_scores', {}):
                continue
            
            raw = results['raw_scores'].get(scale, '-')
            init_br = results['initial_br_scores'].get(scale, '-')
            final_br = results['final_br_scores'][scale]
            
            from mcmi_br_tables import interpret_br_score
            interp = interpret_br_score(final_br)
            
            ws.cell(row=row, column=1, value=scale).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value=SCALE_NAMES.get(scale, ''))
            ws.cell(row=row, column=3, value=cat_name)
            ws.cell(row=row, column=4, value=raw).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=init_br).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=final_br).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=7, value=interp)
            
            # Conditional formatting
            br_cell = ws.cell(row=row, column=6)
            if final_br >= 85:
                br_cell.fill = RED_FILL
                br_cell.font = Font(bold=True, color="FFFFFF")
            elif final_br >= 75:
                br_cell.fill = ORANGE_FILL
                br_cell.font = Font(bold=True)
            elif final_br >= 60:
                br_cell.fill = YELLOW_FILL
            
            # Add border to all cells in row
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = THIN_BORDER
            
            row += 1
        
        row += 1  # Space between categories
    
    # Elevated Scales Summary
    row += 2
    ws.cell(row=row, column=1, value="ELEVATED SCALES SUMMARY").font = HEADER_FONT
    row += 1
    
    elevated = results.get('elevated_scales', {})
    
    if elevated.get('prominent_85+'):
        ws.cell(row=row, column=1, value="PROMINENT (BR >= 85):").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = RED_FILL
        row += 1
        for s in elevated['prominent_85+']:
            ws.cell(row=row, column=2, value=s)
            row += 1
    
    if elevated.get('significant_75_84'):
        ws.cell(row=row, column=1, value="SIGNIFICANT (BR 75-84):").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = ORANGE_FILL
        row += 1
        for s in elevated['significant_75_84']:
            ws.cell(row=row, column=2, value=s)
            row += 1
    
    # Flat profile warning
    if results.get('flat_profile'):
        row += 1
        ws.cell(row=row, column=1, value="WARNING: FLAT PROFILE - UNINTERPRETABLE")
        ws.cell(row=row, column=1).font = Font(bold=True, color="FF0000", size=12)
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 28


if __name__ == '__main__':
    # Test with demo data
    from mcmi_main import run_demo
    print("Generating Excel workbook with demo data...")
    
    from mcmi_scoring_engine import MCMIIIScorer
    import random
    random.seed(42)
    
    # Quick demo setup
    responses = {i: False for i in range(1, 176)}
    for item in [2, 20, 22, 41, 50, 53, 56, 71, 81, 98, 121, 130, 135, 158, 172,
                 16, 35, 44, 48, 83, 87, 92, 107, 114, 118, 122, 136, 154, 163,
                 46, 78, 147, 5, 15, 33, 60, 70, 75, 80, 86, 91,
                 9, 30, 72, 109, 148, 175, 76, 96, 100,
                 4, 68, 74, 82, 95, 99,
                 3, 10, 19, 25, 39, 45, 49, 57, 119, 125, 150, 160, 165, 166, 169]:
        responses[item] = True
    responses[65] = False
    responses[110] = False
    responses[157] = False
    
    scorer = MCMIIIScorer()
    scorer.set_responses(responses)
    scorer.set_patient_info(age=35, setting='OPD', gender='F')
    results = scorer.score()
    
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "MCMI_III_Scoring_Workbook.xlsx")
    generate_excel_workbook(results, output_path)
