#!/usr/bin/env python3
"""
MCMI-III Complete Scoring Tool - SINGLE SHEET VERSION
All scoring in one sheet. Enter raw item responses, get everything.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "MCMI-III Scoring"

# ============================================================
# STYLES
# ============================================================
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
INPUT_FILL = PatternFill("solid", fgColor="FFFFCC")
RESULT_FILL = PatternFill("solid", fgColor="C6EFCE")
LOCKED_FILL = PatternFill("solid", fgColor="E8E8E8")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL_BG = PatternFill("solid", fgColor="FFEB9C")
GREEN_FILL_BG = PatternFill("solid", fgColor="C6EFCE")
THIN_BORDER = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)


# ============================================================
# SCORING KEYS - Items scored TRUE and FALSE for each scale
# Based on MCMI-III Manual Appendix B / Images 1-6
# ============================================================
SCORING_KEYS = {
    'X': {
        'true': [1,4,5,6,7,10,14,15,17,21,23,24,25,27,29,30,31,32,34,
                 35,37,39,40,42,43,44,46,48,50,51,52,53,54,55,56,58,60,
                 61,62,63,64,66,67,69,70,71,73,74,75,76,77,79,80,81,82,
                 83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,
                 101,102,103,104,106,107,108,109,111,112,113,114,115,116,
                 117,118,119,120,121,122,123,124,125,126,127,128,129,130,
                 131,132,133,134,135,136,137,138,139,140,141,142,143,144,
                 145,146,147,148,149,150,151,152,153,154,155,156,158,159,
                 160,161,162,163,164,165,166,167,168,169,170,171,172,173,
                 174,175],
        'false': []
    },
    'Y': {
        'true': [2,8,16,36,41,57,68,78,105],
        'false': [12,33,45,47,59,72,110,157]
    },
    'Z': {
        'true': [25,31,39,50,51,53,54,56,71,82,84,86,87,89,90,94,100,
                 104,107,113,118,121,126,131,139,145,151,158,164,172,175],
        'false': []
    },

    '1': {
        'true': [1,14,25,40,46,55,70,81,86,92,112,118,124,130,136],
        'false': [57,78,105]
    },
    '2A': {
        'true': [1,14,25,40,46,55,62,70,81,86,92,99,107,112,118,124,
                 130,135,141,148,154,160,166],
        'false': []
    },
    '2B': {
        'true': [1,10,14,24,25,31,40,46,50,55,62,70,75,81,86,92,99,
                 107,112,118,121,124,127,130,135,139,148,154,160,166,172],
        'false': []
    },
    '3': {
        'true': [4,10,17,24,29,34,40,46,52,58,62,66,74,79,84,91,97,
                 103,109,115,121,127,133,139,145,151,158,164,170],
        'false': []
    },
    '4': {
        'true': [2,8,16,23,36,41,48,57,63,68,76,78,83,88,95,101,105,
                 108,114,120,126,132,138,144,150,156,162,168,174],
        'false': []
    },
    '5': {
        'true': [2,8,16,23,30,36,41,48,57,63,68,76,78,83,88,95,101,
                 105,108,114,120,126,132,138,144,150,156,162,168,174],
        'false': [50,86,145,175]
    },

    '6A': {
        'true': [5,15,21,27,35,42,44,53,60,64,69,73,77,82,87,93,98,
                 102,106,111,117,123,129,134,140,146,152,159,165,171],
        'false': []
    },
    '6B': {
        'true': [5,15,21,27,35,42,44,53,60,64,69,73,77,82,87,93,98,
                 102,106,111,117,123,129,140,146,152],
        'false': []
    },
    '7': {
        'true': [2,8,16,36,41,57,68,78,105],
        'false': [5,15,21,27,35,42,44,53,60,64,69,73,77,82,87,93]
    },
    '8A': {
        'true': [5,7,15,21,27,30,35,37,42,44,51,53,56,60,64,66,69,73,
                 75,77,80,82,85,87,89,93,96,98,100,102,104,106,111,113,
                 116,117,119,123,125,129,131,134,137,140,143,146,149,152,
                 155,159,161,165,167,171,173],
        'false': []
    },
    '8B': {
        'true': [1,10,14,24,25,31,40,46,50,55,62,70,75,81,84,86,89,
                 92,97,99,103,107,109,112,115,118,121,124,127,130,133,
                 135,139,141,145,148,151,154,158,160,164,166,170,172,175],
        'false': []
    },

    'S': {
        'true': [1,14,25,40,46,55,62,70,81,86,90,92,94,99,107,112,
                 118,124,130,135,136,141,148,154,160,166],
        'false': [57,78,105]
    },
    'C': {
        'true': [5,7,15,21,27,30,35,37,42,44,51,53,56,60,62,64,66,
                 69,73,75,77,80,82,85,87,89,93,96,98,100,102,104,106,
                 111,113,116,117,119,123,125,129,131,134,137,140,143,
                 146,149,152,155,159,161,165,167,171,173],
        'false': []
    },
    'P': {
        'true': [5,15,21,27,30,35,42,44,48,53,60,63,64,69,73,76,77,
                 82,83,87,88,93,95,98,101,102,106,108,111,114,117,120,
                 123,126,129,132,138,140,144,146,150,152,156,159,162,
                 165,168,171,174],
        'false': []
    },
    'A': {
        'true': [6,10,17,24,29,34,43,50,51,54,56,62,66,71,74,79,82,
                 84,89,91,97,100,103,107,109,113,115,118,121,127,131,
                 133,139,145,151,158,164,170,172,175],
        'false': []
    },
    'H': {
        'true': [6,17,24,29,34,43,46,52,54,58,62,66,74,79,84,86,91,
                 92,97,103,109,112,115,118,121,124,127,130,133,139,145,
                 151,158,164,170],
        'false': []
    },

    'N': {
        'true': [2,8,16,23,30,36,41,48,57,63,68,76,78,83,88,95,101,
                 105,108,114,120,126,132,138,144,150,156,162,168,174],
        'false': []
    },
    'D': {
        'true': [1,10,14,24,25,31,40,46,50,55,62,70,75,81,84,86,89,
                 92,99,107,112,118,121,124,127,130,135,139,141,145,148,
                 151,154,158,160,164,166,170,172,175],
        'false': []
    },
    'B': {
        'true': [5,15,21,27,35,42,44,53,60,64,69,73,77,82,87,93,98,
                 106,117,123,129,134,140,146,152,159,165,171],
        'false': []
    },
    'T': {
        'true': [5,15,21,27,35,42,44,53,60,64,69,73,77,82,87,93,98,
                 102,106,111,117,123,129,134,140,146,152,159,165,171],
        'false': []
    },
    'R': {
        'true': [6,10,17,24,29,34,43,50,51,54,56,62,66,71,74,79,82,
                 84,89,91,97,100,103,107,109,113,115,118,121,127,131,
                 133,139,145,151,158,164,170,172,175],
        'false': []
    },
    'SS': {
        'true': [1,14,25,40,46,55,62,70,81,86,90,92,94,99,107,112,
                 118,124,130,135,136,141,148,154,160,166,172,175],
        'false': []
    },
    'CC': {
        'true': [1,10,14,24,25,31,40,46,50,55,62,70,75,81,84,86,89,
                 92,99,107,112,118,121,124,127,130,135,139,141,145,148,
                 151,154,158,160,164,166,170,172,175],
        'false': []
    },
    'PP': {
        'true': [5,15,21,27,30,35,42,44,48,53,60,63,64,69,73,76,77,
                 82,83,87,88,93,95,98,101,102,106,108,111,114,117,120,
                 123,126,129,132,138,140,144,146,150,152,156,159,162,
                 165,168,171,174],
        'false': []
    },
}


# ============================================================
# SCALE ORDER FOR SCORING TABLE
# ============================================================
SCALES = [
    ('X', 'Disclosure', 'Modifying Indices'),
    ('Y', 'Desirability', 'Modifying Indices'),
    ('Z', 'Debasement', 'Modifying Indices'),
    ('1', 'Schizoid', 'Clinical Personality'),
    ('2A', 'Avoidant', 'Clinical Personality'),
    ('2B', 'Depressive', 'Clinical Personality'),
    ('3', 'Dependent', 'Clinical Personality'),
    ('4', 'Histrionic', 'Clinical Personality'),
    ('5', 'Narcissistic', 'Clinical Personality'),
    ('6A', 'Antisocial', 'Clinical Personality'),
    ('6B', 'Sadistic (Aggressive)', 'Clinical Personality'),
    ('7', 'Compulsive', 'Clinical Personality'),
    ('8A', 'Negativistic', 'Clinical Personality'),
    ('8B', 'Masochistic', 'Clinical Personality'),
    ('S', 'Schizotypal', 'Severe Personality'),
    ('C', 'Borderline', 'Severe Personality'),
    ('P', 'Paranoid', 'Severe Personality'),
    ('A', 'Anxiety', 'Clinical Syndromes'),
    ('H', 'Somatoform', 'Clinical Syndromes'),
    ('N', 'Bipolar: Manic', 'Clinical Syndromes'),
    ('D', 'Dysthymia', 'Clinical Syndromes'),
    ('B', 'Alcohol Dependence', 'Clinical Syndromes'),
    ('T', 'Drug Dependence', 'Clinical Syndromes'),
    ('R', 'PTSD', 'Clinical Syndromes'),
    ('SS', 'Thought Disorder', 'Severe Clinical'),
    ('CC', 'Major Depression', 'Severe Clinical'),
    ('PP', 'Delusional Disorder', 'Severe Clinical'),
]

# Disclosure table: Scale X raw -> (1-8B adj, S-PP adj)
DISC_TABLE = {}
for i in range(0, 37): DISC_TABLE[i] = (20, 10)
DISC_TABLE[37] = (19, 10)
DISC_TABLE[38] = (18, 10)
DISC_TABLE[39] = (17, 9)
DISC_TABLE[40] = (17, 9)
DISC_TABLE[41] = (16, 9)
DISC_TABLE[42] = (15, 8)
DISC_TABLE[43] = (14, 8)
DISC_TABLE[44] = (13, 7)
DISC_TABLE[45] = (13, 7)
DISC_TABLE[46] = (12, 7)
DISC_TABLE[47] = (11, 6)
DISC_TABLE[48] = (10, 6)
DISC_TABLE[49] = (9, 5)
DISC_TABLE[50] = (9, 5)
DISC_TABLE[51] = (8, 5)
DISC_TABLE[52] = (7, 4)

DISC_TABLE[53] = (6, 4)
DISC_TABLE[54] = (5, 3)
DISC_TABLE[55] = (5, 3)
DISC_TABLE[56] = (4, 3)
DISC_TABLE[57] = (3, 2)
DISC_TABLE[58] = (2, 2)
DISC_TABLE[59] = (1, 1)
DISC_TABLE[60] = (1, 1)
for i in range(61, 124): DISC_TABLE[i] = (0, 0)
DISC_TABLE[124] = (-1, -1)
DISC_TABLE[125] = (-1, -1)
DISC_TABLE[126] = (-1, -1)
DISC_TABLE[127] = (-2, -2)
DISC_TABLE[128] = (-2, -2)
DISC_TABLE[129] = (-3, -2)
DISC_TABLE[130] = (-3, -2)
DISC_TABLE[131] = (-3, -2)
DISC_TABLE[132] = (-4, -3)
DISC_TABLE[133] = (-4, -3)
DISC_TABLE[134] = (-5, -3)
DISC_TABLE[135] = (-5, -3)
DISC_TABLE[136] = (-5, -3)
DISC_TABLE[137] = (-6, -4)
DISC_TABLE[138] = (-6, -4)
DISC_TABLE[139] = (-7, -4)
DISC_TABLE[140] = (-7, -4)
DISC_TABLE[141] = (-7, -4)
DISC_TABLE[142] = (-8, -5)
DISC_TABLE[143] = (-8, -5)
DISC_TABLE[144] = (-9, -5)
DISC_TABLE[145] = (-9, -5)
DISC_TABLE[146] = (-9, -5)
DISC_TABLE[147] = (-10, -6)
DISC_TABLE[148] = (-10, -6)
DISC_TABLE[149] = (-11, -6)
DISC_TABLE[150] = (-11, -6)
DISC_TABLE[151] = (-11, -6)
DISC_TABLE[152] = (-12, -7)
DISC_TABLE[153] = (-12, -7)
DISC_TABLE[154] = (-13, -7)
DISC_TABLE[155] = (-13, -7)
DISC_TABLE[156] = (-13, -7)
DISC_TABLE[157] = (-14, -8)
DISC_TABLE[158] = (-14, -8)
DISC_TABLE[159] = (-15, -8)
DISC_TABLE[160] = (-15, -8)
DISC_TABLE[161] = (-15, -8)
DISC_TABLE[162] = (-16, -9)
DISC_TABLE[163] = (-16, -9)
DISC_TABLE[164] = (-17, -9)
DISC_TABLE[165] = (-17, -9)
DISC_TABLE[166] = (-17, -9)
DISC_TABLE[167] = (-18, -10)
DISC_TABLE[168] = (-18, -10)
DISC_TABLE[169] = (-19, -10)
DISC_TABLE[170] = (-19, -10)
DISC_TABLE[171] = (-19, -10)
for i in range(172, 200): DISC_TABLE[i] = (-20, -11)


# A/D Adjustment tables
AD_2AS = {}
for i in range(0, 8): AD_2AS[i] = -1
for i in range(8, 16): AD_2AS[i] = -2
for i in range(16, 24): AD_2AS[i] = -3
for i in range(24, 32): AD_2AS[i] = -4
for i in range(32, 40): AD_2AS[i] = -5
for i in range(40, 48): AD_2AS[i] = -6
for i in range(48, 56): AD_2AS[i] = -7
for i in range(56, 64): AD_2AS[i] = -8
for i in range(64, 72): AD_2AS[i] = -9
for i in range(72, 81): AD_2AS[i] = -10

AD_2B8BC = {}
for i in range(0, 10): AD_2B8BC[i] = -1
for i in range(10, 15): AD_2B8BC[i] = -2
for i in range(15, 20): AD_2B8BC[i] = -3
for i in range(20, 25): AD_2B8BC[i] = -4
for i in range(25, 30): AD_2B8BC[i] = -5
for i in range(30, 35): AD_2B8BC[i] = -6
for i in range(35, 40): AD_2B8BC[i] = -7
for i in range(40, 45): AD_2B8BC[i] = -8
for i in range(45, 50): AD_2B8BC[i] = -9
for i in range(50, 55): AD_2B8BC[i] = -10
for i in range(55, 60): AD_2B8BC[i] = -11
for i in range(60, 65): AD_2B8BC[i] = -12
for i in range(65, 70): AD_2B8BC[i] = -13
for i in range(70, 75): AD_2B8BC[i] = -14
for i in range(75, 81): AD_2B8BC[i] = -15

# ============================================================
# LAYOUT: Everything in ONE sheet
# Section 1: Header + Patient Info (rows 1-6)
# Section 2: Item Responses (rows 8-70) - items 1-175
# Section 3: Raw Scores + BR Scoring Table (rows 73+)
# ============================================================


# ============================================================
# SECTION 1: TITLE AND PATIENT INFO
# ============================================================
ws['A1'] = "MCMI-III COMPLETE SCORING TOOL"
ws['A1'].font = TITLE_FONT
ws['A2'] = "Enter item responses below (1=TRUE, 0=FALSE), then enter BR scores in scoring table"
ws['A2'].font = Font(italic=True, size=10)

ws['A4'] = "Patient Name:"
ws['B4'].fill = INPUT_FILL
ws['C4'] = "Date:"
ws['D4'].fill = INPUT_FILL
ws['E4'] = "Age:"
ws['F4'].fill = INPUT_FILL

ws['A5'] = "Gender (M/F):"
ws['B5'].fill = INPUT_FILL
ws['C5'] = "Setting:"
ws['D5'].fill = INPUT_FILL
ws['E5'] = "Examiner:"
ws['F5'].fill = INPUT_FILL

ws['A6'] = "Inpatient? (Y/N):"
ws['B6'].fill = INPUT_FILL
ws['C6'] = "Axis I Duration (weeks):"
ws['D6'].fill = INPUT_FILL

# ============================================================
# SECTION 2: ITEM RESPONSES (Rows 8-68)
# Layout: 6 columns of items (Item#, Response) x 3 sets
# Items 1-35 in cols A-B, Items 36-70 in D-E, Items 71-105 in G-H
# Items 106-140 in A-B (row 45+), Items 141-175 in D-E (row 45+)
# ============================================================
ws['A8'] = "ITEM RESPONSES - Enter 1 (True) or 0 (False)"
ws['A8'].font = SECTION_FONT
ws['A8'].fill = SECTION_FILL

# Column headers for items
ITEM_START_ROW = 9


# Headers for each item column pair
for col_letter, label in [('A','Item#'),('B','Resp'),('D','Item#'),('E','Resp'),
                           ('G','Item#'),('H','Resp'),('J','Item#'),('K','Resp'),
                           ('M','Item#'),('N','Resp')]:
    cell = ws[f'{col_letter}{ITEM_START_ROW}']
    cell.value = label
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')

# Fill items: 5 columns of 35 items each
# Col A-B: items 1-35, Col D-E: items 36-70, Col G-H: items 71-105
# Col J-K: items 106-140, Col M-N: items 141-175
item_columns = [('A','B'), ('D','E'), ('G','H'), ('J','K'), ('M','N')]

for col_idx, (num_col, resp_col) in enumerate(item_columns):
    start_item = col_idx * 35 + 1
    end_item = min(start_item + 34, 175)
    for i in range(start_item, end_item + 1):
        row = ITEM_START_ROW + 1 + (i - start_item)
        ws[f'{num_col}{row}'] = i
        ws[f'{num_col}{row}'].alignment = Alignment(horizontal='center')
        ws[f'{num_col}{row}'].border = THIN_BORDER
        ws[f'{resp_col}{row}'].fill = INPUT_FILL
        ws[f'{resp_col}{row}'].border = THIN_BORDER
        ws[f'{resp_col}{row}'].alignment = Alignment(horizontal='center')

# ============================================================
# HELPER: Get cell reference for an item number
# Items 1-35: column B, rows 10-44
# Items 36-70: column E, rows 10-44
# Items 71-105: column H, rows 10-44
# Items 106-140: column K, rows 10-44
# Items 141-175: column N, rows 10-44
# ============================================================
def get_cell(item_num):
    """Returns absolute cell reference for item response."""
    if 1 <= item_num <= 35:
        return f"$B${ITEM_START_ROW + 1 + (item_num - 1)}"
    elif 36 <= item_num <= 70:
        return f"$E${ITEM_START_ROW + 1 + (item_num - 36)}"
    elif 71 <= item_num <= 105:
        return f"$H${ITEM_START_ROW + 1 + (item_num - 71)}"
    elif 106 <= item_num <= 140:
        return f"$K${ITEM_START_ROW + 1 + (item_num - 106)}"
    elif 141 <= item_num <= 175:
        return f"$N${ITEM_START_ROW + 1 + (item_num - 141)}"
    return "$B$10"


# ============================================================
# SECTION 3: SCORING TABLE (starts after items)
# ============================================================
SCORE_START_ROW = ITEM_START_ROW + 1 + 35 + 2  # row 47

ws[f'A{SCORE_START_ROW}'] = "SCORING TABLE"
ws[f'A{SCORE_START_ROW}'].font = SECTION_FONT
ws[f'A{SCORE_START_ROW}'].fill = SECTION_FILL

# Validity check row
val_row = SCORE_START_ROW + 1
ws[f'A{val_row}'] = "VALIDITY:"
ws[f'A{val_row}'].font = Font(bold=True)

# V score formula (items 65, 110, 157 scored TRUE)
v_formula = f"={get_cell(65)}+{get_cell(110)}+{get_cell(157)}"
ws[f'B{val_row}'] = "V Score ="
ws[f'C{val_row}'] = v_formula
ws[f'C{val_row}'].fill = RESULT_FILL
ws[f'D{val_row}'] = f'=IF(C{val_row}>1,"INVALID (V>1)","Valid")'
ws[f'D{val_row}'].font = Font(bold=True)

# Scale X validity check will use raw score from scoring table
ws[f'F{val_row}'] = "Scale X validity checked in scoring table below"

# Scoring table headers
hdr_row = SCORE_START_ROW + 3
headers = ['Scale', 'Name', 'Category', 'Raw Score\n(auto)',
           'Initial BR\n(ENTER from\nAppendix C)', 'Disclosure\nAdj',
           'A/D Adj', 'Inpatient\nAdj', 'Denial/\nComplaint',
           'FINAL BR', 'Significance']

for col_idx, h in enumerate(headers):
    cell = ws.cell(row=hdr_row, column=col_idx + 1)
    cell.value = h
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True,
                               vertical='center')
    cell.border = THIN_BORDER

ws.row_dimensions[hdr_row].height = 40


# ============================================================
# BUILD RAW SCORE FORMULAS AND SCORING ROWS
# ============================================================
# We need to track where each scale's row is for cross-references
scale_rows = {}  # scale_code -> row number
data_start_row = hdr_row + 1

for idx, (code, name, category) in enumerate(SCALES):
    row = data_start_row + idx
    scale_rows[code] = row
    
    # Col A: Scale code
    ws.cell(row=row, column=1, value=code)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Col B: Scale name
    ws.cell(row=row, column=2, value=name)
    ws.cell(row=row, column=2).border = THIN_BORDER
    
    # Col C: Category
    ws.cell(row=row, column=3, value=category)
    ws.cell(row=row, column=3).border = THIN_BORDER
    
    # Col D: Raw Score (auto-calculated from item responses)
    true_items = SCORING_KEYS.get(code, {}).get('true', [])
    false_items = SCORING_KEYS.get(code, {}).get('false', [])
    
    parts = []
    for item in true_items:
        parts.append(get_cell(item))
    for item in false_items:
        parts.append(f"(1-{get_cell(item)})")
    
    if parts:
        raw_formula = "=" + "+".join(parts)
    else:
        raw_formula = "=0"
    
    # Scale 5: multiply by 2/3
    if code == '5':
        raw_formula = f"=ROUND(({raw_formula[1:]})*2/3,0)"
    
    ws.cell(row=row, column=4, value=raw_formula)
    ws.cell(row=row, column=4).fill = RESULT_FILL
    ws.cell(row=row, column=4).border = THIN_BORDER
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=4).font = Font(bold=True)


    # Col E: Initial BR Score - USER ENTERS from Appendix C
    ws.cell(row=row, column=5).fill = INPUT_FILL
    ws.cell(row=row, column=5).border = THIN_BORDER
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')

# ============================================================
# NOW SET UP ADJUSTMENT FORMULAS (need scale_rows populated)
# ============================================================
# Scales that use 1-8B disclosure adj:
PERS_SCALES = ['1','2A','2B','3','4','5','6A','6B','7','8A','8B']
# Scales using S-PP disclosure adj:
SPP_SCALES = ['S','C','P','A','H','N','D','B','T','R','SS','CC','PP']
# A/D adj scales (2A,S use AD_2AS; 2B,8B,C use AD_2B8BC)
AD_2AS_SCALES = ['2A', 'S']
AD_2B8BC_SCALES = ['2B', '8B', 'C']

# X raw score cell (for disclosure lookup)
x_raw_cell = f"D{scale_rows['X']}"

# A and D initial BR + disclosure adj cells (for A/D adjustment calc)
a_row = scale_rows['A']
d_row = scale_rows['D']
# After disclosure: E{row} + F{row}
a_adj = f"(E{a_row}+F{a_row})"
d_adj = f"(E{d_row}+F{d_row})"

for idx, (code, name, category) in enumerate(SCALES):
    row = data_start_row + idx


    # Col F: Disclosure Adjustment
    if code in ['X', 'Y', 'Z']:
        # No disclosure adj for modifying indices
        ws.cell(row=row, column=6, value=0)
        ws.cell(row=row, column=6).fill = LOCKED_FILL
    elif code in PERS_SCALES:
        # Use VLOOKUP on embedded lookup range (we'll put lookup in cols P-R)
        disc_formula = f'=IF({x_raw_cell}="",0,VLOOKUP(MIN(MAX({x_raw_cell},0),199),P:Q,2,TRUE))'
        ws.cell(row=row, column=6, value=disc_formula)
    else:
        # S-PP adjustment (column R in lookup)
        disc_formula = f'=IF({x_raw_cell}="",0,VLOOKUP(MIN(MAX({x_raw_cell},0),199),P:R,3,TRUE))'
        ws.cell(row=row, column=6, value=disc_formula)
    ws.cell(row=row, column=6).border = THIN_BORDER
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
    
    # Col G: A/D Adjustment
    if code in AD_2AS_SCALES:
        # Calculate: if both A and D adjusted BR < 75, adj = 0
        # Otherwise sum portions above 75, look up in AD_2AS table (cols S-T)
        ad_formula = (
            f'=IF(OR(E{a_row}="",E{d_row}=""),0,'
            f'IF(AND({a_adj}<75,{d_adj}<75),0,'
            f'VLOOKUP(MIN(MAX(IF({a_adj}>=75,{a_adj}-75,0)+IF({d_adj}>=75,{d_adj}-75,0),0),80),S:T,2,TRUE)))'
        )
        ws.cell(row=row, column=7, value=ad_formula)
    elif code in AD_2B8BC_SCALES:
        ad_formula = (
            f'=IF(OR(E{a_row}="",E{d_row}=""),0,'
            f'IF(AND({a_adj}<75,{d_adj}<75),0,'
            f'VLOOKUP(MIN(MAX(IF({a_adj}>=75,{a_adj}-75,0)+IF({d_adj}>=75,{d_adj}-75,0),0),80),U:V,2,TRUE)))'
        )
        ws.cell(row=row, column=7, value=ad_formula)
    else:
        ws.cell(row=row, column=7, value=0)
        ws.cell(row=row, column=7).fill = LOCKED_FILL
    ws.cell(row=row, column=7).border = THIN_BORDER
    ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')


    # Col H: Inpatient Adjustment
    # Only for SS, CC, PP - based on B6 (Y/N) and D6 (weeks)
    if code == 'SS':
        inp_formula = '=IF($B$6="Y",IF($D$6<1,6,IF($D$6<=4,4,0)),0)'
        ws.cell(row=row, column=8, value=inp_formula)
    elif code == 'CC':
        inp_formula = '=IF($B$6="Y",IF($D$6<1,10,IF($D$6<=4,8,0)),0)'
        ws.cell(row=row, column=8, value=inp_formula)
    elif code == 'PP':
        inp_formula = '=IF($B$6="Y",IF($D$6<1,4,IF($D$6<=4,2,0)),0)'
        ws.cell(row=row, column=8, value=inp_formula)
    else:
        ws.cell(row=row, column=8, value=0)
        ws.cell(row=row, column=8).fill = LOCKED_FILL
    ws.cell(row=row, column=8).border = THIN_BORDER
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')
    
    # Col I: Denial/Complaint Adjustment (manual entry for user)
    # User enters 8 if highest 1-8B scale is 4,5,7; else 0
    ws.cell(row=row, column=9, value=0)
    if code in PERS_SCALES:
        ws.cell(row=row, column=9).fill = INPUT_FILL
    else:
        ws.cell(row=row, column=9).fill = LOCKED_FILL
    ws.cell(row=row, column=9).border = THIN_BORDER
    ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
    
    # Col J: FINAL BR SCORE = Initial BR + all adjustments, clamped 0-115
    final_formula = (
        f'=IF(E{row}="","",MIN(MAX(E{row}+F{row}+G{row}+H{row}+I{row},0),115))'
    )
    ws.cell(row=row, column=10, value=final_formula)
    ws.cell(row=row, column=10).fill = RESULT_FILL
    ws.cell(row=row, column=10).border = THIN_BORDER
    ws.cell(row=row, column=10).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=10).font = Font(bold=True, size=12)
    
    # Col K: Clinical Significance
    sig_formula = (
        f'=IF(J{row}="","",IF(J{row}>=85,"PROMINENT",'
        f'IF(J{row}>=75,"PRESENT",IF(J{row}>=60,"Suggestive","Normal"))))'
    )
    ws.cell(row=row, column=11, value=sig_formula)
    ws.cell(row=row, column=11).border = THIN_BORDER
    ws.cell(row=row, column=11).alignment = Alignment(horizontal='center')


# ============================================================
# CONDITIONAL FORMATTING for Final BR column (J)
# ============================================================
first_data_row = data_start_row
last_data_row = data_start_row + len(SCALES) - 1
red_font = Font(bold=True, color="9C0006")
yellow_font = Font(bold=True, color="9C6500")
green_font = Font(color="006100")

range_str = f"J{first_data_row}:J{last_data_row}"
ws.conditional_formatting.add(range_str, CellIsRule(
    operator='greaterThanOrEqual', formula=['85'],
    fill=RED_FILL, font=red_font))
ws.conditional_formatting.add(range_str, CellIsRule(
    operator='between', formula=['75', '84'],
    fill=YELLOW_FILL_BG, font=yellow_font))
ws.conditional_formatting.add(range_str, CellIsRule(
    operator='lessThan', formula=['60'],
    fill=GREEN_FILL_BG, font=green_font))

# Also for significance column K
range_k = f"K{first_data_row}:K{last_data_row}"
ws.conditional_formatting.add(range_k, CellIsRule(
    operator='equal', formula=['"PROMINENT"'],
    fill=RED_FILL, font=red_font))
ws.conditional_formatting.add(range_k, CellIsRule(
    operator='equal', formula=['"PRESENT"'],
    fill=YELLOW_FILL_BG, font=yellow_font))

# ============================================================
# LOOKUP TABLES (embedded in columns P-V, hidden from view)
# Col P: Scale X raw score (0-199)
# Col Q: 1-8B Disclosure Adj
# Col R: S-PP Disclosure Adj
# Col S: A/D value (0-80)
# Col T: 2A, S Adjustment
# Col U: A/D value (0-80) [duplicate for VLOOKUP]
# Col V: 2B, 8B, C Adjustment
# ============================================================
# Header row for lookup tables
lookup_hdr_row = 1
ws.cell(row=lookup_hdr_row, column=16, value="X_Raw")
ws.cell(row=lookup_hdr_row, column=17, value="1-8B_Adj")
ws.cell(row=lookup_hdr_row, column=18, value="S-PP_Adj")
ws.cell(row=lookup_hdr_row, column=19, value="AD_Val")
ws.cell(row=lookup_hdr_row, column=20, value="2A_S_Adj")
ws.cell(row=lookup_hdr_row, column=21, value="AD_Val2")
ws.cell(row=lookup_hdr_row, column=22, value="2B_8B_C_Adj")

for i in range(200):
    r = lookup_hdr_row + 1 + i
    adj_18b, adj_spp = DISC_TABLE.get(i, (0, 0))
    ws.cell(row=r, column=16, value=i)
    ws.cell(row=r, column=17, value=adj_18b)
    ws.cell(row=r, column=18, value=adj_spp)

for i in range(81):
    r = lookup_hdr_row + 1 + i
    ws.cell(row=r, column=19, value=i)
    ws.cell(row=r, column=20, value=AD_2AS[i])
    ws.cell(row=r, column=21, value=i)
    ws.cell(row=r, column=22, value=AD_2B8BC[i])


# ============================================================
# SECTION 4: VALIDITY SUMMARY (after scoring table)
# ============================================================
validity_row = last_data_row + 2
ws.cell(row=validity_row, column=1, value="VALIDITY SUMMARY")
ws.cell(row=validity_row, column=1).font = SECTION_FONT
ws.cell(row=validity_row, column=1).fill = SECTION_FILL

ws.cell(row=validity_row+1, column=1, value="V Score (items 65,110,157):")
v_formula2 = f"={get_cell(65)}+{get_cell(110)}+{get_cell(157)}"
ws.cell(row=validity_row+1, column=3, value=v_formula2)
ws.cell(row=validity_row+1, column=3).fill = RESULT_FILL
ws.cell(row=validity_row+1, column=4,
        value=f'=IF(C{validity_row+1}>1,"INVALID","Valid")')

ws.cell(row=validity_row+2, column=1, value="Scale X Raw Score:")
ws.cell(row=validity_row+2, column=3, value=f"=D{scale_rows['X']}")
ws.cell(row=validity_row+2, column=3).fill = RESULT_FILL
ws.cell(row=validity_row+2, column=4,
        value=f'=IF(OR(C{validity_row+2}<34,C{validity_row+2}>178),"INVALID","Valid")')

ws.cell(row=validity_row+3, column=1, value="Protocol Status:")
ws.cell(row=validity_row+3, column=3,
        value=f'=IF(OR(C{validity_row+1}>1,C{validity_row+2}<34,C{validity_row+2}>178),"INVALID - DO NOT INTERPRET","VALID - Proceed")')
ws.cell(row=validity_row+3, column=3).font = Font(bold=True, size=12)

# ============================================================
# SECTION 5: INTERPRETATION KEY
# ============================================================
interp_row = validity_row + 5
ws.cell(row=interp_row, column=1, value="INTERPRETATION KEY")
ws.cell(row=interp_row, column=1).font = SECTION_FONT
ws.cell(row=interp_row, column=1).fill = SECTION_FILL

keys = [
    ("BR 85-115", "PROMINENT / PATHOLOGICAL", "Disorder present - primary focus"),
    ("BR 75-84", "PRESENT / TRAIT LEVEL", "Clinically significant features"),
    ("BR 60-74", "Suggestive", "Mild/subclinical features - monitor"),
    ("BR 0-59", "Normal", "Not clinically significant"),
]
for ki, (br_range, classif, meaning) in enumerate(keys):
    r = interp_row + 1 + ki
    ws.cell(row=r, column=1, value=br_range)
    ws.cell(row=r, column=2, value=classif)
    ws.cell(row=r, column=3, value=meaning)
    ws.cell(row=r, column=1).font = Font(bold=True)


# ============================================================
# SECTION 6: ADJUSTMENT INSTRUCTIONS
# ============================================================
adj_row = interp_row + 7
ws.cell(row=adj_row, column=1, value="ADJUSTMENT RULES")
ws.cell(row=adj_row, column=1).font = SECTION_FONT
ws.cell(row=adj_row, column=1).fill = SECTION_FILL

adj_text = [
    "1. Disclosure Adj (Col F): Auto-calculated from Scale X raw score",
    "   - Scales 1-8B: 1-8B adjustment | Scales S-PP: S-PP adjustment",
    "2. A/D Adj (Col G): Auto-calculated when Anxiety or Dysthymia BR+Disc >= 75",
    "   - Applies to 2A, 2B, 8B, S, C only",
    "3. Inpatient Adj (Col H): Auto from Inpatient Y/N and duration",
    "   - SS: +6(<1wk)/+4(1-4wk) | CC: +10/+8 | PP: +4/+2",
    "4. Denial/Complaint (Col I): MANUAL - enter 8 if highest 1-8B is scale 4,5,7",
    "",
    "HOW TO USE:",
    "1. Enter 1 or 0 for all 175 items in the yellow cells above",
    "2. Raw scores calculate automatically (Column D in scoring table)",
    "3. Look up each raw score in MCMI-III Manual Appendix C",
    "4. Enter the BR score in column E (yellow cells)",
    "5. All adjustments auto-calculate; Final BR appears in column J",
    "6. Check column K for clinical significance",
]
for ai, line in enumerate(adj_text):
    ws.cell(row=adj_row + 1 + ai, column=1, value=line)

# ============================================================
# COLUMN WIDTHS
# ============================================================
col_widths = {
    'A': 10, 'B': 8, 'C': 16, 'D': 8, 'E': 8, 'F': 8,
    'G': 8, 'H': 8, 'I': 5, 'J': 8, 'K': 8,
    'L': 5, 'M': 8, 'N': 8
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Scoring table column widths (override)
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 9
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 11
ws.column_dimensions['K'].width = 14

# Hide lookup columns P-V
for col in ['P','Q','R','S','T','U','V']:
    ws.column_dimensions[col].hidden = True


# ============================================================
# SAVE
# ============================================================
output = "/projects/sandbox/Dango-kiro/MCMI-III_Scoring_Tool_V2.xlsx"
wb.save(output)
print(f"SUCCESS! File saved: {output}")
print(f"File size: {__import__('os').path.getsize(output)} bytes")
print(f"\nLayout (single sheet):")
print(f"  Rows 1-6: Patient info")
print(f"  Rows {ITEM_START_ROW}-{ITEM_START_ROW+36}: Item responses (175 items in 5 columns)")
print(f"  Row {SCORE_START_ROW}: Scoring table header")
print(f"  Rows {data_start_row}-{last_data_row}: All 27 scales with formulas")
print(f"  Rows {validity_row}-{validity_row+3}: Validity summary")
print(f"  Rows {interp_row}-{interp_row+5}: Interpretation key")
print(f"  Rows {adj_row}-{adj_row+15}: Instructions")
print(f"  Cols P-V (hidden): Lookup tables for VLOOKUP formulas")
print(f"\nScales: {len(SCALES)}")
print(f"Lookup table entries: {len(DISC_TABLE)} disclosure, {len(AD_2AS)} AD_2AS, {len(AD_2B8BC)} AD_2B8BC")
print(f"\nScale row mapping:")
for code, r in scale_rows.items():
    print(f"  {code}: row {r}")
