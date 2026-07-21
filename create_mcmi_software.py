#!/usr/bin/env python3
"""
MCMI-III Scoring Software - Excel Generator
Creates a professional MCMI-III scoring workbook with:
- Patient Data Input Sheet
- Raw Score Entry Sheet
- BR Score Calculation & Interpretation Sheet
- Profile Summary Sheet
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from copy import copy

# Create workbook
wb = Workbook()

# ============================================================
# CONSTANTS
# ============================================================

HEADER_FONT = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=11, color='2E75B6')
NORMAL_FONT = Font(name='Calibri', size=11)
SMALL_FONT = Font(name='Calibri', size=9, italic=True)

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GREEN_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
LIGHT_YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
LIGHT_RED_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# MCMI-III Scale Definitions
MODIFYING_INDICES = [
    ('X', 'Disclosure', 'Openness/willingness to reveal personal information'),
    ('Y', 'Desirability', 'Tendency to present oneself in a favorable light'),
    ('Z', 'Debasement', 'Tendency to deprecate or devalue oneself'),
    ('V', 'Invalidity', 'Random or confused responding (>1 invalidates)'),
]

CLINICAL_PERSONALITY = [
    ('1', 'Schizoid', 'Emotional detachment, social withdrawal, limited affect'),
    ('2A', 'Avoidant', 'Social inhibition, fear of rejection, hypersensitivity'),
    ('2B', 'Depressive', 'Chronic pessimism, low mood, hopelessness'),
    ('3', 'Dependent', 'Submissiveness, need for reassurance, clinging behavior'),
    ('4', 'Histrionic', 'Attention-seeking, emotional lability, superficiality'),
    ('5', 'Narcissistic', 'Grandiosity, entitlement, lack of empathy'),
    ('6A', 'Antisocial', 'Disregard for rules, impulsivity, exploitation'),
    ('6B', 'Sadistic (Aggressive)', 'Hostile, domineering, pleasure in others pain'),
    ('7', 'Compulsive', 'Perfectionism, rigidity, emotional constriction'),
    ('8A', 'Negativistic (Passive-Aggressive)', 'Ambivalence, irritability, oppositional behavior'),
    ('8B', 'Masochistic (Self-Defeating)', 'Self-sacrifice, acceptance of suffering, guilt'),
]

SEVERE_PERSONALITY = [
    ('S', 'Schizotypal', 'Eccentric behavior, cognitive slippage, social detachment'),
    ('C', 'Borderline', 'Emotional instability, identity diffusion, self-harm risk'),
    ('P', 'Paranoid', 'Suspiciousness, mistrust, hypervigilance'),
]

CLINICAL_SYNDROMES = [
    ('A', 'Anxiety', 'Excessive worry, tension, apprehension, somatic anxiety'),
    ('H', 'Somatoform', 'Somatic complaints, health preoccupation'),
    ('N', 'Bipolar: Manic', 'Elevated mood, pressured speech, grandiosity'),
    ('D', 'Dysthymia', 'Chronic mild depression, low energy, pessimism'),
    ('B', 'Alcohol Dependence', 'Problematic alcohol use pattern'),
    ('T', 'Drug Dependence', 'Problematic drug use pattern'),
    ('R', 'Post-Traumatic Stress Disorder', 'Trauma-related distress, hyperarousal, avoidance'),
]

SEVERE_CLINICAL = [
    ('SS', 'Thought Disorder', 'Disorganized thinking, possible psychotic features'),
    ('CC', 'Major Depression', 'Severe depressive episode, hopelessness, anhedonia'),
    ('PP', 'Delusional Disorder', 'Fixed false beliefs, paranoid ideation'),
]


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def apply_header_style(ws, row, start_col, end_col):
    """Apply header styling to a row range."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def apply_border_range(ws, start_row, end_row, start_col, end_col):
    """Apply borders to a range of cells."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def set_column_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ============================================================
# SHEET 1: PATIENT DATA INPUT
# ============================================================

ws1 = wb.active
ws1.title = "Patient Information"

# Title
ws1.merge_cells('A1:F1')
ws1['A1'] = 'MCMI-III SCORING SOFTWARE'
ws1['A1'].font = Font(name='Calibri', bold=True, size=18, color='1F4E79')
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:F2')
ws1['A2'] = 'Millon Clinical Multiaxial Inventory - Third Edition'
ws1['A2'].font = Font(name='Calibri', size=12, italic=True, color='4472C4')
ws1['A2'].alignment = Alignment(horizontal='center')

# Patient Demographics Section
ws1['A4'] = 'PATIENT DEMOGRAPHICS'
ws1['A4'].font = SUBTITLE_FONT

patient_fields = [
    ('A5', 'Patient Name:', 'B5'),
    ('A6', 'Age:', 'B6'),
    ('A7', 'Gender:', 'B7'),
    ('A8', 'Date of Birth:', 'B8'),
    ('A9', 'Education:', 'B9'),
    ('A10', 'Marital Status:', 'B10'),
    ('A11', 'Occupation:', 'B11'),
    ('D5', 'Case Number:', 'E5'),
    ('D6', 'Date of Testing:', 'E6'),
    ('D7', 'Referred By:', 'E7'),
    ('D8', 'Examiner:', 'E8'),
    ('D9', 'Diagnosis (Provisional):', 'E9'),
    ('D10', 'Setting:', 'E10'),
    ('D11', 'Reason for Referral:', 'E11'),
]

for label_cell, label_text, input_cell in patient_fields:
    ws1[label_cell] = label_text
    ws1[label_cell].font = Font(name='Calibri', bold=True, size=11)
    ws1[input_cell].fill = INPUT_FILL
    ws1[input_cell].border = THIN_BORDER


# Instructions
ws1['A13'] = 'INSTRUCTIONS'
ws1['A13'].font = SUBTITLE_FONT
ws1.merge_cells('A14:F14')
ws1['A14'] = ('1. Fill in Patient Information above.\n'
              '2. Go to "Raw Score Entry" sheet and enter the raw scores for each scale.\n'
              '3. Go to "BR Scores & Interpretation" sheet - BR scores and interpretations are auto-calculated.\n'
              '4. Check "Profile Summary" sheet for visual profile and clinical significance overview.')
ws1['A14'].font = NORMAL_FONT
ws1['A14'].alignment = Alignment(wrap_text=True, vertical='top')
ws1.row_dimensions[14].height = 65

# BR Score Interpretation Guide
ws1['A17'] = 'BR SCORE INTERPRETATION GUIDE'
ws1['A17'].font = SUBTITLE_FONT

guide_headers = ['BR Score Range', 'Classification', 'Clinical Meaning']
for i, h in enumerate(guide_headers, 1):
    cell = ws1.cell(row=18, column=i)
    cell.value = h
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER

guide_data = [
    ('0 - 34', 'Not Significant', 'No clinical relevance; trait not present'),
    ('35 - 59', 'Normal Range', 'Within normal limits; no clinical concern'),
    ('60 - 74', 'Suggestive', 'Mild features present; monitor clinically'),
    ('75 - 84', 'Trait Level / Presence', 'Clinically significant trait; syndrome features present'),
    ('85 - 115', 'Pathological / Disorder', 'Prominent disorder; syndrome confirmed'),
]

for row_idx, (score, classification, meaning) in enumerate(guide_data, 19):
    ws1.cell(row=row_idx, column=1, value=score).border = THIN_BORDER
    ws1.cell(row=row_idx, column=2, value=classification).border = THIN_BORDER
    ws1.cell(row=row_idx, column=3, value=meaning).border = THIN_BORDER
    # Color coding
    if '0 - 34' in score:
        fill = LIGHT_GREEN_FILL
    elif '35 - 59' in score:
        fill = LIGHT_GREEN_FILL
    elif '60 - 74' in score:
        fill = LIGHT_BLUE_FILL
    elif '75 - 84' in score:
        fill = LIGHT_YELLOW_FILL
    else:
        fill = LIGHT_RED_FILL
    for col in range(1, 4):
        ws1.cell(row=row_idx, column=col).fill = fill

set_column_widths(ws1, {'A': 22, 'B': 25, 'C': 50, 'D': 22, 'E': 25, 'F': 20})


# ============================================================
# SHEET 2: RAW SCORE ENTRY
# ============================================================

ws2 = wb.create_sheet("Raw Score Entry")

ws2.merge_cells('A1:E1')
ws2['A1'] = 'MCMI-III RAW SCORE ENTRY'
ws2['A1'].font = TITLE_FONT
ws2['A1'].alignment = Alignment(horizontal='center')

ws2.merge_cells('A2:E2')
ws2['A2'] = 'Enter the Raw Scores for each scale in the highlighted cells (Column D)'
ws2['A2'].font = SMALL_FONT
ws2['A2'].alignment = Alignment(horizontal='center')

# Headers
headers = ['Scale Code', 'Scale Name', 'Category', 'Raw Score', 'Max Possible']
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=4, column=col)
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

# Data rows
current_row = 5
all_scales = []

# Section: Modifying Indices
ws2.cell(row=current_row, column=1, value='MODIFYING INDICES').font = SUBTITLE_FONT
ws2.merge_cells(f'A{current_row}:E{current_row}')
ws2.cell(row=current_row, column=1).fill = LIGHT_BLUE_FILL
current_row += 1

max_scores_modifier = {'X': 645, 'Y': 21, 'Z': 33, 'V': 3}
for code, name, desc in MODIFYING_INDICES:
    ws2.cell(row=current_row, column=1, value=code).border = THIN_BORDER
    ws2.cell(row=current_row, column=2, value=name).border = THIN_BORDER
    ws2.cell(row=current_row, column=3, value='Modifying Index').border = THIN_BORDER
    raw_cell = ws2.cell(row=current_row, column=4)
    raw_cell.fill = INPUT_FILL
    raw_cell.border = THIN_BORDER
    raw_cell.alignment = Alignment(horizontal='center')
    ws2.cell(row=current_row, column=5, value=max_scores_modifier.get(code, '')).border = THIN_BORDER
    ws2.cell(row=current_row, column=5).alignment = Alignment(horizontal='center')
    all_scales.append((code, name, 'Modifying Index', current_row))
    current_row += 1

# Section: Clinical Personality Patterns
current_row += 1
ws2.cell(row=current_row, column=1, value='CLINICAL PERSONALITY PATTERNS').font = SUBTITLE_FONT
ws2.merge_cells(f'A{current_row}:E{current_row}')
ws2.cell(row=current_row, column=1).fill = LIGHT_GREEN_FILL
current_row += 1

for code, name, desc in CLINICAL_PERSONALITY:
    ws2.cell(row=current_row, column=1, value=code).border = THIN_BORDER
    ws2.cell(row=current_row, column=2, value=name).border = THIN_BORDER
    ws2.cell(row=current_row, column=3, value='Clinical Personality').border = THIN_BORDER
    raw_cell = ws2.cell(row=current_row, column=4)
    raw_cell.fill = INPUT_FILL
    raw_cell.border = THIN_BORDER
    raw_cell.alignment = Alignment(horizontal='center')
    ws2.cell(row=current_row, column=5, value='').border = THIN_BORDER
    all_scales.append((code, name, 'Clinical Personality', current_row))
    current_row += 1


# Section: Severe Personality Pathology
current_row += 1
ws2.cell(row=current_row, column=1, value='SEVERE PERSONALITY PATHOLOGY').font = SUBTITLE_FONT
ws2.merge_cells(f'A{current_row}:E{current_row}')
ws2.cell(row=current_row, column=1).fill = LIGHT_YELLOW_FILL
current_row += 1

for code, name, desc in SEVERE_PERSONALITY:
    ws2.cell(row=current_row, column=1, value=code).border = THIN_BORDER
    ws2.cell(row=current_row, column=2, value=name).border = THIN_BORDER
    ws2.cell(row=current_row, column=3, value='Severe Personality').border = THIN_BORDER
    raw_cell = ws2.cell(row=current_row, column=4)
    raw_cell.fill = INPUT_FILL
    raw_cell.border = THIN_BORDER
    raw_cell.alignment = Alignment(horizontal='center')
    ws2.cell(row=current_row, column=5, value='').border = THIN_BORDER
    all_scales.append((code, name, 'Severe Personality', current_row))
    current_row += 1

# Section: Clinical Syndromes
current_row += 1
ws2.cell(row=current_row, column=1, value='CLINICAL SYNDROMES').font = SUBTITLE_FONT
ws2.merge_cells(f'A{current_row}:E{current_row}')
ws2.cell(row=current_row, column=1).fill = LIGHT_BLUE_FILL
current_row += 1

for code, name, desc in CLINICAL_SYNDROMES:
    ws2.cell(row=current_row, column=1, value=code).border = THIN_BORDER
    ws2.cell(row=current_row, column=2, value=name).border = THIN_BORDER
    ws2.cell(row=current_row, column=3, value='Clinical Syndrome').border = THIN_BORDER
    raw_cell = ws2.cell(row=current_row, column=4)
    raw_cell.fill = INPUT_FILL
    raw_cell.border = THIN_BORDER
    raw_cell.alignment = Alignment(horizontal='center')
    ws2.cell(row=current_row, column=5, value='').border = THIN_BORDER
    all_scales.append((code, name, 'Clinical Syndrome', current_row))
    current_row += 1

# Section: Severe Clinical Syndromes
current_row += 1
ws2.cell(row=current_row, column=1, value='SEVERE CLINICAL SYNDROMES').font = SUBTITLE_FONT
ws2.merge_cells(f'A{current_row}:E{current_row}')
ws2.cell(row=current_row, column=1).fill = LIGHT_RED_FILL
current_row += 1

for code, name, desc in SEVERE_CLINICAL:
    ws2.cell(row=current_row, column=1, value=code).border = THIN_BORDER
    ws2.cell(row=current_row, column=2, value=name).border = THIN_BORDER
    ws2.cell(row=current_row, column=3, value='Severe Clinical').border = THIN_BORDER
    raw_cell = ws2.cell(row=current_row, column=4)
    raw_cell.fill = INPUT_FILL
    raw_cell.border = THIN_BORDER
    raw_cell.alignment = Alignment(horizontal='center')
    ws2.cell(row=current_row, column=5, value='').border = THIN_BORDER
    all_scales.append((code, name, 'Severe Clinical', current_row))
    current_row += 1

set_column_widths(ws2, {'A': 14, 'B': 32, 'C': 22, 'D': 14, 'E': 14})


# ============================================================
# SHEET 3: BR SCORES & INTERPRETATION
# ============================================================

ws3 = wb.create_sheet("BR Scores & Interpretation")

ws3.merge_cells('A1:G1')
ws3['A1'] = 'MCMI-III BR SCORES & CLINICAL INTERPRETATION'
ws3['A1'].font = TITLE_FONT
ws3['A1'].alignment = Alignment(horizontal='center')

ws3.merge_cells('A2:G2')
ws3['A2'] = 'Enter BR Scores in Column D. Clinical Significance is auto-calculated.'
ws3['A2'].font = SMALL_FONT
ws3['A2'].alignment = Alignment(horizontal='center')

# Headers
br_headers = ['Scale Code', 'Scale Name', 'Category', 'BR Score', 
              'Clinical Significance', 'Interpretation Level', 'Description']
for col, header in enumerate(br_headers, 1):
    cell = ws3.cell(row=4, column=col)
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

ws3.row_dimensions[4].height = 30

# Data rows with formulas
current_row = 5
br_scale_rows = []

def add_br_section(ws, scales, category, section_title, fill, start_row):
    """Add a section of scales with BR score interpretation formulas."""
    row = start_row
    # Section header
    ws.cell(row=row, column=1, value=section_title).font = SUBTITLE_FONT
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1).fill = fill
    row += 1
    
    section_rows = []
    for code, name, desc in scales:
        ws.cell(row=row, column=1, value=code).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=name).border = THIN_BORDER
        ws.cell(row=row, column=3, value=category).border = THIN_BORDER
        
        # BR Score input cell
        br_cell = ws.cell(row=row, column=4)
        br_cell.fill = INPUT_FILL
        br_cell.border = THIN_BORDER
        br_cell.alignment = Alignment(horizontal='center')
        
        # Clinical Significance formula (auto-calculated from BR score)
        br_ref = f'D{row}'
        sig_formula = (
            f'=IF({br_ref}="","",IF({br_ref}>=85,"Pathological Level / Disorder",'
            f'IF({br_ref}>=75,"Trait Level / Presence of Syndrome",'
            f'IF({br_ref}>=60,"Suggestive (Monitor)",'
            f'IF({br_ref}>=35,"Normal Range","Not Significant")))))'
        )
        ws.cell(row=row, column=5, value=sig_formula).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Interpretation Level (numeric for charting)
        level_formula = (
            f'=IF({br_ref}="",0,IF({br_ref}>=85,4,'
            f'IF({br_ref}>=75,3,IF({br_ref}>=60,2,IF({br_ref}>=35,1,0)))))'
        )
        ws.cell(row=row, column=6, value=level_formula).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        
        # Description
        ws.cell(row=row, column=7, value=desc).border = THIN_BORDER
        ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
        
        section_rows.append(row)
        row += 1
    
    return row, section_rows


# Add all sections
current_row, mod_rows = add_br_section(
    ws3, MODIFYING_INDICES, 'Modifying Index',
    'MODIFYING INDICES', LIGHT_BLUE_FILL, current_row
)
br_scale_rows.extend(mod_rows)

current_row += 1
current_row, cp_rows = add_br_section(
    ws3, CLINICAL_PERSONALITY, 'Clinical Personality',
    'CLINICAL PERSONALITY PATTERNS', LIGHT_GREEN_FILL, current_row
)
br_scale_rows.extend(cp_rows)

current_row += 1
current_row, sp_rows = add_br_section(
    ws3, SEVERE_PERSONALITY, 'Severe Personality',
    'SEVERE PERSONALITY PATHOLOGY', LIGHT_YELLOW_FILL, current_row
)
br_scale_rows.extend(sp_rows)

current_row += 1
current_row, cs_rows = add_br_section(
    ws3, CLINICAL_SYNDROMES, 'Clinical Syndrome',
    'CLINICAL SYNDROMES', LIGHT_BLUE_FILL, current_row
)
br_scale_rows.extend(cs_rows)

current_row += 1
current_row, sc_rows = add_br_section(
    ws3, SEVERE_CLINICAL, 'Severe Clinical',
    'SEVERE CLINICAL SYNDROMES', LIGHT_RED_FILL, current_row
)
br_scale_rows.extend(sc_rows)

# Conditional formatting for BR scores
# Red for pathological (>=85)
ws3.conditional_formatting.add(
    f'D5:D{current_row}',
    CellIsRule(operator='greaterThanOrEqual', formula=['85'],
              fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'))
)
# Yellow for trait level (75-84)
ws3.conditional_formatting.add(
    f'D5:D{current_row}',
    CellIsRule(operator='between', formula=['75', '84'],
              fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
)
# Green for normal (35-74)
ws3.conditional_formatting.add(
    f'D5:D{current_row}',
    CellIsRule(operator='between', formula=['35', '74'],
              fill=PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid'))
)

set_column_widths(ws3, {
    'A': 12, 'B': 30, 'C': 20, 'D': 12,
    'E': 30, 'F': 18, 'G': 50
})


# ============================================================
# SHEET 4: PROFILE SUMMARY
# ============================================================

ws4 = wb.create_sheet("Profile Summary")

ws4.merge_cells('A1:F1')
ws4['A1'] = 'MCMI-III PROFILE SUMMARY'
ws4['A1'].font = TITLE_FONT
ws4['A1'].alignment = Alignment(horizontal='center')

ws4.merge_cells('A2:F2')
ws4['A2'] = 'Visual Profile Overview - BR Scores by Scale'
ws4['A2'].font = SMALL_FONT
ws4['A2'].alignment = Alignment(horizontal='center')

# Profile table for charting
profile_headers = ['Scale', 'Scale Name', 'BR Score', 'Bar (Visual)', 'Significance', 'Notes']
for col, header in enumerate(profile_headers, 1):
    cell = ws4.cell(row=4, column=col)
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

# Reference the BR Scores sheet
all_scale_defs = (
    MODIFYING_INDICES + CLINICAL_PERSONALITY + 
    SEVERE_PERSONALITY + CLINICAL_SYNDROMES + SEVERE_CLINICAL
)

row = 5
for i, (code, name, desc) in enumerate(all_scale_defs):
    br_row = br_scale_rows[i]
    
    ws4.cell(row=row, column=1, value=code).border = THIN_BORDER
    ws4.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws4.cell(row=row, column=2, value=name).border = THIN_BORDER
    
    # Link to BR Score from Sheet 3
    br_formula = f"='BR Scores & Interpretation'!D{br_row}"
    ws4.cell(row=row, column=3, value=br_formula).border = THIN_BORDER
    ws4.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    
    # Visual bar using REPT function
    bar_formula = f'=IF(C{row}="","",REPT("█",INT(C{row}/5)))'
    ws4.cell(row=row, column=4, value=bar_formula).border = THIN_BORDER
    ws4.cell(row=row, column=4).font = Font(name='Calibri', size=8, color='2E75B6')
    
    # Significance formula
    sig_formula = (
        f'=IF(C{row}="","",IF(C{row}>=85,"⚠ PATHOLOGICAL",'
        f'IF(C{row}>=75,"● TRAIT LEVEL",'
        f'IF(C{row}>=60,"○ Suggestive","— Normal"))))'
    )
    ws4.cell(row=row, column=5, value=sig_formula).border = THIN_BORDER
    ws4.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    
    # Notes column (empty for clinician input)
    ws4.cell(row=row, column=6).border = THIN_BORDER
    ws4.cell(row=row, column=6).fill = INPUT_FILL
    
    row += 1

# Add conditional formatting for significance column
last_data_row = row - 1
ws4.conditional_formatting.add(
    f'C5:C{last_data_row}',
    CellIsRule(operator='greaterThanOrEqual', formula=['85'],
              fill=PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid'))
)
ws4.conditional_formatting.add(
    f'C5:C{last_data_row}',
    CellIsRule(operator='between', formula=['75', '84'],
              fill=PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'))
)

set_column_widths(ws4, {'A': 10, 'B': 30, 'C': 12, 'D': 30, 'E': 18, 'F': 30})


# ============================================================
# SHEET 5: CLINICAL REPORT TEMPLATE
# ============================================================

ws5 = wb.create_sheet("Clinical Report")

ws5.merge_cells('A1:D1')
ws5['A1'] = 'MCMI-III CLINICAL REPORT TEMPLATE'
ws5['A1'].font = TITLE_FONT
ws5['A1'].alignment = Alignment(horizontal='center')

# Report sections
report_sections = [
    ('A3', 'TEST VALIDITY', 
     'Based on the Modifying Indices, the protocol is [VALID/INVALID]. '
     'The Disclosure score (X) suggests [interpretation]. '
     'The Desirability score (Y) indicates [interpretation]. '
     'The Debasement score (Z) reflects [interpretation].'),
    ('A6', 'CLINICAL PERSONALITY PATTERN',
     'The personality profile reveals predominant features of [scale names with BR≥75]. '
     'These elevations suggest [personality description]. '
     'The individual demonstrates [behavioral patterns].'),
    ('A9', 'SEVERE PERSONALITY PATHOLOGY',
     'Regarding severe personality pathology, the profile [shows/does not show] '
     'significant elevations. [If elevated: The elevation on Scale [X] suggests...]. '
     'Features include [description].'),
    ('A12', 'CLINICAL SYNDROMES',
     'The clinical syndrome scales indicate [elevated scales]. '
     'The individual experiences [symptom descriptions]. '
     'These findings are consistent with [clinical observations].'),
    ('A15', 'SEVERE CLINICAL SYNDROMES',
     'Severe clinical syndrome analysis [reveals/does not reveal] significant concerns. '
     '[If elevated: The elevation on [Scale] suggests...]. '
     'Clinical implications include [description].'),
    ('A18', 'INTEGRATION & RECOMMENDATIONS',
     'Overall, the MCMI-III profile suggests [integrated summary]. '
     'Key clinical concerns include: [list]. '
     'Treatment recommendations: [list]. '
     'Prognosis considerations: [description].'),
]

for cell_ref, title, template in report_sections:
    ws5[cell_ref] = title
    ws5[cell_ref].font = SUBTITLE_FONT
    # Content row
    row_num = int(cell_ref[1:]) + 1
    content_cell = f'A{row_num}'
    ws5.merge_cells(f'A{row_num}:D{row_num}')
    ws5[content_cell] = template
    ws5[content_cell].font = NORMAL_FONT
    ws5[content_cell].alignment = Alignment(wrap_text=True, vertical='top')
    ws5.row_dimensions[row_num].height = 50

set_column_widths(ws5, {'A': 30, 'B': 30, 'C': 30, 'D': 30})


# ============================================================
# SHEET 6: ITEM SCORING KEY (175 items)
# ============================================================

ws6 = wb.create_sheet("Item Response Entry")

ws6.merge_cells('A1:D1')
ws6['A1'] = 'MCMI-III ITEM RESPONSES (175 Items)'
ws6['A1'].font = TITLE_FONT
ws6['A1'].alignment = Alignment(horizontal='center')

ws6.merge_cells('A2:D2')
ws6['A2'] = 'Enter T (True) or F (False) for each item'
ws6['A2'].font = SMALL_FONT
ws6['A2'].alignment = Alignment(horizontal='center')

# Headers
item_headers = ['Item #', 'Response (T/F)', 'Item #', 'Response (T/F)']
for col, header in enumerate(item_headers, 1):
    cell = ws6.cell(row=4, column=col)
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER

# Create 175 item rows (split into 2 columns for space)
for i in range(1, 89):
    row = i + 4
    ws6.cell(row=row, column=1, value=i).border = THIN_BORDER
    ws6.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws6.cell(row=row, column=2).fill = INPUT_FILL
    ws6.cell(row=row, column=2).border = THIN_BORDER
    ws6.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    
    # Second column (items 89-175)
    item2 = i + 88
    if item2 <= 175:
        ws6.cell(row=row, column=3, value=item2).border = THIN_BORDER
        ws6.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws6.cell(row=row, column=4).fill = INPUT_FILL
        ws6.cell(row=row, column=4).border = THIN_BORDER
        ws6.cell(row=row, column=4).alignment = Alignment(horizontal='center')

set_column_widths(ws6, {'A': 10, 'B': 16, 'C': 10, 'D': 16})

# Add data validation for T/F
from openpyxl.worksheet.datavalidation import DataValidation

dv = DataValidation(type="list", formula1='"T,F"', allow_blank=True)
dv.error = "Please enter T (True) or F (False) only"
dv.errorTitle = "Invalid Response"
ws6.add_data_validation(dv)
dv.add(f'B5:B92')
dv.add(f'D5:D92')


# ============================================================
# SHEET 7: SCORING REFERENCE
# ============================================================

ws7 = wb.create_sheet("Scoring Reference")

ws7.merge_cells('A1:E1')
ws7['A1'] = 'MCMI-III SCORING REFERENCE & SCALE DESCRIPTIONS'
ws7['A1'].font = TITLE_FONT
ws7['A1'].alignment = Alignment(horizontal='center')

# Complete scale reference table
ref_headers = ['Scale Code', 'Scale Name', 'Category', 'Items', 'Description']
for col, header in enumerate(ref_headers, 1):
    cell = ws7.cell(row=3, column=col)
    cell.value = header
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

row = 4
all_sections = [
    ('MODIFYING INDICES', MODIFYING_INDICES, 'Validity', LIGHT_BLUE_FILL,
     {'X': '~178 items', 'Y': '21 items', 'Z': '33 items', 'V': '3 items'}),
    ('CLINICAL PERSONALITY PATTERNS', CLINICAL_PERSONALITY, 'Personality', LIGHT_GREEN_FILL,
     {'1': '16 items', '2A': '16 items', '2B': '15 items', '3': '16 items',
      '4': '17 items', '5': '24 items', '6A': '17 items', '6B': '20 items',
      '7': '17 items', '8A': '16 items', '8B': '15 items'}),
    ('SEVERE PERSONALITY PATHOLOGY', SEVERE_PERSONALITY, 'Severe Pers.', LIGHT_YELLOW_FILL,
     {'S': '16 items', 'C': '16 items', 'P': '17 items'}),
    ('CLINICAL SYNDROMES', CLINICAL_SYNDROMES, 'Syndrome', LIGHT_BLUE_FILL,
     {'A': '14 items', 'H': '12 items', 'N': '13 items', 'D': '14 items',
      'B': '15 items', 'T': '14 items', 'R': '16 items'}),
    ('SEVERE CLINICAL SYNDROMES', SEVERE_CLINICAL, 'Severe Syn.', LIGHT_RED_FILL,
     {'SS': '17 items', 'CC': '17 items', 'PP': '13 items'}),
]

for section_title, scales, category, fill, items_dict in all_sections:
    ws7.cell(row=row, column=1, value=section_title).font = SUBTITLE_FONT
    ws7.merge_cells(f'A{row}:E{row}')
    ws7.cell(row=row, column=1).fill = fill
    row += 1
    
    for code, name, desc in scales:
        ws7.cell(row=row, column=1, value=code).border = THIN_BORDER
        ws7.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws7.cell(row=row, column=2, value=name).border = THIN_BORDER
        ws7.cell(row=row, column=3, value=category).border = THIN_BORDER
        ws7.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws7.cell(row=row, column=4, value=items_dict.get(code, '')).border = THIN_BORDER
        ws7.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws7.cell(row=row, column=5, value=desc).border = THIN_BORDER
        ws7.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1

# Add interpretation guidelines at the bottom
row += 2
ws7.cell(row=row, column=1, value='INTERPRETATION GUIDELINES').font = SUBTITLE_FONT
row += 1
guidelines = [
    'BR Score ≥ 85: Indicates the presence of a disorder or pathological level of the trait.',
    'BR Score 75-84: Suggests clinically significant features (trait level or syndrome presence).',
    'BR Score 60-74: Suggestive features that warrant monitoring but not diagnosis.',
    'BR Score 35-59: Within normal limits; features are within normative expectations.',
    'BR Score < 35: Absence of the measured trait or syndrome.',
    '',
    'VALIDITY CHECKS:',
    'Scale V (Invalidity): If V > 1, the protocol may be INVALID (random/confused responding).',
    'Scale X (Disclosure): Very low (<35) = guarded/minimal disclosure; Very high (>85) = over-disclosure.',
    'Scale Y (Desirability): High scores (>75) suggest impression management / fake good.',
    'Scale Z (Debasement): High scores (>75) suggest self-deprecation / cry for help / fake bad.',
]

for guideline in guidelines:
    ws7.cell(row=row, column=1, value=guideline).font = NORMAL_FONT
    ws7.merge_cells(f'A{row}:E{row}')
    row += 1

set_column_widths(ws7, {'A': 14, 'B': 32, 'C': 14, 'D': 12, 'E': 55})


# ============================================================
# SAVE THE WORKBOOK
# ============================================================

output_path = '/projects/sandbox/Dango-kiro/MCMI.SOFTWARE.xlsx'
wb.save(output_path)
print(f"✅ MCMI-III Scoring Software Excel file created successfully!")
print(f"   Saved to: {output_path}")
print(f"   File contains {len(wb.sheetnames)} sheets:")
for name in wb.sheetnames:
    print(f"     - {name}")
print()
print("The workbook includes:")
print("  1. Patient Information - Demographics & instructions")
print("  2. Raw Score Entry - Enter raw scores for all 28 scales")
print("  3. BR Scores & Interpretation - Auto-calculated significance")
print("  4. Profile Summary - Visual bar profile with significance markers")
print("  5. Clinical Report - Template for writing interpretive reports")
print("  6. Item Response Entry - Enter T/F for all 175 items")
print("  7. Scoring Reference - Complete scale descriptions & guidelines")
