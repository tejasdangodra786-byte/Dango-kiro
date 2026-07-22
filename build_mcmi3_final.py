#!/usr/bin/env python3
"""
MCMI-III COMPLETE SCORING TOOL - FINAL CORRECT VERSION
Based on Appendix B (Scale Composition & Item Weighting)
Items have WEIGHTS: Prototypal=2, Nonprototypal=1, False=1
All in ONE sheet. BR tables from Appendix C.1 and C.2.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()
ws = wb.active
ws.title = "MCMI-III Scoring"

# Styles
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
HDR_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
INPUT_FILL = PatternFill("solid", fgColor="FFFFCC")
RESULT_FILL = PatternFill("solid", fgColor="C6EFCE")
GREY_FILL = PatternFill("solid", fgColor="E8E8E8")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YEL_FILL = PatternFill("solid", fgColor="FFEB9C")
GRN_FILL = PatternFill("solid", fgColor="C6EFCE")
THIN = Border(left=Side('thin'), right=Side('thin'),
              top=Side('thin'), bottom=Side('thin'))


# ============================================================
# SCORING KEYS FROM APPENDIX B (pages 181-184)
# Format: 'scale': {'proto': [items weight=2], 'nonproto': [items weight=1], 'false': [items weight=1]}
# Raw Score = sum(proto*2) + sum(nonproto*1) + sum(false_answered_false*1)
# ============================================================
KEYS = {
    '1': {  # Schizoid (16 items)
        'proto': [10, 27, 46, 92, 105, 148, 165],
        'nonproto': [4, 38, 48, 101, 142, 156, 167],
        'false': [32, 57]
    },
    '2A': {  # Avoidant (16 items)
        'proto': [18, 40, 69, 84, 99, 127, 141, 174],
        'nonproto': [47, 48, 146, 148, 151, 158],
        'false': [57, 80]
    },
    '2B': {  # Depressive (15 items)
        'proto': [20, 25, 47, 112, 123, 133, 145, 151],
        'nonproto': [24, 43, 83, 86, 142, 148, 154],
        'false': []
    },
    '3': {  # Dependent (16 items)
        'proto': [16, 35, 45, 73, 94, 108, 135, 169],
        'nonproto': [47, 56, 84, 120, 133, 141, 151],
        'false': [82]
    },
    '4': {  # Histrionic (17 items)
        'proto': [12, 21, 32, 51, 57, 80, 88],
        'nonproto': [10, 24, 27, 48, 69, 92, 99, 123, 127, 174],
        'false': []
    },
    '5': {  # Narcissistic (24 items)
        'proto': [5, 26, 31, 67, 85, 93, 144, 159],
        'nonproto': [21, 38, 57, 80, 88, 116],
        'false': [35, 40, 47, 69, 84, 86, 94, 99, 141, 169]
    },
    '6A': {  # Antisocial (17 items)
        'proto': [17, 38, 53, 101, 113, 139, 166],
        'nonproto': [7, 13, 14, 21, 41, 52, 93, 122, 136],
        'false': [172]
    },
    '6B': {  # Sadistic/Aggressive (20 items)
        'proto': [9, 14, 28, 64, 87, 95, 116],
        'nonproto': [7, 13, 17, 33, 36, 39, 41, 49, 53, 79, 93, 96, 166],
        'false': []
    },
    '7': {  # Compulsive (17 items)
        'proto': [2, 29, 59, 82, 97, 114, 137, 172],
        'nonproto': [],
        'false': [7, 14, 22, 41, 53, 72, 101, 139, 166]
    },
    '8A': {  # Negativistic/Passive-Aggressive (16 items)
        'proto': [7, 15, 22, 36, 50, 60, 79, 115, 126],
        'nonproto': [6, 42, 83, 98, 122, 133, 166],
        'false': []
    },
    '8B': {  # Masochistic/Self-Defeating (15 items)
        'proto': [19, 43, 70, 90, 104, 122, 161],
        'nonproto': [18, 24, 25, 35, 40, 98, 148, 169],
        'false': []
    },
    'S': {  # Schizotypal (16 items)
        'proto': [8, 48, 71, 76, 117, 138, 156, 158, 162],
        'nonproto': [69, 99, 102, 134, 141, 148, 151],
        'false': []
    },
    'C': {  # Borderline (16 items)
        'proto': [30, 41, 72, 83, 98, 120, 134, 142, 154],
        'nonproto': [7, 22, 122, 135, 161, 166, 171],
        'false': []
    },
    'P': {  # Paranoid (17 items)
        'proto': [6, 33, 42, 49, 89, 103, 146, 167, 175],
        'nonproto': [8, 48, 60, 63, 115, 138, 158, 159],
        'false': []
    },
    'A': {  # Anxiety (14 items)
        'proto': [58, 75, 124, 147, 164, 170],
        'nonproto': [40, 61, 76, 108, 109, 135, 145, 149],
        'false': []
    },
    'H': {  # Somatoform (12 items)
        'proto': [4, 11, 37, 55, 74],
        'nonproto': [1, 75, 107, 111, 130, 145, 148],
        'false': []
    },
    'N': {  # Bipolar: Manic (13 items)
        'proto': [3, 54, 96, 106, 125],
        'nonproto': [22, 41, 51, 83, 117, 134, 166, 170],
        'false': []
    },
    'D': {  # Dysthymia (14 items)
        'proto': [24, 56, 62, 86, 111, 130],
        'nonproto': [15, 25, 55, 83, 104, 141, 142, 148],
        'false': []
    },
    'B': {  # Alcohol Dependence (15 items)
        'proto': [52, 77, 100, 131, 152],
        'nonproto': [14, 41, 64, 93, 101, 113, 122, 139, 166],
        'false': [23]
    },
    'T': {  # Drug Dependence (14 items)
        'proto': [13, 39, 66, 91, 118, 136],
        'nonproto': [7, 21, 38, 41, 53, 101, 113, 139],
        'false': []
    },
    'R': {  # PTSD (16 items)
        'proto': [109, 129, 149, 160, 173],
        'nonproto': [62, 76, 83, 123, 133, 142, 147, 148, 151, 154, 164],
        'false': []
    },
    'SS': {  # Thought Disorder (17 items)
        'proto': [34, 61, 68, 78, 102, 168],
        'nonproto': [22, 56, 72, 76, 83, 117, 134, 142, 148, 151, 162],
        'false': []
    },
    'CC': {  # Major Depression (17 items)
        'proto': [1, 44, 107, 128, 150, 171],
        'nonproto': [4, 34, 55, 74, 104, 111, 130, 142, 148, 149, 154],
        'false': []
    },
    'PP': {  # Delusional Disorder (13 items)
        'proto': [63, 119, 140, 153],
        'nonproto': [5, 38, 49, 67, 89, 103, 138, 159, 175],
        'false': []
    },
    'Y': {  # Desirability (21 items)
        'true': [32, 51, 57, 59, 80, 82, 88, 97, 137, 172],
        'false': [20, 35, 40, 69, 104, 112, 123, 141, 142, 148, 151]
    },
    'Z': {  # Debasement (33 items)
        'true': [1, 4, 8, 15, 22, 24, 30, 34, 36, 37, 44, 55, 56, 58, 62, 63,
                 70, 74, 75, 76, 83, 84, 86, 99, 111, 123, 128, 133, 134, 142,
                 145, 150, 171],
        'false': []
    },
    'V': {  # Invalidity (3 items)
        'true': [65, 110, 157],
        'false': []
    },
}


# ============================================================
# SCALE X (Disclosure) - special: ALL 175 items, raw = count of TRUE
# From Appendix B: Scale X uses all items endorsed True
# The raw score for X = total number of items answered TRUE
# ============================================================
# Scale X is simply the COUNT of all items answered TRUE (weight=1 each)
# It's listed separately because its BR table is unique (Appendix C.2)

# ============================================================
# BR TRANSFORMATION TABLE - Appendix C.1 (page 186)
# Format: {scale: {raw_score: br_score}}
# From the image: Raw Score 0-26 for scales 1,2A,2B,3,4,5,6A,6B,7,8A,8B,S,C,P
# ============================================================
BR_C1 = {
    '1':  [0,12,24,36,48,60,62,64,66,68,70,72,75,78,81,85,89,93,97,101,105,109,112,115,115,115,115],
    '2A': [0,12,24,36,48,60,63,67,71,75,77,79,81,83,85,88,91,94,97,100,103,106,109,112,115,115,115],
    '2B': [0,10,20,30,40,50,60,65,70,75,77,79,81,83,85,89,92,96,99,103,106,109,112,115,115,115,115],
    '3':  [0,10,20,30,40,50,60,65,70,75,78,81,83,85,87,89,91,94,97,100,103,106,109,112,115,115,115],
    '4':  [0,4,8,12,16,20,24,28,32,36,40,44,48,51,54,57,60,63,66,69,72,75,79,83,88,88,88],
    '5':  [0,5,10,15,20,25,30,35,40,44,48,52,56,60,63,67,71,75,85,89,93,97,101,105,110,115,115],
    '6A': [0,12,24,36,48,60,62,64,66,69,71,73,75,75,79,82,85,89,92,96,99,103,106,109,112,115,115],
    '6B': [0,12,24,36,48,60,62,64,66,68,69,70,71,72,73,74,75,78,80,83,85,90,95,100,105,110,115],
    '7':  [0,4,8,12,16,20,24,28,32,36,39,42,45,48,51,54,57,60,63,66,69,72,75,79,83,87,87],
    '8A': [0,10,20,30,40,50,60,62,64,66,68,70,72,72,75,75,77,77,79,81,83,85,89,93,97,101,105,110,115],
    '8B': [0,20,40,60,63,66,69,75,78,80,82,85,88,91,94,97,100,103,106,109,112,115,115,115,115,115,115],
    'S':  [0,20,40,60,62,64,64,67,68,69,70,71,72,73,74,75,78,81,85,90,95,99,103,107,111,115,115],
    'C':  [0,12,24,36,48,60,63,66,69,72,75,77,77,79,81,83,85,88,91,94,97,100,103,106,109,112,115],
    'P':  [0,15,30,45,60,61,63,64,66,67,69,70,72,73,75,77,79,81,83,85,90,95,100,105,110,115,115],
}

# Appendix C.1 continued (page 187) - Clinical syndromes + Y, Z
BR_C1_2 = {
    'A':  [0,20,40,60,75,77,79,81,83,85,87,89,91,94,97,100,103,106,109,112,115,115,115,115],
    'H':  [0,15,30,45,60,62,64,66,68,70,72,73,74,75,80,85,100,115,115,115,115,115,115,115],
    'N':  [0,12,24,36,48,60,63,66,69,72,75,80,85,90,95,100,105,110,115,115,115,115,115,115],
    'D':  [0,12,24,36,48,60,62,64,66,69,72,75,78,80,82,85,91,97,103,109,115,115,115,115],
    'B':  [0,20,40,60,63,67,71,75,77,80,83,85,88,91,95,99,103,107,111,115,115,115,115,115],
    'T':  [0,20,40,60,63,67,71,75,76,77,78,79,81,83,85,90,95,100,105,110,115,115,115,115],
    'R':  [0,15,30,45,60,62,63,65,66,68,69,71,73,75,77,79,81,83,85,95,105,115,115,115],
    'SS': [0,15,30,45,60,62,64,66,67,68,69,70,71,72,73,74,75,79,82,85,93,100,108,115,115,115],
    'CC': [0,15,30,45,60,65,70,75,78,81,87,89,91,93,95,97,99,101,103,106,109,112,115,115,115,115],
    'PP': [0,30,60,62,65,68,70,72,75,80,85,90,95,100,105,110,115,115,115,115,115,115,115,115],
    'Y':  [0,5,10,15,20,25,30,35,39,43,47,51,55,59,63,67,71,75,80,85,93,100],
    'Z':  [0,18,35,38,40,42,45,47,49,52,54,56,59,61,63,66,68,70,73,75,76,78,79,81,82,84,85,86,88,90,92,94,96,98,100],
}


# ============================================================
# BR TABLE for Scale X (Appendix C.2, page 188)
# Raw score ranges map to BR scores
# ============================================================
BR_X = {}
# Raw 34-38 -> BR 0
for r in range(34, 39): BR_X[r] = 0
BR_X[39] = 2
BR_X[40] = 3
BR_X[41] = 5
BR_X[42] = 6
BR_X[43] = 8
BR_X[44] = 9
BR_X[45] = 11
BR_X[46] = 12
BR_X[47] = 14
BR_X[48] = 15
BR_X[49] = 17
BR_X[50] = 18
BR_X[51] = 20
BR_X[52] = 21
BR_X[53] = 23
BR_X[54] = 24
BR_X[55] = 26
BR_X[56] = 27
BR_X[57] = 29
BR_X[58] = 30
BR_X[59] = 32
BR_X[60] = 33
BR_X[61] = 35
BR_X[62] = 36
for r in range(63, 65): BR_X[r] = 37
BR_X[65] = 38
for r in range(66, 68): BR_X[r] = 39
BR_X[68] = 40
for r in range(69, 71): BR_X[r] = 41
BR_X[71] = 42
for r in range(72, 74): BR_X[r] = 43
BR_X[74] = 44
for r in range(75, 77): BR_X[r] = 45
BR_X[77] = 46
for r in range(78, 80): BR_X[r] = 47
BR_X[80] = 48
for r in range(81, 83): BR_X[r] = 49
BR_X[83] = 50
for r in range(84, 86): BR_X[r] = 51
BR_X[86] = 52
for r in range(87, 89): BR_X[r] = 53
BR_X[89] = 54
for r in range(90, 92): BR_X[r] = 55
BR_X[92] = 56
for r in range(93, 95): BR_X[r] = 57
BR_X[95] = 58
for r in range(96, 98): BR_X[r] = 59
BR_X[98] = 60
for r in range(99, 101): BR_X[r] = 61
BR_X[101] = 62
for r in range(102, 104): BR_X[r] = 63
BR_X[104] = 64
for r in range(105, 107): BR_X[r] = 65
BR_X[107] = 66
for r in range(108, 110): BR_X[r] = 67
BR_X[110] = 68
for r in range(111, 113): BR_X[r] = 69
BR_X[113] = 70
for r in range(114, 116): BR_X[r] = 71
BR_X[116] = 72
for r in range(117, 119): BR_X[r] = 73
for r in range(119, 121): BR_X[r] = 74
for r in range(121, 124): BR_X[r] = 75
for r in range(124, 127): BR_X[r] = 76
for r in range(127, 130): BR_X[r] = 77
for r in range(130, 133): BR_X[r] = 78
for r in range(133, 136): BR_X[r] = 79
for r in range(136, 139): BR_X[r] = 80
for r in range(139, 142): BR_X[r] = 81
for r in range(142, 145): BR_X[r] = 82
for r in range(145, 147): BR_X[r] = 83
for r in range(147, 149): BR_X[r] = 84
for r in range(149, 151): BR_X[r] = 85
for r in range(151, 153): BR_X[r] = 86
BR_X[153] = 87
for r in range(154, 156): BR_X[r] = 88
for r in range(156, 158): BR_X[r] = 89
for r in range(158, 160): BR_X[r] = 90
BR_X[160] = 91
for r in range(161, 163): BR_X[r] = 92
for r in range(163, 165): BR_X[r] = 93
for r in range(165, 167): BR_X[r] = 94
for r in range(167, 169): BR_X[r] = 95
for r in range(169, 171): BR_X[r] = 96
for r in range(171, 173): BR_X[r] = 97
BR_X[173] = 98
BR_X[174] = 99
for r in range(175, 179): BR_X[r] = 100


# ============================================================
# DISCLOSURE ADJUSTMENT TABLE (from existing MCMI.SOFTWARE.xlsx)
# Scale X raw score -> (adjustment for scales 1-8B, adjustment for S-PP)
# ============================================================
DISC_ADJ = {}
for i in range(0, 37): DISC_ADJ[i] = (20, 10)
DISC_ADJ[37] = (19, 10); DISC_ADJ[38] = (18, 10)
DISC_ADJ[39] = (17, 9); DISC_ADJ[40] = (17, 9)
DISC_ADJ[41] = (16, 9); DISC_ADJ[42] = (15, 8)
DISC_ADJ[43] = (14, 8); DISC_ADJ[44] = (13, 7)
DISC_ADJ[45] = (13, 7); DISC_ADJ[46] = (12, 7)
DISC_ADJ[47] = (11, 6); DISC_ADJ[48] = (10, 6)
DISC_ADJ[49] = (9, 5); DISC_ADJ[50] = (9, 5)
DISC_ADJ[51] = (8, 5); DISC_ADJ[52] = (7, 4)
DISC_ADJ[53] = (6, 4); DISC_ADJ[54] = (5, 3)
DISC_ADJ[55] = (5, 3); DISC_ADJ[56] = (4, 3)
DISC_ADJ[57] = (3, 2); DISC_ADJ[58] = (2, 2)
DISC_ADJ[59] = (1, 1); DISC_ADJ[60] = (1, 1)
for i in range(61, 124): DISC_ADJ[i] = (0, 0)
DISC_ADJ[124]=(-1,-1); DISC_ADJ[125]=(-1,-1); DISC_ADJ[126]=(-1,-1)
DISC_ADJ[127]=(-2,-2); DISC_ADJ[128]=(-2,-2)
DISC_ADJ[129]=(-3,-2); DISC_ADJ[130]=(-3,-2); DISC_ADJ[131]=(-3,-2)
DISC_ADJ[132]=(-4,-3); DISC_ADJ[133]=(-4,-3)
DISC_ADJ[134]=(-5,-3); DISC_ADJ[135]=(-5,-3); DISC_ADJ[136]=(-5,-3)
DISC_ADJ[137]=(-6,-4); DISC_ADJ[138]=(-6,-4)
DISC_ADJ[139]=(-7,-4); DISC_ADJ[140]=(-7,-4); DISC_ADJ[141]=(-7,-4)
DISC_ADJ[142]=(-8,-5); DISC_ADJ[143]=(-8,-5)
DISC_ADJ[144]=(-9,-5); DISC_ADJ[145]=(-9,-5); DISC_ADJ[146]=(-9,-5)
DISC_ADJ[147]=(-10,-6); DISC_ADJ[148]=(-10,-6)
DISC_ADJ[149]=(-11,-6); DISC_ADJ[150]=(-11,-6); DISC_ADJ[151]=(-11,-6)
DISC_ADJ[152]=(-12,-7); DISC_ADJ[153]=(-12,-7)
DISC_ADJ[154]=(-13,-7); DISC_ADJ[155]=(-13,-7); DISC_ADJ[156]=(-13,-7)
DISC_ADJ[157]=(-14,-8); DISC_ADJ[158]=(-14,-8)
DISC_ADJ[159]=(-15,-8); DISC_ADJ[160]=(-15,-8); DISC_ADJ[161]=(-15,-8)
DISC_ADJ[162]=(-16,-9); DISC_ADJ[163]=(-16,-9)
DISC_ADJ[164]=(-17,-9); DISC_ADJ[165]=(-17,-9); DISC_ADJ[166]=(-17,-9)
DISC_ADJ[167]=(-18,-10); DISC_ADJ[168]=(-18,-10)
DISC_ADJ[169]=(-19,-10); DISC_ADJ[170]=(-19,-10); DISC_ADJ[171]=(-19,-10)
for i in range(172, 200): DISC_ADJ[i] = (-20, -11)

# A/D Adjustment tables
AD_2AS = {}  # For scales 2A, S
for i in range(0,8): AD_2AS[i]=-1
for i in range(8,16): AD_2AS[i]=-2
for i in range(16,24): AD_2AS[i]=-3
for i in range(24,32): AD_2AS[i]=-4
for i in range(32,40): AD_2AS[i]=-5
for i in range(40,48): AD_2AS[i]=-6
for i in range(48,56): AD_2AS[i]=-7
for i in range(56,64): AD_2AS[i]=-8
for i in range(64,72): AD_2AS[i]=-9
for i in range(72,81): AD_2AS[i]=-10

AD_2B8BC = {}  # For scales 2B, 8B, C
for i in range(0,10): AD_2B8BC[i]=-1
for i in range(10,15): AD_2B8BC[i]=-2
for i in range(15,20): AD_2B8BC[i]=-3
for i in range(20,25): AD_2B8BC[i]=-4
for i in range(25,30): AD_2B8BC[i]=-5
for i in range(30,35): AD_2B8BC[i]=-6
for i in range(35,40): AD_2B8BC[i]=-7
for i in range(40,45): AD_2B8BC[i]=-8
for i in range(45,50): AD_2B8BC[i]=-9
for i in range(50,55): AD_2B8BC[i]=-10
for i in range(55,60): AD_2B8BC[i]=-11
for i in range(60,65): AD_2B8BC[i]=-12
for i in range(65,70): AD_2B8BC[i]=-13
for i in range(70,75): AD_2B8BC[i]=-14
for i in range(75,81): AD_2B8BC[i]=-15

# ============================================================
# SCALE ORDER
# ============================================================
SCALES = [
    ('X', 'Disclosure', 'Modifying'),
    ('Y', 'Desirability', 'Modifying'),
    ('Z', 'Debasement', 'Modifying'),
    ('1', 'Schizoid', 'Personality'),
    ('2A', 'Avoidant', 'Personality'),
    ('2B', 'Depressive', 'Personality'),
    ('3', 'Dependent', 'Personality'),
    ('4', 'Histrionic', 'Personality'),
    ('5', 'Narcissistic', 'Personality'),
    ('6A', 'Antisocial', 'Personality'),
    ('6B', 'Sadistic', 'Personality'),
    ('7', 'Compulsive', 'Personality'),
    ('8A', 'Negativistic', 'Personality'),
    ('8B', 'Masochistic', 'Personality'),
    ('S', 'Schizotypal', 'Severe Pers'),
    ('C', 'Borderline', 'Severe Pers'),
    ('P', 'Paranoid', 'Severe Pers'),
    ('A', 'Anxiety', 'Clinical Syn'),
    ('H', 'Somatoform', 'Clinical Syn'),
    ('N', 'Bipolar Manic', 'Clinical Syn'),
    ('D', 'Dysthymia', 'Clinical Syn'),
    ('B', 'Alcohol Dep', 'Clinical Syn'),
    ('T', 'Drug Dep', 'Clinical Syn'),
    ('R', 'PTSD', 'Clinical Syn'),
    ('SS', 'Thought Dis', 'Severe Clin'),
    ('CC', 'Major Dep', 'Severe Clin'),
    ('PP', 'Delusional', 'Severe Clin'),
]


# ============================================================
# SECTION 1: HEADER + PATIENT INFO (Rows 1-5)
# ============================================================
ws['A1'] = "MCMI-III COMPLETE SCORING TOOL"
ws['A1'].font = TITLE_FONT
ws['A2'] = "Enter 1=TRUE or 0=FALSE for each item. Scoring is automatic."
ws['A2'].font = Font(italic=True, size=9)

ws['A4'] = "Patient:"
ws['B4'].fill = INPUT_FILL
ws['D4'] = "Date:"
ws['E4'].fill = INPUT_FILL
ws['G4'] = "Age:"
ws['H4'].fill = INPUT_FILL
ws['A5'] = "Inpatient (Y/N):"
ws['B5'].fill = INPUT_FILL
ws['D5'] = "Axis I weeks:"
ws['E5'].fill = INPUT_FILL
ws['G5'] = "Gender:"
ws['H5'].fill = INPUT_FILL

# ============================================================
# SECTION 2: ITEM RESPONSES (Rows 7-47)
# 175 items in 5 groups of 35
# Cols: A/B, D/E, G/H, J/K, M/N
# ============================================================
ws['A7'] = "ITEM RESPONSES"
ws['A7'].font = SECTION_FONT
ws['A7'].fill = SECTION_FILL

ITEM_ROW_START = 8  # Header row
item_cols = [('A','B'), ('D','E'), ('G','H'), ('J','K'), ('M','N')]

# Headers
for nc, rc in item_cols:
    ws[f'{nc}{ITEM_ROW_START}'] = "#"
    ws[f'{nc}{ITEM_ROW_START}'].font = HDR_FONT
    ws[f'{nc}{ITEM_ROW_START}'].fill = HDR_FILL
    ws[f'{nc}{ITEM_ROW_START}'].border = THIN
    ws[f'{rc}{ITEM_ROW_START}'] = "T/F"
    ws[f'{rc}{ITEM_ROW_START}'].font = HDR_FONT
    ws[f'{rc}{ITEM_ROW_START}'].fill = HDR_FILL
    ws[f'{rc}{ITEM_ROW_START}'].border = THIN

# Fill items 1-175
for col_idx, (nc, rc) in enumerate(item_cols):
    start = col_idx * 35 + 1
    end = min(start + 34, 175)
    for item in range(start, end + 1):
        r = ITEM_ROW_START + 1 + (item - start)
        ws[f'{nc}{r}'] = item
        ws[f'{nc}{r}'].alignment = Alignment(horizontal='center')
        ws[f'{nc}{r}'].border = THIN
        ws[f'{rc}{r}'].fill = INPUT_FILL
        ws[f'{rc}{r}'].border = THIN
        ws[f'{rc}{r}'].alignment = Alignment(horizontal='center')

# Item cell mapping function
def cell_ref(item_num):
    """Get absolute cell reference for an item's response."""
    col_idx = (item_num - 1) // 35
    row_offset = (item_num - 1) % 35
    resp_col = item_cols[col_idx][1]
    row = ITEM_ROW_START + 1 + row_offset
    return f"${resp_col}${row}"

# Verify: item 1 -> $B$9, item 35 -> $B$43, item 36 -> $E$9, etc.


# ============================================================
# SECTION 3: SCORING TABLE (Row 46 onwards)
# ============================================================
SCORE_SEC_ROW = ITEM_ROW_START + 1 + 35 + 1  # row 45
ws[f'A{SCORE_SEC_ROW}'] = "SCORING RESULTS"
ws[f'A{SCORE_SEC_ROW}'].font = SECTION_FONT
ws[f'A{SCORE_SEC_ROW}'].fill = SECTION_FILL

# Column headers
HDR_ROW = SCORE_SEC_ROW + 1
col_headers = ['Scale', 'Name', 'Cat', 'Raw\nScore',
               'BR Score\n(auto)', 'Disc\nAdj', 'A/D\nAdj',
               'Inp\nAdj', 'Den/\nComp', 'FINAL\nBR', 'Signif.']
for ci, h in enumerate(col_headers):
    c = ws.cell(row=HDR_ROW, column=ci+1)
    c.value = h
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.border = THIN
    c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
ws.row_dimensions[HDR_ROW].height = 30

# ============================================================
# BUILD RAW SCORE FORMULAS (WEIGHTED)
# Raw = sum(proto_items * 2) + sum(nonproto_items * 1) + sum((1-false_items) * 1)
# For Y and Z: just count of TRUE/FALSE items (weight=1 each)
# For X: count ALL 175 items answered TRUE
# ============================================================
DATA_START = HDR_ROW + 1
scale_rows = {}

for idx, (code, name, cat) in enumerate(SCALES):
    row = DATA_START + idx
    scale_rows[code] = row
    
    ws.cell(row=row, column=1, value=code).border = THIN
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=2, value=name).border = THIN
    ws.cell(row=row, column=3, value=cat).border = THIN
    
    # RAW SCORE FORMULA (Column D)
    if code == 'X':
        # Scale X = count of all 175 TRUE responses
        parts = [cell_ref(i) for i in range(1, 176)]
        raw_formula = "=" + "+".join(parts)
    elif code in ('Y', 'Z', 'V'):
        # Y and Z use weight=1 for all (true items count, false items score for NOT answering)
        key = KEYS[code]
        parts = []
        for item in key.get('true', []):
            parts.append(cell_ref(item))
        for item in key.get('false', []):
            parts.append(f"(1-{cell_ref(item)})")
        raw_formula = "=" + "+".join(parts) if parts else "=0"
    else:
        # Weighted: proto*2, nonproto*1, false=(1-response)*1
        key = KEYS[code]
        parts = []
        for item in key['proto']:
            parts.append(f"2*{cell_ref(item)}")
        for item in key['nonproto']:
            parts.append(cell_ref(item))
        for item in key['false']:
            parts.append(f"(1-{cell_ref(item)})")
        raw_formula = "=" + "+".join(parts) if parts else "=0"
    
    ws.cell(row=row, column=4, value=raw_formula)
    ws.cell(row=row, column=4).fill = RESULT_FILL
    ws.cell(row=row, column=4).border = THIN
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=4).font = Font(bold=True)


# ============================================================
# EMBED BR LOOKUP TABLES IN HIDDEN COLUMNS (P onwards)
# Col P: raw scores 0-33 (row number for lookup)
# Col Q onwards: BR for each scale
# We'll put the C.1 tables for personality/clinical scales
# And C.2 for Scale X in a separate range
# ============================================================

# --- BR Lookup for scales (Cols P-AQ) ---
# Put labels in row 1
LOOKUP_START_COL = 16  # Column P
lookup_scales_order = ['1','2A','2B','3','4','5','6A','6B','7','8A','8B',
                       'S','C','P','A','H','N','D','B','T','R','SS','CC','PP','Y','Z']

# Header row
ws.cell(row=1, column=LOOKUP_START_COL, value="Raw")
for si, sc in enumerate(lookup_scales_order):
    ws.cell(row=1, column=LOOKUP_START_COL + 1 + si, value=sc)

# Fill data (rows 2 onwards = raw score 0, 1, 2, ...)
max_raw = 35  # enough for all scales
for raw in range(max_raw + 1):
    ws.cell(row=2 + raw, column=LOOKUP_START_COL, value=raw)
    for si, sc in enumerate(lookup_scales_order):
        br_list = BR_C1.get(sc, BR_C1_2.get(sc, []))
        if raw < len(br_list):
            br_val = br_list[raw]
        else:
            br_val = br_list[-1] if br_list else 0
        ws.cell(row=2 + raw, column=LOOKUP_START_COL + 1 + si, value=br_val)

# Scale X BR lookup (separate cols)
X_LOOKUP_COL = LOOKUP_START_COL + len(lookup_scales_order) + 2  # Skip a col
ws.cell(row=1, column=X_LOOKUP_COL, value="X_Raw")
ws.cell(row=1, column=X_LOOKUP_COL + 1, value="X_BR")
for raw in range(34, 179):
    r = 2 + (raw - 34)
    ws.cell(row=r, column=X_LOOKUP_COL, value=raw)
    ws.cell(row=r, column=X_LOOKUP_COL + 1, value=BR_X.get(raw, 100))

# Disclosure adjustment lookup
DISC_LOOKUP_COL = X_LOOKUP_COL + 3
ws.cell(row=1, column=DISC_LOOKUP_COL, value="X_Raw_D")
ws.cell(row=1, column=DISC_LOOKUP_COL + 1, value="Adj_18B")
ws.cell(row=1, column=DISC_LOOKUP_COL + 2, value="Adj_SPP")
for raw in range(200):
    r = 2 + raw
    adj18, adjsp = DISC_ADJ.get(raw, (0, 0))
    ws.cell(row=r, column=DISC_LOOKUP_COL, value=raw)
    ws.cell(row=r, column=DISC_LOOKUP_COL + 1, value=adj18)
    ws.cell(row=r, column=DISC_LOOKUP_COL + 2, value=adjsp)

# A/D lookup tables
AD2AS_COL = DISC_LOOKUP_COL + 4
ws.cell(row=1, column=AD2AS_COL, value="AD_Val")
ws.cell(row=1, column=AD2AS_COL + 1, value="Adj_2AS")
for val in range(81):
    ws.cell(row=2 + val, column=AD2AS_COL, value=val)
    ws.cell(row=2 + val, column=AD2AS_COL + 1, value=AD_2AS[val])

AD2B_COL = AD2AS_COL + 3
ws.cell(row=1, column=AD2B_COL, value="AD_Val2")
ws.cell(row=1, column=AD2B_COL + 1, value="Adj_2B8BC")
for val in range(81):
    ws.cell(row=2 + val, column=AD2B_COL, value=val)
    ws.cell(row=2 + val, column=AD2B_COL + 1, value=AD_2B8BC[val])

# Get column letters for formulas
from openpyxl.utils import get_column_letter
RAW_COL_LTR = get_column_letter(LOOKUP_START_COL)
X_RAW_LTR = get_column_letter(X_LOOKUP_COL)
X_BR_LTR = get_column_letter(X_LOOKUP_COL + 1)
DISC_RAW_LTR = get_column_letter(DISC_LOOKUP_COL)
DISC_18B_LTR = get_column_letter(DISC_LOOKUP_COL + 1)
DISC_SPP_LTR = get_column_letter(DISC_LOOKUP_COL + 2)
AD2AS_VAL_LTR = get_column_letter(AD2AS_COL)
AD2AS_ADJ_LTR = get_column_letter(AD2AS_COL + 1)
AD2B_VAL_LTR = get_column_letter(AD2B_COL)
AD2B_ADJ_LTR = get_column_letter(AD2B_COL + 1)

# Build column letter for each scale's BR lookup
scale_br_col_ltr = {}
for si, sc in enumerate(lookup_scales_order):
    scale_br_col_ltr[sc] = get_column_letter(LOOKUP_START_COL + 1 + si)


# ============================================================
# NOW ADD BR SCORE AND ADJUSTMENT FORMULAS
# ============================================================
PERS_18B = ['1','2A','2B','3','4','5','6A','6B','7','8A','8B']
SPP_SCALES = ['S','C','P','A','H','N','D','B','T','R','SS','CC','PP']
AD_2AS_CODES = ['2A', 'S']
AD_2B8BC_CODES = ['2B', '8B', 'C']

# X raw cell for disclosure adjustment
x_raw = f"D{scale_rows['X']}"

# A and D rows for A/D adjustment
a_row = scale_rows['A']
d_row = scale_rows['D']

for idx, (code, name, cat) in enumerate(SCALES):
    row = DATA_START + idx
    
    # Column E: BR Score (auto from lookup)
    if code == 'X':
        # Use Scale X BR table (Appendix C.2)
        br_formula = (
            f'=IF(D{row}<34,0,IF(D{row}>178,100,'
            f'VLOOKUP(D{row},${X_RAW_LTR}:${X_BR_LTR},2,TRUE)))'
        )
    elif code in scale_br_col_ltr:
        # Use C.1 table via VLOOKUP
        br_col = scale_br_col_ltr[code]
        br_formula = (
            f'=VLOOKUP(MIN(D{row},{max_raw}),'
            f'${RAW_COL_LTR}:${br_col},COLUMN(${br_col}$1)-COLUMN(${RAW_COL_LTR}$1)+1,TRUE)'
        )
    else:
        br_formula = ""
    
    ws.cell(row=row, column=5, value=br_formula)
    ws.cell(row=row, column=5).fill = RESULT_FILL
    ws.cell(row=row, column=5).border = THIN
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    
    # Column F: Disclosure Adjustment
    if code in ['X', 'Y', 'Z']:
        ws.cell(row=row, column=6, value=0)
        ws.cell(row=row, column=6).fill = GREY_FILL
    elif code in PERS_18B:
        disc_f = (
            f'=IF({x_raw}="",0,'
            f'VLOOKUP(MIN(MAX({x_raw},0),199),'
            f'${DISC_RAW_LTR}:${DISC_18B_LTR},2,TRUE))'
        )
        ws.cell(row=row, column=6, value=disc_f)
    else:  # S-PP scales
        disc_f = (
            f'=IF({x_raw}="",0,'
            f'VLOOKUP(MIN(MAX({x_raw},0),199),'
            f'${DISC_RAW_LTR}:${DISC_SPP_LTR},3,TRUE))'
        )
        ws.cell(row=row, column=6, value=disc_f)
    ws.cell(row=row, column=6).border = THIN
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
    
    # Column G: A/D Adjustment
    # A adjusted = E{a_row} + F{a_row}; D adjusted = E{d_row} + F{d_row}
    a_adj_expr = f"(E{a_row}+F{a_row})"
    d_adj_expr = f"(E{d_row}+F{d_row})"
    
    if code in AD_2AS_CODES:
        ad_f = (
            f'=IF(AND({a_adj_expr}<75,{d_adj_expr}<75),0,'
            f'VLOOKUP(MIN(MAX(IF({a_adj_expr}>=75,{a_adj_expr}-75,0)'
            f'+IF({d_adj_expr}>=75,{d_adj_expr}-75,0),0),80),'
            f'${AD2AS_VAL_LTR}:${AD2AS_ADJ_LTR},2,TRUE))'
        )
        ws.cell(row=row, column=7, value=ad_f)
    elif code in AD_2B8BC_CODES:
        ad_f = (
            f'=IF(AND({a_adj_expr}<75,{d_adj_expr}<75),0,'
            f'VLOOKUP(MIN(MAX(IF({a_adj_expr}>=75,{a_adj_expr}-75,0)'
            f'+IF({d_adj_expr}>=75,{d_adj_expr}-75,0),0),80),'
            f'${AD2B_VAL_LTR}:${AD2B_ADJ_LTR},2,TRUE))'
        )
        ws.cell(row=row, column=7, value=ad_f)
    else:
        ws.cell(row=row, column=7, value=0)
        ws.cell(row=row, column=7).fill = GREY_FILL
    ws.cell(row=row, column=7).border = THIN
    ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
    
    # Column H: Inpatient Adjustment
    if code == 'SS':
        ws.cell(row=row, column=8, value='=IF($B$5="Y",IF($E$5<1,6,IF($E$5<=4,4,0)),0)')
    elif code == 'CC':
        ws.cell(row=row, column=8, value='=IF($B$5="Y",IF($E$5<1,10,IF($E$5<=4,8,0)),0)')
    elif code == 'PP':
        ws.cell(row=row, column=8, value='=IF($B$5="Y",IF($E$5<1,4,IF($E$5<=4,2,0)),0)')
    else:
        ws.cell(row=row, column=8, value=0)
        ws.cell(row=row, column=8).fill = GREY_FILL
    ws.cell(row=row, column=8).border = THIN
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')
    
    # Column I: Denial/Complaint (manual entry)
    ws.cell(row=row, column=9, value=0)
    if code in PERS_18B:
        ws.cell(row=row, column=9).fill = INPUT_FILL
    else:
        ws.cell(row=row, column=9).fill = GREY_FILL
    ws.cell(row=row, column=9).border = THIN
    ws.cell(row=row, column=9).alignment = Alignment(horizontal='center')
    
    # Column J: FINAL BR = E + F + G + H + I, clamped 0-115
    final_f = f'=MIN(MAX(E{row}+F{row}+G{row}+H{row}+I{row},0),115)'
    ws.cell(row=row, column=10, value=final_f)
    ws.cell(row=row, column=10).border = THIN
    ws.cell(row=row, column=10).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=10).font = Font(bold=True, size=11)
    
    # Column K: Clinical Significance
    sig_f = (
        f'=IF(J{row}>=85,"PATHOLOGICAL",'
        f'IF(J{row}>=75,"PRESENT",'
        f'IF(J{row}>=60,"Suggestive","Normal")))'
    )
    ws.cell(row=row, column=11, value=sig_f)
    ws.cell(row=row, column=11).border = THIN
    ws.cell(row=row, column=11).alignment = Alignment(horizontal='center')


# ============================================================
# CONDITIONAL FORMATTING
# ============================================================
last_data_row = DATA_START + len(SCALES) - 1
rng = f"J{DATA_START}:J{last_data_row}"
ws.conditional_formatting.add(rng, CellIsRule(
    operator='greaterThanOrEqual', formula=['85'],
    fill=RED_FILL, font=Font(bold=True, color="9C0006")))
ws.conditional_formatting.add(rng, CellIsRule(
    operator='between', formula=['75','84'],
    fill=YEL_FILL, font=Font(bold=True, color="9C6500")))
ws.conditional_formatting.add(rng, CellIsRule(
    operator='lessThan', formula=['60'],
    fill=GRN_FILL, font=Font(color="006100")))

# ============================================================
# VALIDITY SECTION (below scoring table)
# ============================================================
val_row = last_data_row + 2
ws.cell(row=val_row, column=1, value="VALIDITY").font = SECTION_FONT
ws.cell(row=val_row, column=1).fill = SECTION_FILL

ws.cell(row=val_row+1, column=1, value="V Score:")
ws.cell(row=val_row+1, column=2, value=f"={cell_ref(65)}+{cell_ref(110)}+{cell_ref(157)}")
ws.cell(row=val_row+1, column=2).fill = RESULT_FILL
ws.cell(row=val_row+1, column=3, value=f'=IF(B{val_row+1}>1,"INVALID","OK")')

ws.cell(row=val_row+2, column=1, value="X Raw:")
ws.cell(row=val_row+2, column=2, value=f"=D{scale_rows['X']}")
ws.cell(row=val_row+2, column=2).fill = RESULT_FILL
ws.cell(row=val_row+2, column=3, value=f'=IF(OR(B{val_row+2}<34,B{val_row+2}>178),"INVALID","OK")')

ws.cell(row=val_row+3, column=1, value="STATUS:")
ws.cell(row=val_row+3, column=1).font = Font(bold=True)
ws.cell(row=val_row+3, column=2,
    value=f'=IF(OR(B{val_row+1}>1,B{val_row+2}<34,B{val_row+2}>178),"INVALID PROTOCOL","VALID")')
ws.cell(row=val_row+3, column=2).font = Font(bold=True, size=12)

# ============================================================
# INTERPRETATION KEY
# ============================================================
int_row = val_row + 5
ws.cell(row=int_row, column=1, value="INTERPRETATION").font = SECTION_FONT
ws.cell(row=int_row, column=1).fill = SECTION_FILL
ws.cell(row=int_row+1, column=1, value="BR 85-115 = PATHOLOGICAL (Disorder present)")
ws.cell(row=int_row+2, column=1, value="BR 75-84 = PRESENT (Clinically significant trait)")
ws.cell(row=int_row+3, column=1, value="BR 60-74 = Suggestive (Subclinical)")
ws.cell(row=int_row+4, column=1, value="BR 0-59 = Normal")

# Instructions
inst_row = int_row + 6
ws.cell(row=inst_row, column=1, value="INSTRUCTIONS").font = SECTION_FONT
ws.cell(row=inst_row, column=1).fill = SECTION_FILL
instructions = [
    "1. Enter 1 (TRUE) or 0 (FALSE) for all 175 items in yellow cells above",
    "2. Raw scores auto-calculate with correct WEIGHTS (proto=2, nonproto=1, false=1)",
    "3. BR scores auto-calculate from Appendix C lookup tables",
    "4. Disclosure, A/D, and Inpatient adjustments are automatic",
    "5. For Denial/Complaint (Col I): enter 8 if highest 1-8B scale is 4, 5, or 7",
    "6. Final BR and Significance appear automatically",
    "7. Scale 5 (Narcissistic) does NOT multiply by 2/3 - it has 24 items with weights",
    "8. Validity: V>1 or X<34 or X>178 = INVALID protocol",
]
for i, txt in enumerate(instructions):
    ws.cell(row=inst_row + 1 + i, column=1, value=txt)

# ============================================================
# COLUMN WIDTHS
# ============================================================
widths = {'A':7,'B':8,'C':5,'D':8,'E':8,'F':8,'G':8,'H':8,
          'I':5,'J':8,'K':8,'L':3,'M':8,'N':8}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Scoring table widths (override)
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 11
ws.column_dimensions['D'].width = 7
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 6
ws.column_dimensions['G'].width = 6
ws.column_dimensions['H'].width = 5
ws.column_dimensions['I'].width = 5
ws.column_dimensions['J'].width = 7
ws.column_dimensions['K'].width = 13

# Hide lookup columns (P onwards)
for col_num in range(LOOKUP_START_COL, AD2B_COL + 2):
    ws.column_dimensions[get_column_letter(col_num)].hidden = True

# ============================================================
# SAVE
# ============================================================
output = "/projects/sandbox/Dango-kiro/MCMI-III_Scoring_Tool_V2.xlsx"
wb.save(output)
import os
print(f"SUCCESS! Saved: {output}")
print(f"Size: {os.path.getsize(output)} bytes")
print(f"\nScoring table rows: {DATA_START}-{last_data_row}")
print(f"Scales: {len(SCALES)}")
print(f"Lookup columns: {get_column_letter(LOOKUP_START_COL)}-{get_column_letter(AD2B_COL+1)} (hidden)")
print(f"\nScale row map:")
for code, r in scale_rows.items():
    print(f"  {code}: row {r}")
